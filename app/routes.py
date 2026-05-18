from __future__ import annotations
from functools import wraps
from datetime import date, timedelta, datetime
import pytz
import secrets, json, re, os
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from pathlib import Path

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file, send_from_directory, current_app
import pandas as pd
from io import BytesIO
from sqlalchemy import or_, func, text, bindparam
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage

from . import db
from .models import (
    User,
    UserPtAccess,
    Vehicle,
    Company,
    KirRecord,
    ServiceRecord,
    LoanTransaction,
    VehicleChangeHistory,
    UserHistory,
    LoanHistory,
    AuditLog,
)
from .utils import DEFAULT_REMINDER_DAYS, audit, current_user, is_due_soon, is_overdue

bp=Blueprint("main",__name__)

UPLOAD_HEADERS=["No","PT","PT pemilik aset","NAMA AKTIVA","Nama sesuai asset PT","Nama Aset baru","MERK","TYPE","JENIS","NO POLISI LAMA","NO POLISI BARU","TAHUN PEMAKAIAN","USER LAMA","USER BARU","Status","PT Pemakai 1","PT Pemakai 2","Kondisi terkini","Lokasi","Tambahan Keterangan"]
HEADER_ALIASES={
"no":"no","pt":"pt","pt pemilik aset":"pt_pemilik_aset","ptpemilikaset":"pt_pemilik_aset","nama aktiva":"nama_aktiva",
"nama sesuai asset pt":"nama_sesuai_asset_pt","nama aset baru":"nama_aset_baru","merk":"merk","type":"type","jenis":"jenis",
"no polisi lama":"no_polisi_lama","nopol lama":"no_polisi_lama","no polisi baru":"no_polisi_baru","nopol baru":"no_polisi_baru",
"tahun pemakaian":"tahun_pemakaian","tahun pakai":"tahun_pemakaian","user lama":"user_lama","user baru":"user_baru","status":"status","pt pemakai":"pt_pemakai","kondisi terkini":"kondisi_terkini","lokasi":"lokasi","tambahan keterangan":"tambahan_keterangan"
}

def _display_name(user)->str:
    if not user:
        return "-"
    for attr in ("full_name", "name", "username"):
        value = getattr(user, attr, None)
        if value and str(value).strip():
            return str(value).strip()
    return "-"

def _session_user_id():
    return session.get("user_id")

def _current_db_user():
    user_id = _session_user_id()
    if not user_id:
        return None
    try:
        return db.session.get(User, int(user_id))
    except Exception:
        return None

def _is_master_session()->bool:
    return bool(session.get("is_master"))

def _is_admin_pt_session()->bool:
    """Return True jika user login sebagai admin PT (bukan master)."""
    return bool(session.get("user_id") and not _is_master_session())

def _user_allowed_pts(user=None)->list[str]:
    user = user or _current_db_user()
    if not user or user.is_master():
        return []
    values = []
    for item in getattr(user, "pt_accesses", []) or []:
        pt = (getattr(item, "pt_name", None) or "").strip()
        if pt:
            values.append(pt)
    return sorted(set(values))

def get_me():
    return _current_db_user()

def is_logged_in()->bool:
    return bool(session.get("user_id"))

def master_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_logged_in():
            flash("Silakan login dulu.", "warning")
            return redirect(url_for("main.login"))
        return fn(*args, **kwargs)
    return wrapper

def _scope_pt_names()->list[str]:
    if _is_master_session():
        return []
    return _user_allowed_pts()

def _is_allowed_pt_name(pt_name:str|None)->bool:
    allowed_pts = _scope_pt_names()
    if not allowed_pts:
        return True
    clean = (pt_name or "").strip().lower()
    return bool(clean and clean in {x.strip().lower() for x in allowed_pts})

def _sanitize_company_filter(company:str="") -> str:
    clean = (company or "").strip()
    if clean and not _is_allowed_pt_name(clean):
        return ""
    return clean

def _scoped_company_choices()->list[str]:
    companies = _vehicle_pt_choices()
    allowed_pts = _scope_pt_names()
    if not allowed_pts:
        return companies
    allowed_lookup = {x.strip().lower() for x in allowed_pts if x and x.strip()}
    return [pt for pt in companies if pt and pt.strip().lower() in allowed_lookup]

def _apply_vehicle_scope(query):
    allowed_pts = _scope_pt_names()
    if allowed_pts:
        lowered = [pt.strip().lower() for pt in allowed_pts if pt and pt.strip()]
        if lowered:
            query = query.filter(func.lower(func.trim(Vehicle.pt)).in_(lowered))
        else:
            query = query.filter(Vehicle.id == -1)
    return query

def _assert_vehicle_scope(vehicle:Vehicle):
    allowed_pts = _scope_pt_names()
    if not allowed_pts:
        return
    pt = (getattr(vehicle, "pt", None) or "").strip().lower()
    allowed = {x.strip().lower() for x in allowed_pts}
    if pt not in allowed:
        raise PermissionError("Anda tidak punya akses ke PT ini.")

def format_tgl_id_full(d)->str|None:
    if not d:return None
    if hasattr(d,"strftime"):
        day=d.strftime("%d");year=d.strftime("%Y");month_num=int(d.strftime("%m"))
    else:return str(d)
    bulan={1:"Januari",2:"Februari",3:"Maret",4:"April",5:"Mei",6:"Juni",7:"Juli",8:"Agustus",9:"September",10:"Oktober",11:"November",12:"Desember"}[month_num]
    return f"{day} {bulan} {year}"

def _coerce_datetime_value(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        text_value = value.strip()
        if not text_value:
            return None
        try:
            return datetime.fromisoformat(text_value.replace("Z", "+00:00"))
        except Exception:
            pass
        for fmt in (
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
        ):
            try:
                return datetime.strptime(text_value, fmt)
            except Exception:
                continue
    return None

def _history_timestamp_value(row, *field_names):
    for field_name in field_names:
        if not field_name:
            continue
        value = getattr(row, field_name, None)
        coerced = _coerce_datetime_value(value)
        if coerced:
            return coerced
    return None

def _get_now_jkt():
    return datetime.now(pytz.timezone('Asia/Jakarta'))

def _now_naive() -> datetime:
    """Waktu sekarang (WIB) sebagai naive datetime tanpa tzinfo.
    Gunakan ini untuk semua kolom DATETIME/TIMESTAMP di MySQL agar tidak
    ada konversi UTC yang menyebabkan jam bergeser atau tampil 00:00."""
    return _get_now_jkt().replace(microsecond=0, tzinfo=None)

def _assign_model_timestamp(instance, primary_field:str, fallback_fields:tuple[str,...]=()):
    now = _now_naive()
    assigned = False
    for field_name in (primary_field, *fallback_fields):
        if field_name and hasattr(instance, field_name):
            setattr(instance, field_name, now)
            assigned = True
    return assigned

def _looks_like_date_only_string(value: str) -> bool:
    """True jika string hanya berisi tanggal, tanpa komponen jam."""
    text_value = (value or "").strip()
    if not text_value:
        return False
    # Format tanggal yang biasa dipakai aplikasi: YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY.
    return bool(re.fullmatch(r"(?:\d{4}-\d{2}-\d{2}|\d{2}-\d{2}-\d{4}|\d{2}/\d{2}/\d{4})", text_value))


def _has_real_time(value) -> bool:
    """Cek apakah value punya jam yang memang tersimpan, bukan default 00:00 dari kolom DATE."""
    if isinstance(value, datetime):
        return (value.hour, value.minute, value.second, value.microsecond) != (0, 0, 0, 0)
    if isinstance(value, date):
        return False
    if isinstance(value, str):
        return not _looks_like_date_only_string(value) and bool(re.search(r"\d{1,2}:\d{2}", value))
    return False


def format_tgl_jam_id_full(dt)->str|None:
    if not dt:
        return None

    # Penting: kolom DATE dari database akan dibaca Python sebagai date dan jamnya
    # otomatis menjadi 00:00 jika dipaksa ke datetime. Jangan tampilkan jam palsu ini,
    # supaya history Pajak/KIR/Servis/Peminjaman tidak terlihat berubah menjadi 00:00
    # setelah halaman dibuka besoknya.
    if isinstance(dt, datetime):
        dt_value = dt
        if dt_value.tzinfo:
            dt_value = dt_value.astimezone(pytz.timezone('Asia/Jakarta'))
        return f"{format_tgl_id_full(dt_value.date())} {dt_value.strftime('%H:%M')}"

    if isinstance(dt, date):
        return format_tgl_id_full(dt)

    if isinstance(dt, str) and _looks_like_date_only_string(dt):
        d_value = _parse_date(dt)
        return format_tgl_id_full(d_value) if d_value else dt

    dt_value = _coerce_datetime_value(dt)
    if dt_value:
        if dt_value.tzinfo:
            dt_value = dt_value.astimezone(pytz.timezone('Asia/Jakarta'))
        # Kalau hasil parsing tetap 00:00 dari string tanggal saja, tampilkan tanggalnya saja.
        if not _has_real_time(dt):
            return format_tgl_id_full(dt_value.date())
        return f"{format_tgl_id_full(dt_value.date())} {dt_value.strftime('%H:%M')}"
    return str(dt)

def rupiah(v)->str:
    if v is None or v=="":return "Rp 0"
    try:n=int(float(v))
    except Exception:return f"Rp {v}"
    return "Rp "+f"{n:,}".replace(",",".")

def vehicle_display_name(v:Vehicle|None)->str:
    if not v:return "Kendaraan"
    name = "Kendaraan"
    # Prioritaskan active_name (Nama Aktiva) sesuai permintaan user
    for attr in ("active_name", "new_asset_name", "name_as_asset_pt"):
        val = getattr(v, attr, None)
        if val and str(val).strip():
            name = str(val).strip()
            break
    else:
        merk = getattr(v, "merk", None)
        tipe = getattr(v, "type", None)
        if merk and tipe:
            name = f"{merk} {tipe}".strip()
        elif merk:
            name = str(merk).strip()
        else:
            jenis = getattr(v, "jenis", None)
            if jenis:
                name = str(jenis).strip()
    
    # Tambahkan informasi PT jika ada
    pt = getattr(v, "pt", None)
    if pt and str(pt).strip():
        return f"[{str(pt).strip()}] {name}"
    return name

def vehicle_plate(v:Vehicle|None)->str:
    if not v:return "-"
    for attr in ("plate_new","plate_old","plate","no_polisi"):
        val=getattr(v,attr,None)
        if val and str(val).strip():return str(val).strip()
    return "-"


def _normalize_plate_value(value) -> str:
    """Normalisasi nomor polisi: kosongkan nilai placeholder seperti '-'."""
    clean = _normalize_text(value).upper()
    if clean.lower() in {"-", "none", "null", "nan"}:
        return ""
    return " ".join(clean.split())

def days_left(d):
    if not d:return None
    return (d-date.today()).days

def _add_6_months(d):
    """Tambah 6 bulan kalender: tgl 13 Jan → tgl 13 Jul, dst."""
    if not d:
        return None
    month = d.month + 6
    year = d.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    import calendar
    max_day = calendar.monthrange(year, month)[1]
    return d.replace(year=year, month=month, day=min(d.day, max_day))

def service_due_date(r):
    if not r: return None
    
    # Ambil jenis servis dari berbagai kemungkinan atribut model
    st = (getattr(r, "service_type", None) or 
          getattr(r, "jenis_servis", None) or 
          getattr(r, "jenis_service", None) or 
          getattr(r, "note", "") or "")
    
    # Permintaan user: Hanya servis rutin yang merubah tanggal servis rutin berikutnya
    # Jika mengandung kata 'berat', maka ini servis berat dan tidak merubah jadwal rutin
    if "berat" in str(st).lower():
        return None
        
    # Default: jika tidak ada kata 'berat', kita anggap ini rutin (6 bulanan)
    # Ini memastikan kolom Next Servis di history terisi otomatis untuk semua servis rutin.
    base = getattr(r, "service_date", None)
    return _add_6_months(base) if base else None

def _service_status(next_due):
    if not next_due:
        return "empty"
    days=(next_due-date.today()).days
    if days < 0:
        return "overdue"
    if days <= 30:
        return "soon"
    return "safe"

def _normalize_text(value)->str:
    if value is None:return ""
    text=str(value).strip()
    if text.lower() in {"nan","none","null"}:return ""
    return " ".join(text.split())

def _normalize_header(value)->str:
    text=_normalize_text(value).lower()
    return HEADER_ALIASES.get(text,text.replace(".","").replace("_"," ").strip())

def _parse_year(value):
    if value in (None,""):return None
    if isinstance(value,(datetime,date)):return value.year
    text=_normalize_text(value)
    if not text:return None
    try:num=int(float(text.replace(",","")))
    except Exception:return None
    if 1900<=num<=2100:return num
    return None

def _parse_date(value):
    if value in (None,""):return None
    if isinstance(value,date):return value
    text=_normalize_text(value)
    if not text:return None
    for fmt in ("%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%Y/%m/%d"):
        try:return datetime.strptime(text,fmt).date()
        except Exception:pass
    try:return datetime.fromisoformat(text).date()
    except Exception:return None

def _parse_date_as_datetime(value):
    """Parse tanggal dari form input (YYYY-MM-DD) dan gabungkan dengan jam SEKARANG (WIB).
    Dipakai untuk kolom paid_date, done_date, service_date, date_out, date_return_actual
    agar jam tersimpan real-time, bukan 00:00.

    Catatan: mengembalikan naive datetime (tanpa tzinfo) agar kompatibel dengan
    kolom DATETIME di MySQL — timezone sudah dikonversi ke WIB sebelum disimpan.
    """
    d = _parse_date(value)
    if not d:
        return None
    now_jkt = _get_now_jkt()
    # Ambil jam WIB sekarang, gabungkan dengan tanggal dari form, hasilkan naive datetime
    return datetime.combine(d, now_jkt.time().replace(microsecond=0, tzinfo=None))

def _today_as_datetime():
    """Return waktu sekarang (WIB) sebagai naive datetime — untuk default paid_date/done_date kalau form kosong."""
    now_jkt = _get_now_jkt()
    return now_jkt.replace(microsecond=0, tzinfo=None)

def _split_foto_urls(value: str | None) -> list[str]:
    """Pecah daftar foto yang tersimpan di keterangan TERJUAL."""
    if not value:
        return []
    text_value = str(value).strip()
    if not text_value or text_value in {'-', 'None', 'null'}:
        return []
    parts = re.split(r"\s*(?:\|\||;|,)\s*", text_value)
    urls = []
    seen = set()
    for part in parts:
        normalized = _normalize_foto_url(part.strip())
        if normalized and normalized not in seen:
            seen.add(normalized)
            urls.append(normalized)
    return urls


def _join_foto_urls(urls: list[str] | tuple[str, ...] | None) -> str:
    """Gabungkan daftar foto untuk disimpan ke blok TERJUAL."""
    cleaned = []
    seen = set()
    for url in urls or []:
        normalized = _normalize_foto_url(url)
        if normalized and normalized not in seen:
            seen.add(normalized)
            cleaned.append(normalized)
    return " || ".join(cleaned)


def _parse_sold_info(keterangan: str) -> dict:
    info = {"sold_date": None, "sold_to": "-", "price": 0, "terbilang": "-", "note": "-", "nik": "-", "npwp": "-", "foto_url": None, "foto_urls": []}
    if not keterangan:
        return info

    sold_parts = re.findall(r"TERJUAL pada .*?(?=\s*\|\s*TERJUAL pada|$)", keterangan, flags=re.IGNORECASE | re.DOTALL)
    source = sold_parts[-1] if sold_parts else keterangan

    match_date = re.search(r"TERJUAL pada (\d{4}-\d{2}-\d{2})", source, flags=re.IGNORECASE)
    match_to   = re.search(r"kepada (.+?) seharga", source, flags=re.IGNORECASE | re.DOTALL)
    match_prc  = re.search(r"seharga Rp ([\d.,]+)", source, flags=re.IGNORECASE)
    match_ter  = re.search(r"Rp [\d.,]+ \((.+?)\)", source, flags=re.IGNORECASE | re.DOTALL)
    match_nik  = re.search(r"NIK/KTP: (.+?)(?:\.\s*(?:NPWP:|Catatan:|Foto:)|$)", source, flags=re.IGNORECASE | re.DOTALL)
    match_npwp = re.search(r"NPWP: (.+?)(?:\.\s*(?:Catatan:|Foto:)|$)", source, flags=re.IGNORECASE | re.DOTALL)
    match_note = re.search(r"Catatan: (.+?)(?:\.\s*Foto:|$)", source, flags=re.IGNORECASE | re.DOTALL)
    match_foto = re.search(r"Foto: (.+?)$", source, flags=re.IGNORECASE | re.DOTALL)

    if match_date: info["sold_date"] = _parse_date(match_date.group(1))
    if match_to:   info["sold_to"] = match_to.group(1).strip()
    if match_prc:  info["price"] = _parse_float(match_prc.group(1).replace(".", "")) or 0
    if match_ter:  info["terbilang"] = match_ter.group(1).strip()
    if match_nik:  info["nik"] = match_nik.group(1).strip()
    if match_npwp: info["npwp"] = match_npwp.group(1).strip()
    if match_note: info["note"] = match_note.group(1).strip()
    if match_foto:
        foto_urls = _split_foto_urls(match_foto.group(1).strip())
        info["foto_urls"] = foto_urls
        info["foto_url"] = foto_urls[0] if foto_urls else None
    return info

def _normalize_foto_url(value: str | None) -> str | None:
    if not value:
        return None
    url = str(value).strip().replace('\\', '/')
    if not url or url in {'-', 'None', 'null'}:
        return None
    if url.startswith('http://') or url.startswith('https://'):
        return url
    if '/kendaraan/foto/' in url:
        filename = url.rsplit('/kendaraan/foto/', 1)[-1].split('?', 1)[0].split('#', 1)[0]
        return url_for('main.kendaraan_foto', filename=os.path.basename(filename), _external=False)
    filename = os.path.basename(url.split('?', 1)[0].split('#', 1)[0])
    if filename:
        return url_for('main.kendaraan_foto', filename=filename, _external=False)
    return None

def _resolve_foto_path(foto_url: str | None) -> Path | None:
    """Cari file foto dari nilai yang tersimpan di database."""
    if not foto_url:
        return None
    raw = str(foto_url).strip().strip('"').strip("'").replace('\\', '/')
    if not raw or raw.lower() in {'-', 'none', 'null'}:
        return None
    if raw.startswith('http://') or raw.startswith('https://'):
        return None
    normalized = _normalize_foto_url(raw)
    filename = os.path.basename((normalized or raw).rsplit('/', 1)[-1].split('?', 1)[0].split('#', 1)[0])
    if not filename:
        return None
    try:
        app_root = Path(current_app.root_path)
    except RuntimeError:
        app_root = Path(__file__).resolve().parent
    cwd = Path.cwd()
    candidates = [
        _foto_upload_dir() / filename,
        app_root / 'static' / 'uploads' / 'terjual_foto' / filename,
        app_root / 'static' / 'uploads' / 'kendaraan' / 'foto' / filename,
        app_root / 'static' / 'kendaraan' / 'foto' / filename,
        app_root / 'uploads' / 'terjual_foto' / filename,
        app_root / 'uploads' / 'kendaraan' / 'foto' / filename,
        cwd / 'static' / 'uploads' / 'terjual_foto' / filename,
        cwd / 'static' / 'uploads' / 'kendaraan' / 'foto' / filename,
        cwd / 'uploads' / 'terjual_foto' / filename,
        cwd / 'uploads' / 'kendaraan' / 'foto' / filename,
        Path(raw) if Path(raw).is_absolute() else None,
    ]
    seen = set()
    for path in candidates:
        if not path:
            continue
        try:
            key = str(path.resolve())
            if key in seen:
                continue
            seen.add(key)
            if path.exists() and path.is_file():
                return path
        except Exception:
            continue
    for base in (app_root / 'static', app_root / 'uploads', cwd / 'static', cwd / 'uploads'):
        try:
            if base.exists():
                found = next(base.rglob(filename), None)
                if found and found.is_file():
                    return found
        except Exception:
            continue
    return None
def _parse_int(value):
    if value in (None,""):return None
    try:return int(float(str(value).replace(",","").strip()))
    except Exception:return None

def _parse_float(value):
    if value in (None,""):return None
    try:return float(str(value).replace(",","").strip())
    except Exception:return None

def _active_vehicle_base_query():
    """Query kendaraan aktif: belum dihapus dan bukan status TERJUAL.
    Dipakai untuk daftar kendaraan, pilihan PT export, dan export Excel agar data terjual
    hanya muncul di laporan terjual, bukan di data kendaraan aktif.
    """
    return Vehicle.query.filter(
        Vehicle.is_deleted == False,
        or_(Vehicle.status.is_(None), ~func.lower(func.trim(Vehicle.status)).like("%terjual%"))
    )

def _vehicle_pt_choices()->list[str]:
    rows=(
        _active_vehicle_base_query()
        .with_entities(Vehicle.pt)
        .filter(Vehicle.pt.isnot(None),func.trim(Vehicle.pt)!="")
        .distinct()
        .order_by(Vehicle.pt.asc())
        .all()
    )
    return [r[0] for r in rows if r[0] and str(r[0]).strip()]

def _find_company_by_name(name:str):
    clean=_normalize_text(name)
    if not clean:return None
    return Company.query.filter(func.lower(Company.name)==clean.lower()).first()

def _get_or_create_company(name:str):
    clean=_normalize_text(name)
    if not clean:return None
    company=_find_company_by_name(clean)
    if company:return company
    session_db=Vehicle.query.session
    company=Company(name=clean)
    session_db.add(company)
    session_db.flush()
    return company

def _company_name(company):
    if not company:return ""
    if isinstance(company,int):
        c=Company.query.get(company)
        return c.name if c else ""
    return company.name if getattr(company,"name",None) else ""

def _vehicle_snapshot(v:Vehicle, include_custom=True):
    """Snapshot data kendaraan untuk perbandingan history. Sekarang include kolom kustom."""
    snapshot = {
        "no":getattr(v,"no",None),
        "pt":v.pt or "",
        "asset_owner_company_name":_company_name(getattr(v,"asset_owner_company",None)),
        "pt_pemakai_company_name":_company_name(getattr(v,"pt_pemakai_company",None)),
        "active_name":v.active_name or "",
        "name_as_asset_pt":v.name_as_asset_pt or "",
        "new_asset_name":v.new_asset_name or "",
        "merk":v.merk or "",
        "type":v.type or "",
        "jenis":v.jenis or "",
        "plate_old":v.plate_old or "",
        "plate_new":v.plate_new or "",
        "year_of_use":v.year_of_use,
        "user_old":v.user_old or "",
        "user_new":v.user_new or "",
        "status":v.status or "",
        "kondisi_terkini":v.kondisi_terkini or "",
        "lokasi":v.lokasi or "",
        "tambahan_keterangan":v.tambahan_keterangan or "",
    }
    
    # Tambahkan kolom kustom jika include_custom=True
    if include_custom:
        custom_values = _vehicle_custom_values_map([v.id]).get(v.id, {})
        snapshot.update(custom_values)
    
    return snapshot

def _label_map(include_custom=True):
    """Map nama field ke label display. Sekarang include kolom kustom."""
    labels = {
        "no":"No","pt":"PT","asset_owner_company_name":"PT Pemilik Aset","pt_pemakai_company_name":"PT Pemakai",
        "active_name":"NAMA AKTIVA","name_as_asset_pt":"Nama sesuai asset PT","new_asset_name":"Nama Aset baru","merk":"MERK","type":"TYPE","jenis":"JENIS",
        "plate_old":"NO POLISI LAMA","plate_new":"NO POLISI BARU","year_of_use":"TAHUN PEMAKAIAN","user_old":"USER LAMA","user_new":"USER BARU",
        "status":"Status","kondisi_terkini":"Kondisi terkini","lokasi":"Lokasi","tambahan_keterangan":"Tambahan Keterangan"
    }
    
    # Tambahkan label untuk kolom kustom
    if include_custom:
        custom_columns = _vehicle_custom_columns()
        for col in custom_columns:
            labels[col['column_key']] = col['column_label']
    
    return labels

def _changed_fields(before:dict, after:dict, include_custom=True):
    """Deteksi field mana saja yang berubah. Sekarang include kolom kustom."""
    labels = _label_map(include_custom=include_custom)
    changes = []
    
    for k in labels.keys():
        b = before.get(k)
        a = after.get(k)
        bcmp = "" if b is None else str(b)
        acmp = "" if a is None else str(a)
        if bcmp != acmp:
            changes.append({
                "field_name": k,
                "field_label": labels[k],
                "old_value": b,
                "new_value": a
            })
    
    return changes

def _log_vehicle_edit(vehicle:Vehicle,before:dict,after:dict,changed_by:str="system"):
    session_db=Vehicle.query.session
    changes=_changed_fields(before,after,include_custom=True)  # <-- Sekarang include custom
    if not changes:return
    for item in changes:
        vh = VehicleChangeHistory(
            vehicle_id=vehicle.id,
            field_name=item["field_name"],
            field_label=item["field_label"],
            old_value=None if item["old_value"] is None else str(item["old_value"]),
            new_value=None if item["new_value"] is None else str(item["new_value"]),
            change_type="update_master",
            changed_by=changed_by,
            note=f'Perubahan {item["field_label"]}'
        )
        _assign_model_timestamp(vh, "changed_at", ("change_date", "created_at", "updated_at"))
        session_db.add(vh)
        if any(x["field_name"] in {"user_old", "user_new"} for x in changes):
            user_history_row = UserHistory(
                vehicle_id=vehicle.id,
                user_lama=(before.get("user_new") or before.get("user_old") or ""),
                user_baru=(after.get("user_new") or after.get("user_old") or ""),
                note="Perubahan user dari data master kendaraan"
            )
            _assign_model_timestamp(user_history_row, "change_date", ("changed_at", "created_at", "updated_at"))
            session_db.add(user_history_row)

    payload={"kind":"vehicle_edit","vehicle_id":vehicle.id,"vehicle_name":vehicle_display_name(vehicle),"plate":vehicle_plate(vehicle),"changed_at":_get_now_jkt().strftime("%Y-%m-%d %H:%M:%S"),"editor":changed_by,"changes":[{"field":x["field_name"],"label":x["field_label"],"before":x["old_value"],"after":x["new_value"]} for x in changes]}
    note=json.dumps(payload,ensure_ascii=False)
    try:
        audit("UPDATE", "Vehicle", vehicle.id, note=note)
    except Exception as exc:
        print("Audit fallback gagal/lewat:", exc)
def _vehicle_query_filtered(q:str="",company:str=""):
    query=_apply_vehicle_scope(_active_vehicle_base_query())
    if q:
        like=f"%{q}%"
        query=query.filter(or_(
            Vehicle.pt.ilike(like),Vehicle.active_name.ilike(like),Vehicle.name_as_asset_pt.ilike(like),Vehicle.new_asset_name.ilike(like),
            Vehicle.merk.ilike(like),Vehicle.type.ilike(like),Vehicle.jenis.ilike(like),Vehicle.plate_old.ilike(like),Vehicle.plate_new.ilike(like),
            Vehicle.user_old.ilike(like),Vehicle.user_new.ilike(like),Vehicle.status.ilike(like),Vehicle.kondisi_terkini.ilike(like),
            Vehicle.lokasi.ilike(like),Vehicle.tambahan_keterangan.ilike(like)
        ))
    if company:query=query.filter(func.lower(func.trim(Vehicle.pt))==company.strip().lower())
    return query

def _filtered_vehicle_ids(q:str="",company:str="")->list[int]|None:
    if not q and not company:return None
    rows=_vehicle_query_filtered(q,company).with_entities(Vehicle.id).all()
    return [vid for (vid,) in rows]


def _kir_status(next_due):
    if not next_due:
        return "empty"
    days=(next_due-date.today()).days
    if days < 0:
        return "overdue"
    if days <= DEFAULT_REMINDER_DAYS:
        return "soon"
    return "safe"


def _kir_display_result(record):
    """Hasil KIR hanya tampil kalau benar-benar diisi lewat proses KIR.

    Tombol Edit Tanggal & Catatan dipakai sebagai anchor/jadwal awal, bukan
    bukti kendaraan sudah KIR. Versi lama sempat otomatis mengisi
    result=Lulus dan status=Selesai saat edit tanggal; pola legacy itu
    disembunyikan supaya kendaraan yang belum diproses KIR tidak terlihat
    sudah lulus.
    """
    if not record:
        return None
    result = getattr(record, "result", None)
    status = getattr(record, "status", None)
    pay_date = getattr(record, "pay_date", None)
    note = getattr(record, "note", None)
    legacy_auto_result = (
        _normalize_text(result).lower() == "lulus"
        and _normalize_text(status).lower() == "selesai"
        and not pay_date
        and not note
    )
    if legacy_auto_result:
        return None
    return result

def _latest_kir_map(vehicle_ids:list[int]):
    """
    Mengambil data KIR terakhir untuk setiap kendaraan.
    Kita urutkan berdasarkan ID terbaru (id.desc()) agar data yang baru saja 
    diproses atau diedit langsung muncul di tabel 'Data KIR' utama.
    """
    rows={}
    if not vehicle_ids:return rows
    kir_rows=(
        KirRecord.query
        .filter(KirRecord.vehicle_id.in_(vehicle_ids))
        .order_by(KirRecord.vehicle_id.asc(), KirRecord.id.desc())
        .all()
    )
    for r in kir_rows:
        if r.vehicle_id not in rows:rows[r.vehicle_id]=r
    return rows

def _days_until(d):
    if not d:
        return None
    return (d - date.today()).days


def _status_from_days(days):
    if days is None:
        return "empty"
    if days < 0:
        return "overdue"
    if days <= 14:
        return "soon"
    return "safe"


def _alert_sisa_text(days):
    if days is None:
        return "-"
    if days < 0:
        return f"Terlambat {abs(days)} hari"
    if days == 0:
        return "Hari ini"
    return f"{days} hari lagi"


def _alert_sisa_class(days):
    status = _status_from_days(days)
    if status == "overdue":
        return "sisa-merah"
    if status == "soon":
        return "sisa-kuning"
    return "sisa-hijau"


def _build_alert_entry(category:str, vehicle, due_date, link:str, label:str|None=None, extra:dict|None=None):
    days = _days_until(due_date)
    entry = {
        "type": category,
        "label": label or category.title(),
        "vehicle": vehicle,
        "due": due_date,
        "due_display": format_tgl_id_full(due_date),
        "days_left": days,
        "status": _status_from_days(days),
        "sisa_text": _alert_sisa_text(days),
        "sisa_class": _alert_sisa_class(days),
        "link": link,
    }
    if extra:
        entry.update(extra)
    return entry


def _home_alert_groups(q:str="", company:str=""):
    today = date.today()
    vehicles = (
        _vehicle_query_filtered(q, company)
        .order_by(Vehicle.updated_at.desc(), Vehicle.id.desc())
        .limit(1000)
        .all()
    )
    vehicle_map = {v.id: v for v in vehicles}
    vehicle_ids = list(vehicle_map.keys())

    groups = {"tax": [], "kir": [], "service": [], "loan": []}

    five_map = _latest_five_tax_map(vehicle_ids)
    for vehicle_id in vehicle_ids:
        vehicle = vehicle_map[vehicle_id]
        five = five_map.get(vehicle_id)
        due = five.get("due_date") if five else None
        paid_date = five.get("paid_date") if five else None
        if not due or paid_date:
            continue
        days = _days_until(due)
        if days is not None and days <= 60:
            groups["tax"].append(_build_alert_entry(
                "tax",
                vehicle,
                due,
                url_for("main.go_data_pajak") + f"?highlight={vehicle_id}",
                label="Pajak 5 Tahunan",
                extra={
                    "kind": "five",
                    "record_id": five.get("id"),
                    "amount": five.get("amount") or 0,
                    "note": five.get("note"),
                    "plate_before": five.get("plate_before") or getattr(vehicle, "plate_old", None),
                    "plate_after": five.get("plate_after") or getattr(vehicle, "plate_new", None),
                }
            ))

    # ── Pajak TAHUNAN ──
    # Pakai _annual_tax_cycle agar kendaraan yang belum pernah bayar pun ikut dicek
    # main_due_date tahunan = five_year due_date dikurangi 5 tahun (via five_year_due_to_annual_due)
    annual_map_alert = _annual_payment_map(vehicle_ids)
    for vehicle_id in vehicle_ids:
        vehicle = vehicle_map[vehicle_id]
        five = five_map.get(vehicle_id)
        # main_due_date untuk annual cycle = due date 5 tahunan (patokan siklus)
        main_due = five.get("due_date") if five else None
        if not main_due:
            # Fallback: cek di Vehicle model langsung
            main_due = getattr(vehicle, "annual_due_date", None) or getattr(vehicle, "tax_due_date", None)
        if not main_due:
            continue
        try:
            annual_info = _annual_tax_cycle(vehicle_id, main_due, annual_map_alert)
        except Exception:
            continue
        annual_due = annual_info.get("display_due_date") or annual_info.get("current_due_date")
        annual_status = annual_info.get("display_status", "empty")
        if not annual_due or annual_status == "paid":
            continue
        days = _days_until(annual_due)
        # Pajak tahunan muncul di notifikasi saat H-30 atau sudah lewat jatuh tempo,
        # selama belum lunas. Pajak 5 tahunan tetap H-60.
        if days is not None and days <= 30:
            current_payment = annual_info.get("current_payment")
            groups["tax"].append(_build_alert_entry(
                "tax",
                vehicle,
                annual_due,
                url_for("main.go_data_pajak") + f"?highlight={vehicle_id}",
                label="Pajak Tahunan",
                extra={
                    "kind": "annual",
                    "record_id": current_payment.get("id") if current_payment else None,
                    "amount": (current_payment.get("amount") if current_payment else 0) or 0,
                }
            ))

    latest_kir_map = _latest_kir_map(vehicle_ids)
    for vehicle_id, record in latest_kir_map.items():
        due = getattr(record, "due_date", None)
        days = _days_until(due)
        if due and days is not None and days <= 30:
            groups["kir"].append(_build_alert_entry("kir", vehicle_map[vehicle_id], due, url_for("main.go_data_kir") + f"?highlight={vehicle_id}", label="KIR"))

    # ── SERVIS ──
    # Alert servis harus mengikuti JATUH TEMPO servis yang disimpan lewat tombol Edit
    # di Data Servis (record anchor [JADWAL_SERVIS]). Sebelumnya alert dihitung dari
    # transaksi servis terakhir sehingga jadwal yang baru dibuat/diubah bisa tidak muncul
    # di Beranda.
    latest_service_anchor_map = {}
    latest_processed_service_map = {}
    service_rows = (
        ServiceRecord.query
        .filter(ServiceRecord.vehicle_id.in_(vehicle_ids))
        .order_by(ServiceRecord.vehicle_id.asc(), ServiceRecord.service_date.desc(), ServiceRecord.id.desc())
        .all()
    ) if vehicle_ids else []
    for record in service_rows:
        note_text = _normalize_text(getattr(record, "note", None)).lower()
        is_anchor = "jadwal_servis" in note_text
        if is_anchor:
            if record.vehicle_id not in latest_service_anchor_map:
                latest_service_anchor_map[record.vehicle_id] = record
        elif record.vehicle_id not in latest_processed_service_map:
            latest_processed_service_map[record.vehicle_id] = record

    for vehicle_id in vehicle_ids:
        anchor = latest_service_anchor_map.get(vehicle_id)
        processed = latest_processed_service_map.get(vehicle_id)
        # Prioritas utama: due date anchor. Fallback hanya untuk data lama yang belum punya anchor.
        due = getattr(anchor, "service_date", None) if anchor else service_due_date(processed)
        days = _days_until(due)
        if due and days is not None and days <= 30:
            groups["service"].append(_build_alert_entry(
                "service",
                vehicle_map[vehicle_id],
                due,
                url_for("main.go_data_servis") + f"?highlight={vehicle_id}",
                label="Servis"
            ))

    latest_loan_map = {}
    loan_rows = (
        LoanTransaction.query
        .filter(LoanTransaction.vehicle_id.in_(vehicle_ids))
        .order_by(LoanTransaction.vehicle_id.asc(), LoanTransaction.created_at.desc(), LoanTransaction.id.desc())
        .all()
    ) if vehicle_ids else []
    for record in loan_rows:
        if record.vehicle_id not in latest_loan_map:
            latest_loan_map[record.vehicle_id] = record
    for vehicle_id, record in latest_loan_map.items():
        due = getattr(record, "date_return_plan", None)
        status = _normalize_text(getattr(record, "status", None)).lower()
        days = _days_until(due)
        if due and status != "selesai" and days is not None:
            groups["loan"].append(_build_alert_entry("loan", vehicle_map[vehicle_id], due, url_for("main.peminjaman") + f"?highlight={vehicle_id}", label="Peminjaman"))

    for key in groups.keys():
        groups[key].sort(key=lambda item: (item["due"] or today, vehicle_plate(item["vehicle"])))

    return groups


def _reminder_items(q:str="", company:str=""):
    groups = _home_alert_groups(q=q, company=company)
    items = []
    for key in ("tax", "kir", "service", "loan"):
        items.extend(groups[key])
    items.sort(key=lambda x: (x["due"] or date.today(), x["label"]))
    return items[:50]

def _notif_summary(q:str="", company:str=""):
    groups = _home_alert_groups(q=q, company=company)
    return {
        "all": sum(len(v) for v in groups.values()),
        "tax": len(groups["tax"]),
        "kir": len(groups["kir"]),
        "service": len(groups["service"]),
        "loan": len(groups["loan"]),
    }

def _read_upload_rows(file_storage):
    workbook=load_workbook(file_storage,data_only=True)
    sheet=workbook[workbook.sheetnames[0]]
    rows=list(sheet.iter_rows(values_only=True))
    if not rows:raise ValueError("File Excel kosong.")
    headers=[_normalize_header(cell) for cell in rows[0]]
    required=["no","pt","pt_pemilik_aset","nama_aktiva","nama_sesuai_asset_pt","nama_aset_baru","merk","type","jenis","no_polisi_lama","no_polisi_baru","tahun_pemakaian","user_lama","user_baru","status","pt_pemakai","kondisi_terkini","lokasi","tambahan_keterangan"]
    missing=[k for k in required if k not in headers]
    if missing:raise ValueError("Header kolom tidak sesuai template: "+", ".join(missing))
    parsed_rows=[]
    for excel_row_index,row in enumerate(rows[1:],start=2):
        row_map={}
        for idx,header_key in enumerate(headers):
            if not header_key:continue
            row_map[header_key]=row[idx] if idx<len(row) else None
        if not any(_normalize_text(v) for v in row_map.values()):continue
        parsed_rows.append((excel_row_index,row_map))
    return parsed_rows

def _import_vehicle_rows(file_storage,actor_name="system",allowed_pts:list[str]|None=None):
    """Import kendaraan dari Excel. Jika allowed_pts diberikan (admin PT), hanya baris dengan PT
    yang ada di allowed_pts yang diproses; baris di luar scope di-skip dengan peringatan."""
    session_db=Vehicle.query.session
    parsed_rows=_read_upload_rows(file_storage)
    created=0
    updated=0
    skipped_pts: set[str] = set()

    def find_existing_vehicle(pt,no_number,plate_old,plate_new):
        vehicle=None

        plate_candidates=[]
        if plate_new:
            plate_candidates.append(plate_new)
        if plate_old and plate_old not in plate_candidates:
            plate_candidates.append(plate_old)

        for plate in plate_candidates:
            vehicle=Vehicle.query.filter(
                Vehicle.is_deleted==False,
                or_(
                    Vehicle.plate_new==plate,
                    Vehicle.plate_old==plate,
                )
            ).order_by(Vehicle.id.asc()).first()
            if vehicle:
                return vehicle

        if plate_old and plate_new:
            vehicle=Vehicle.query.filter(
                Vehicle.is_deleted==False,
                or_(
                    Vehicle.plate_new==plate_old,
                    Vehicle.plate_old==plate_new,
                )
            ).order_by(Vehicle.id.asc()).first()
            if vehicle:
                return vehicle

        if no_number is not None and pt:
            vehicle=Vehicle.query.filter(
                Vehicle.is_deleted==False,
                Vehicle.no==no_number,
                func.lower(func.trim(Vehicle.pt))==pt.strip().lower()
            ).order_by(Vehicle.id.asc()).first()
            if vehicle:
                return vehicle

        if pt and plate_new:
            vehicle=Vehicle.query.filter(
                Vehicle.is_deleted==False,
                func.lower(func.trim(Vehicle.pt))==pt.strip().lower(),
                or_(
                    Vehicle.plate_new==plate_new,
                    Vehicle.plate_old==plate_new,
                )
            ).order_by(Vehicle.id.asc()).first()
            if vehicle:
                return vehicle

        if pt and plate_old:
            vehicle=Vehicle.query.filter(
                Vehicle.is_deleted==False,
                func.lower(func.trim(Vehicle.pt))==pt.strip().lower(),
                or_(
                    Vehicle.plate_new==plate_old,
                    Vehicle.plate_old==plate_old,
                )
            ).order_by(Vehicle.id.asc()).first()
            if vehicle:
                return vehicle

        return None

    allowed_pts_lower = {x.strip().lower() for x in allowed_pts} if allowed_pts else None

    for excel_row_index,row in parsed_rows:
        pt=_normalize_text(row.get("pt"))
        if not pt:
            raise ValueError(f"Baris {excel_row_index}: kolom PT wajib diisi.")

        # Scope guard: admin PT hanya boleh upload data untuk PT-nya sendiri
        if allowed_pts_lower and pt.lower() not in allowed_pts_lower:
            skipped_pts.add(pt)
            continue

        no_number=_parse_int(row.get("no"))
        plate_old=_normalize_text(row.get("no_polisi_lama"))
        plate_new=_normalize_text(row.get("no_polisi_baru"))

        vehicle=find_existing_vehicle(
            pt=pt,
            no_number=no_number,
            plate_old=plate_old,
            plate_new=plate_new,
        )

        now=_now_naive()

        if vehicle is None:
            vehicle=Vehicle(created_at=now,is_deleted=False)
            session_db.add(vehicle)
            session_db.flush()
            created+=1
            before={}
        else:
            updated+=1
            before=_vehicle_snapshot(vehicle)

        asset_owner=_get_or_create_company(row.get("pt_pemilik_aset"))
        pt_pemakai=_get_or_create_company(row.get("pt_pemakai"))

        vehicle.no=no_number
        vehicle.pt=pt
        vehicle.asset_owner_company_id=asset_owner.id if asset_owner else None
        vehicle.pt_pemakai_company_id=pt_pemakai.id if pt_pemakai else None
        vehicle.active_name=_normalize_text(row.get("nama_aktiva")) or None
        vehicle.name_as_asset_pt=_normalize_text(row.get("nama_sesuai_asset_pt")) or None
        vehicle.new_asset_name=_normalize_text(row.get("nama_aset_baru")) or None
        vehicle.merk=_normalize_text(row.get("merk")) or None
        vehicle.type=_normalize_text(row.get("type")) or None
        vehicle.jenis=_normalize_text(row.get("jenis")) or None
        vehicle.plate_old=plate_old or None
        vehicle.plate_new=plate_new or None
        vehicle.year_of_use=_parse_year(row.get("tahun_pemakaian"))
        vehicle.user_old=_normalize_text(row.get("user_lama")) or None
        vehicle.user_new=_normalize_text(row.get("user_baru")) or None
        vehicle.status=_normalize_text(row.get("status")) or None
        vehicle.kondisi_terkini=_normalize_text(row.get("kondisi_terkini")) or None
        vehicle.lokasi=_normalize_text(row.get("lokasi")) or None
        vehicle.tambahan_keterangan=_normalize_text(row.get("tambahan_keterangan")) or None
        vehicle.is_deleted=False
        vehicle.updated_at=now

        session_db.flush()

        after=_vehicle_snapshot(vehicle)
        if before:
            _log_vehicle_edit(vehicle,before,after,changed_by=actor_name)

    session_db.commit()
    return created,updated,len(parsed_rows),skipped_pts

def _safe_commit(msg_ok:str,msg_err:str,redirect_endpoint:str,**redirect_values):
    try:
        Vehicle.query.session.commit()
        flash(msg_ok,"success")
    except Exception as exc:
        Vehicle.query.session.rollback()
        flash(f"{msg_err}: {exc}","danger")
    return redirect(url_for(redirect_endpoint,**redirect_values))


EDITABLE_VEHICLE_FIELDS={
    "pt":"text",
    "asset_owner_company_name":"company",
    "active_name":"text",
    "name_as_asset_pt":"text",
    "new_asset_name":"text",
    "merk":"text",
    "type":"text",
    "jenis":"text",
    "plate_old":"text",
    "plate_new":"text",
    "year_of_use":"year",
    "user_old":"text",
    "user_new":"text",
    "status":"text",
    "pt_pemakai_company_name":"company",
    "kondisi_terkini":"text",
    "lokasi":"text",
    "tambahan_keterangan":"text",
}

def _set_vehicle_field_value(vehicle:Vehicle,field_name:str,raw_value):
    kind=EDITABLE_VEHICLE_FIELDS.get(field_name)
    if not kind:
        raise ValueError("Kolom ini tidak bisa diedit dari tabel utama.")

    if kind=="company":
        company=_get_or_create_company(raw_value)
        company_id=company.id if company else None
        if field_name=="asset_owner_company_name":
            vehicle.asset_owner_company_id=company_id
        elif field_name=="pt_pemakai_company_name":
            vehicle.pt_pemakai_company_id=company_id
        return _company_name(company)

    if kind=="year":
        parsed=_parse_year(raw_value)
        vehicle.year_of_use=parsed
        return "" if parsed is None else str(parsed)

    value=_normalize_text(raw_value) or None
    setattr(vehicle,field_name,value)
    return value or ""


@bp.post("/kendaraan/<int:vehicle_id>/field")
@master_required
def kendaraan_update_field(vehicle_id:int):
    v=Vehicle.query.get_or_404(vehicle_id)
    
    # Lewati _assert_vehicle_scope(v) di sini agar jika PT baru saja di-set (misal default AKY),
    # user masih bisa mengubahnya ke PT lain (misal AMBRO) yang masih dalam daftar aksesnya.
    # Scope tetap terjaga karena ada validasi _is_allowed_pt_name di bawah.
    if not _is_master_session():
        # Jika bukan master, cek apakah PT kendaraan saat ini ADA dalam daftar akses user
        # Jika tidak ada sama sekali akses ke PT ini, baru blokir.
        if not _is_allowed_pt_name(v.pt):
            raise PermissionError("Anda tidak punya akses ke data PT ini.")

    session_db=Vehicle.query.session
    field_name=_normalize_text(request.form.get("field_name"))
    raw_value=request.form.get("value")

    # Scope Validation: Jika mengedit field 'pt', pastikan PT baru diizinkan
    if field_name == "pt":
        pt_name = _normalize_text(raw_value)
        if not _is_allowed_pt_name(pt_name):
            return jsonify({
                "success": False, 
                "message": f"Gagal: Anda tidak memiliki akses ke PT {pt_name or '-'}."
            }), 403

    actor_name=session.get("user_name") or "SYSTEM"
    before=_vehicle_snapshot(v)

    try:
        custom_key=None
        if field_name in EDITABLE_VEHICLE_FIELDS:
            display_value=_set_vehicle_field_value(v,field_name,raw_value)
        else:
            custom_key=_normalize_text(field_name).lower().replace(' ','_')
            display_value='' if raw_value is None else str(raw_value)
            # Pastikan before punya key ini (kolom baru = belum ada di snapshot)
            if custom_key not in before:
                before[custom_key]=''
            _set_vehicle_custom_value(v.id, custom_key, display_value)
        v.updated_at=_now_naive()
        session_db.flush()
        if field_name=="plate_new":
            _sync_vehicle_plate_update(v)
        # Patch after secara manual karena raw SQL tidak ikut di-read ulang oleh SQLAlchemy
        after=_vehicle_snapshot(v)
        if custom_key:
            after[custom_key]=display_value
        _log_vehicle_edit(v,before,after,changed_by=actor_name)
        session_db.commit()
        return jsonify({
            "success":True,
            "message":"Data berhasil diperbarui.",
            "display_value":display_value or "-",
            "field_name":field_name,
        })
    except Exception as exc:
        session_db.rollback()
        return jsonify({"success":False,"message":str(exc)}),400

@bp.app_context_processor
def inject_globals():
    current_company = _sanitize_company_filter(request.args.get("company") or "") if request else ""
    raw=_reminder_items(company=current_company);top_notif_items=[]
    for it in raw[:8]:
        v=it["vehicle"];due_raw=it["due"]
        top_notif_items.append({"label":it["label"],"vehicle_name":vehicle_display_name(v),"plate":vehicle_plate(v),"due":due_raw,"due_display":format_tgl_id_full(due_raw),"days_left":days_left(due_raw),"link":it["link"]})
    return {"me":get_me(),"REMINDER_DAYS":DEFAULT_REMINDER_DAYS,"reminder_link":url_for("main.kendaraan",due="1",company=current_company) if current_company else url_for("main.kendaraan",due="1"),"top_notif_items":top_notif_items,"notif_summary":_notif_summary(company=current_company),"fmt_tgl":format_tgl_id_full,"fmt_tgl_jam":format_tgl_jam_id_full,"rupiah":rupiah,"vehicle_plate":vehicle_plate,"vehicle_display_name":vehicle_display_name,"service_due_date":service_due_date,"today":date.today()}

@bp.get("/")
@master_required
def home():
    company = _sanitize_company_filter((request.args.get("company") or "").strip())
    alert_groups = _home_alert_groups(company=company)
    raw_items = []
    for key in ("tax", "kir", "service", "loan"):
        raw_items.extend(alert_groups[key])
    raw_items.sort(key=lambda x: (x["due"] or date.today(), x["label"]))
    top_notif_items = []
    for it in raw_items[:8]:
        v = it["vehicle"]
        due_raw = it["due"]
        top_notif_items.append({
            "label": it["label"],
            "vehicle_name": vehicle_display_name(v),
            "plate": vehicle_plate(v),
            "due": due_raw,
            "due_display": format_tgl_id_full(due_raw),
            "days_left": days_left(due_raw),
            "link": it["link"],
        })
    return render_template(
        "home.html",
        title="Sinar Group - Fleet & Asset Vehicle",
        notif_summary=_notif_summary(company=company),
        top_notif_items=top_notif_items,
        alert_groups=alert_groups,
        company=company,
        companies=_scoped_company_choices(),
    )



def _detail_back_url(default_endpoint: str = "main.kendaraan") -> str:
    raw_next = (request.args.get("next") or request.args.get("back") or "").strip()
    if raw_next.startswith("/") and not raw_next.startswith("//"):
        return raw_next

    source = (request.args.get("from") or request.args.get("source") or "").strip().lower()
    if source in {"servis", "service", "data_servis"}:
        return url_for("main.go_data_servis")
    if source in {"pajak", "tax", "data_pajak"}:
        return url_for("main.go_data_pajak")
    if source in {"kir", "data_kir"}:
        return url_for("main.go_data_kir")
    if source in {"peminjaman", "loan", "data_peminjaman"}:
        return url_for("main.peminjaman")
    if source in {"pt", "perusahaan", "ringkasan"}:
        company = (request.args.get("company") or "").strip()
        if company:
            return url_for("main.perusahaan_ringkasan", company=company)

    return url_for(default_endpoint)


def _request_payload():
    if request.is_json:
        return request.get_json(silent=True) or {}
    data = request.form.to_dict(flat=True)
    if not data:
        try:
            data = json.loads(request.get_data(as_text=True) or '{}')
        except Exception:
            data = {}
    return data


_custom_tables_ensured = False

def _ensure_vehicle_custom_tables():
    global _custom_tables_ensured
    if _custom_tables_ensured:
        return
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS vehicle_custom_columns (
            id INT NOT NULL AUTO_INCREMENT,
            column_key VARCHAR(255) NOT NULL,
            column_label VARCHAR(255) NOT NULL,
            position_index INT NOT NULL DEFAULT 999,
            created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uq_vehicle_custom_columns_key (column_key)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """))
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS vehicle_custom_values (
            id INT NOT NULL AUTO_INCREMENT,
            vehicle_id INT NOT NULL,
            column_key VARCHAR(255) NOT NULL,
            value TEXT NULL,
            updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            UNIQUE KEY uq_vehicle_custom_values_vehicle_column (vehicle_id, column_key),
            KEY idx_vehicle_custom_values_vehicle (vehicle_id),
            CONSTRAINT fk_vehicle_custom_values_vehicle
                FOREIGN KEY (vehicle_id) REFERENCES vehicles (id)
                ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """))
    db.session.commit()
    _custom_tables_ensured = True


def _vehicle_custom_columns():
    _ensure_vehicle_custom_tables()
    rows = db.session.execute(text("SELECT column_key, column_label, position_index FROM vehicle_custom_columns ORDER BY position_index ASC, id ASC")).mappings().all()
    return [dict(r) for r in rows]


def _vehicle_custom_values_map(vehicle_ids:list[int]):
    _ensure_vehicle_custom_tables()
    result = {}
    if not vehicle_ids:
        return result
    rows = db.session.execute(
        text("SELECT vehicle_id, column_key, value FROM vehicle_custom_values WHERE vehicle_id IN :ids").bindparams(bindparam('ids', expanding=True)),
        {'ids': vehicle_ids}
    ).mappings().all()
    for r in rows:
        result.setdefault(r['vehicle_id'], {})[r['column_key']] = r['value'] or ''
    return result


def _set_vehicle_custom_value(vehicle_id:int, column_key:str, value):
    _ensure_vehicle_custom_tables()
    value = '' if value is None else str(value)
    exists = db.session.execute(text("SELECT id FROM vehicle_custom_values WHERE vehicle_id=:vehicle_id AND column_key=:column_key"), {'vehicle_id':vehicle_id,'column_key':column_key}).first()
    if exists:
        db.session.execute(text("UPDATE vehicle_custom_values SET value=:value, updated_at=CURRENT_TIMESTAMP WHERE vehicle_id=:vehicle_id AND column_key=:column_key"), {'value':value,'vehicle_id':vehicle_id,'column_key':column_key})
    else:
        db.session.execute(text("INSERT INTO vehicle_custom_values (vehicle_id, column_key, value) VALUES (:vehicle_id, :column_key, :value)"), {'vehicle_id':vehicle_id,'column_key':column_key,'value':value})


def _delete_vehicle_custom_value(vehicle_id:int, column_key:str):
    _ensure_vehicle_custom_tables()
    db.session.execute(text("DELETE FROM vehicle_custom_values WHERE vehicle_id=:vehicle_id AND column_key=:column_key"), {'vehicle_id':vehicle_id,'column_key':column_key})


def _iso(d):
    return d.isoformat() if d else None


def _date_in_range(d, date_from=None, date_to=None):
    if not d:
        return False if (date_from or date_to) else True
    if date_from and d < date_from:
        return False
    if date_to and d > date_to:
        return False
    return True


def _summarize_vehicle_detail(v, loans, services, annual_taxes, five_year_taxes, kirs, vehicle_history, user_history, loan_history):
    total_service_cost = sum(int(float(x.get('cost') or 0)) for x in services if x.get('cost') not in (None, ''))
    total_tax_amount = sum(int(float(x.get('amount') or 0)) for x in annual_taxes + five_year_taxes if x.get('amount') not in (None, ''))
    total_loan = len(loans)
    unique_borrowers = sorted({(x.get('borrower') or '').strip() for x in loans if (x.get('borrower') or '').strip()})
    latest_service = services[0] if services else None
    latest_kir = kirs[0] if kirs else None
    latest_annual = annual_taxes[0] if annual_taxes else None
    latest_five = five_year_taxes[0] if five_year_taxes else None
    return {
        'vehicle_label': vehicle_display_name(v),
        'plate': vehicle_plate(v),
        'total_loans': total_loan,
        'unique_borrowers': unique_borrowers,
        'loan_people_count': len(unique_borrowers),
        'total_services': len(services),
        'service_total_cost': total_service_cost,
        'total_kirs': len(kirs),
        'total_annual_taxes': len(annual_taxes),
        'total_five_year_taxes': len(five_year_taxes),
        'tax_total_amount': total_tax_amount,
        'total_history_events': len(vehicle_history) + len(user_history) + len(loan_history),
        'latest_service_date': latest_service.get('service_date') if latest_service else None,
        'latest_kir_due_date': latest_kir.get('due_date') if latest_kir else None,
        'latest_annual_due_date': latest_annual.get('due_date') if latest_annual else None,
        'latest_five_due_date': latest_five.get('due_date') if latest_five else None,
    }


def _company_detail_payload(company:str, date_from=None, date_to=None):
    clean = _sanitize_company_filter((company or '').strip())
    vehicles = Vehicle.query.filter(Vehicle.is_deleted==False, func.lower(func.trim(Vehicle.pt))==clean.lower()).order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).all()
    vehicle_ids = [v.id for v in vehicles]
    custom_values = _vehicle_custom_values_map(vehicle_ids)
    vehicle_list = []
    for v in vehicles:
        vehicle_list.append({
            'id': v.id,
            'label': vehicle_display_name(v),
            'plate': vehicle_plate(v),
            'merk': v.merk or '',
            'type': v.type or '',
            'status': v.status or '',
            'user': v.user_new or v.user_old or '',
            'custom_values': custom_values.get(v.id, {})
        })

    loan_rows = LoanTransaction.query.filter(LoanTransaction.vehicle_id.in_(vehicle_ids)).order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).all() if vehicle_ids else []
    loan_place_lookup = _loan_place_map([r.id for r in loan_rows]) if loan_rows else {}
    service_rows = ServiceRecord.query.filter(ServiceRecord.vehicle_id.in_(vehicle_ids)).order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).all() if vehicle_ids else []
    kir_rows = KirRecord.query.filter(KirRecord.vehicle_id.in_(vehicle_ids)).order_by(KirRecord.due_date.desc(), KirRecord.id.desc()).all() if vehicle_ids else []
    change_rows = VehicleChangeHistory.query.filter(VehicleChangeHistory.vehicle_id.in_(vehicle_ids)).order_by(VehicleChangeHistory.changed_at.desc(), VehicleChangeHistory.id.desc()).limit(1000).all() if vehicle_ids else []
    # SQL Manual untuk UserHistory guna menghindari crash kolom pt_pemakai_lama/baru yang tidak sinkron
    user_hist_rows = []
    if vehicle_ids:
        try:
            sql_user = text("""
                SELECT id, vehicle_id, user_lama, user_baru, change_date, note 
                FROM user_histories 
                WHERE vehicle_id IN :ids 
                ORDER BY change_date DESC, id DESC 
                LIMIT 300
            """).bindparams(bindparam('ids', expanding=True))
            user_hist_rows = db.session.execute(sql_user, {'ids': vehicle_ids}).mappings().all()
        except Exception as e:
            print(f"DEBUG: CompanyDetail UserHistory Error: {e}")
            user_hist_rows = []

    loans = []
    for r in loan_rows:
        date_out = getattr(r, 'date_out', None) or getattr(r, 'loan_date_out', None) or getattr(r, 'tanggal_pinjam', None)
        date_plan = getattr(r, 'date_return_plan', None) or getattr(r, 'tanggal_kembali_rencana', None)
        date_actual = getattr(r, 'date_return_actual', None) or getattr(r, 'tanggal_kembali_aktual', None)
        loan_filter_date = date_out or date_plan or date_actual or getattr(r, 'created_at', None)
        if not _date_in_range(loan_filter_date, date_from, date_to):
            continue
        loans.append({
            'vehicle_id': r.vehicle_id,
            'vehicle_label': vehicle_display_name(getattr(r, 'vehicle', None)),
            'date_out': _iso(getattr(r, 'date_out', None)),
            'date_return_plan': _iso(getattr(r, 'date_return_plan', None)),
            'date_return_actual': _iso(getattr(r, 'date_return_actual', None)),
            'borrower': getattr(r, 'borrower_name', None) or getattr(r, 'borrower', None) or '',
            'borrower_company': getattr(r, 'borrower_company', None) or '',
            'place': loan_place_lookup.get(r.id, '') or '',
            'purpose': getattr(r, 'purpose', None) or '',
            'status': getattr(r, 'status', None) or '',
            'note': getattr(r, 'note', None) or ''
        })

    services = []
    for r in service_rows:
        svc_date = getattr(r, 'service_date', None)
        if not _date_in_range(svc_date, date_from, date_to):
            continue
        services.append({
            'vehicle_id': r.vehicle_id,
            'vehicle_label': vehicle_display_name(getattr(r, 'vehicle', None)),
            'service_date': _iso(svc_date),
            'service_type': getattr(r, 'service_type', None) or getattr(r, 'jenis_servis', None) or getattr(r, 'jenis_service', None) or '',
            'vendor': getattr(r, 'vendor', None) or getattr(r, 'bengkel', None) or '',
            'cost': getattr(r, 'cost', None) or getattr(r, 'biaya', None) or 0,
            'note': getattr(r, 'note', None) or ''
        })

    kirs = []
    for r in kir_rows:
        done_date = getattr(r, 'done_date', None) or getattr(r, 'kir_date', None)
        due_date = getattr(r, 'due_date', None)
        kir_filter_date = done_date or due_date
        if not _date_in_range(kir_filter_date, date_from, date_to):
            continue
        kirs.append({
            'vehicle_id': r.vehicle_id,
            'vehicle_label': vehicle_display_name(getattr(r, 'vehicle', None)),
            'done_date': _iso(getattr(r, 'done_date', None)),
            'due_date': _iso(getattr(r, 'due_date', None)),
            'result': getattr(r, 'result', None) or '',
            'note': getattr(r, 'note', None) or ''
        })

    annual_taxes = []
    five_year_taxes = []
    if vehicle_ids:
        try:
            annual_rows = db.session.execute(text("SELECT id, vehicle_id, tax_year, due_date, paid_date, amount, note FROM annual_tax_payments WHERE vehicle_id IN :ids ORDER BY due_date DESC, id DESC").bindparams(bindparam('ids', expanding=True)), {'ids': vehicle_ids}).mappings().all()
            for r in annual_rows:
                d = r['paid_date'] or r['due_date']
                if not _date_in_range(d, date_from, date_to):
                    continue
                annual_taxes.append({
                    'vehicle_id': r['vehicle_id'], 'tax_year': r['tax_year'], 'due_date': _iso(r['due_date']), 'paid_date': _iso(r['paid_date']), 'amount': r['amount'] or 0, 'note': r['note'] or ''
                })
        except Exception:
            pass
        try:
            five_rows = db.session.execute(text("SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note FROM five_year_tax_payments WHERE vehicle_id IN :ids ORDER BY due_date DESC, id DESC").bindparams(bindparam('ids', expanding=True)), {'ids': vehicle_ids}).mappings().all()
            for r in five_rows:
                d = r['paid_date'] or r['due_date']
                if not _date_in_range(d, date_from, date_to):
                    continue
                five_year_taxes.append({
                    'vehicle_id': r['vehicle_id'], 'due_date': _iso(r['due_date']), 'paid_date': _iso(r['paid_date']), 'amount': r['amount'] or 0, 'plate_before': r['plate_before'] or '', 'plate_after': r['plate_after'] or '', 'note': r['note'] or ''
                })
        except Exception:
            pass

    vehicle_history = []
    for r in change_rows:
        d = _history_timestamp_value(r, 'changed_at', 'change_date', 'created_at', 'updated_at')
        if not _date_in_range(d.date() if hasattr(d, 'date') else d, date_from, date_to):
            continue
        vehicle_history.append({
            'vehicle_id': r.vehicle_id,
            'vehicle_label': vehicle_display_name(getattr(r, 'vehicle', None)),
            'changed_at': _iso(d),
            'field_label': getattr(r, 'field_label', None) or getattr(r, 'field_name', None) or '',
            'old_value': getattr(r, 'old_value', None) or '',
            'new_value': getattr(r, 'new_value', None) or '',
            'changed_by': getattr(r, 'changed_by', None) or 'system',
            'note': getattr(r, 'note', None) or ''
        })

    user_history = []
    for r in user_hist_rows:
        d = _coerce_datetime_value(r.get('change_date')) or _coerce_datetime_value(r.get('changed_at')) or _coerce_datetime_value(r.get('created_at'))
        if not _date_in_range(d.date() if hasattr(d, 'date') else d, date_from, date_to):
            continue
        
        v_obj = Vehicle.query.get(r.get('vehicle_id'))
        user_history.append({
            'vehicle_id': r.get('vehicle_id'),
            'vehicle_label': vehicle_display_name(v_obj),
            'change_date': _iso(d),
            'user_lama': r.get('user_lama') or '',
            'user_baru': r.get('user_baru') or '',
            'pt_pemakai_lama': '', # Dikosongkan agar aman
            'pt_pemakai_baru': '', # Dikosongkan agar aman
            'changed_by': 'system',
            'note': r.get('note') or ''
        })

    total_service_cost = sum(int(float(x.get('cost') or 0)) for x in services)
    total_tax_amount = sum(int(float(x.get('amount') or 0)) for x in annual_taxes + five_year_taxes)
    unique_borrowers = sorted({(x.get('borrower') or '').strip() for x in loans if (x.get('borrower') or '').strip()})

    return {
        'company': clean,
        'date_from': _iso(date_from),
        'date_to': _iso(date_to),
        'vehicles': vehicle_list,
        'loans': loans,
        'services': services,
        'kirs': kirs,
        'annual_taxes': annual_taxes,
        'five_year_taxes': five_year_taxes,
        'vehicle_history': vehicle_history,
        'user_history': user_history,
        'summary': {
            'total_vehicles': len(vehicles),
            'total_loans': len(loans),
            'loan_people_count': len(unique_borrowers),
            'unique_borrowers': unique_borrowers,
            'total_services': len(services),
            'service_total_cost': total_service_cost,
            'total_kirs': len(kirs),
            'total_annual_taxes': len(annual_taxes),
            'total_five_year_taxes': len(five_year_taxes),
            'tax_total_amount': total_tax_amount,
            'total_history_events': len(vehicle_history) + len(user_history),
        }
    }

@bp.get("/kendaraan")
@master_required
def kendaraan():
    q=(request.args.get("q") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    due=(request.args.get("due") or "").strip()
    query=_vehicle_query_filtered(q,company)
    due_vehicle_ids=set()
    due_info_map={}
    if due=="1":
        for it in _reminder_items(q=q, company=company):
            vehicle_id = it["vehicle"].id
            due_vehicle_ids.add(vehicle_id)
            due_info_map.setdefault(vehicle_id, []).append({
                "label": it.get("label") or it.get("type", "Jatuh Tempo"),
                "due_display": it.get("due_display") or format_tgl_id_full(it.get("due")),
                "sisa_text": it.get("sisa_text") or _alert_sisa_text(it.get("days_left")),
                "sisa_class": it.get("sisa_class") or _alert_sisa_class(it.get("days_left")),
                "status": it.get("status") or "empty",
            })
        query=query.filter(Vehicle.id.in_(list(due_vehicle_ids))) if due_vehicle_ids else query.filter(Vehicle.id==-1)
    vehicles=query.order_by(Vehicle.updated_at.desc(),Vehicle.id.desc()).limit(300).all()
    companies = _scoped_company_choices()
    custom_columns=_vehicle_custom_columns()
    custom_values_map=_vehicle_custom_values_map([v.id for v in vehicles])
    
    # Tambahan: Kirim metadata kolom ke template agar modal export otomatis
    return render_template("kendaraan.html",
        vehicles=vehicles,
        companies=companies,
        q=q,
        company=company,
        due=due,
        due_vehicle_ids=due_vehicle_ids,
        due_info_map=due_info_map,
        title="Data Kendaraan - Sinar Group",
        custom_columns=custom_columns,
        custom_column_keys=[c['column_key'] for c in custom_columns],
        custom_values_map=custom_values_map,
        column_defs=_export_column_definitions(custom_columns),
        kendaraan_row_create_url=url_for('main.kendaraan_tambah'),
        kendaraan_column_create_url=url_for('main.kendaraan_column_create'),
        kendaraan_column_delete_url_template=url_for('main.kendaraan_column_delete', column_key='__COLUMN_KEY__'),
        allowed_pts=companies,
        is_master=_is_master_session(),
    )

_loan_support_tables_ensured = False

def _ensure_loan_support_tables():
    global _loan_support_tables_ensured
    if _loan_support_tables_ensured:
        return
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS loan_transaction_meta (
            loan_id INT NOT NULL,
            place VARCHAR(255) NULL,
            updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (loan_id),
            CONSTRAINT fk_loan_transaction_meta_loan
                FOREIGN KEY (loan_id) REFERENCES loan_transactions (id)
                ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """))
    db.session.commit()
    _loan_support_tables_ensured = True

def _loan_place_map(loan_ids:list[int]):
    _ensure_loan_support_tables()
    result = {}
    if not loan_ids:
        return result
    rows = db.session.execute(
        text("SELECT loan_id, place FROM loan_transaction_meta WHERE loan_id IN :ids").bindparams(bindparam('ids', expanding=True)),
        {'ids': loan_ids}
    ).mappings().all()
    for r in rows:
        result[r['loan_id']] = _normalize_text(r.get('place'))
    return result

def _loan_place_value(loan_id:int)->str:
    return _loan_place_map([loan_id]).get(loan_id, '')

def _set_loan_place(loan_id:int, place):
    _ensure_loan_support_tables()
    clean_place = _normalize_text(place) or None
    exists = db.session.execute(text("SELECT loan_id FROM loan_transaction_meta WHERE loan_id=:loan_id"), {'loan_id': loan_id}).first()
    if exists:
        db.session.execute(text("UPDATE loan_transaction_meta SET place=:place, updated_at=CURRENT_TIMESTAMP WHERE loan_id=:loan_id"), {'loan_id': loan_id, 'place': clean_place})
    else:
        db.session.execute(text("INSERT INTO loan_transaction_meta (loan_id, place) VALUES (:loan_id, :place)"), {'loan_id': loan_id, 'place': clean_place})

def _base_row_data(idx: int, v) -> dict:
    """Mengambil semua nilai kolom dasar dari sebuah Vehicle untuk export Excel."""
    # KIR latest
    kir = _latest_kir_map([v.id]).get(v.id)
    kir_done = getattr(kir, "done_date", None) if kir else None
    kir_due = getattr(kir, "due_date", None) if kir else None
    kir_days = days_left(kir_due) if kir_due else None
    if kir_due and kir_days is not None:
        if kir_days < 0: kir_status = "Kadaluarsa"
        elif kir_days <= 14: kir_status = "Segera"
        else: kir_status = "Aktif"
    else:
        kir_status = "-"

    # Servis latest
    svc_rows = ServiceRecord.query.filter_by(vehicle_id=v.id).order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).first()
    svc_date = getattr(svc_rows, "service_date", None) if svc_rows else None
    svc_due = service_due_date(svc_rows) if svc_rows else None
    svc_days = days_left(svc_due) if svc_due else None
    if svc_due and svc_days is not None:
        if svc_days < 0: svc_status = "Kadaluarsa"
        elif svc_days <= 14: svc_status = "Segera"
        else: svc_status = "Aktif"
    else:
        svc_status = "-"

    # Peminjaman latest
    loan = LoanTransaction.query.filter_by(vehicle_id=v.id).order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).first()
    loan_place_map = _loan_place_map([loan.id]) if loan else {}
    loan_status = _loan_status_label(loan) if loan else "-"

    # Pajak tahunan latest
    annual_map = _annual_payment_map([v.id])
    latest_annual = None
    for (vid, yr), row in annual_map.items():
        if vid == v.id:
            if latest_annual is None or yr > latest_annual[0]:
                latest_annual = (yr, row)
    annual_row = latest_annual[1] if latest_annual else None

    # Pajak 5 tahunan latest
    five = _latest_five_tax_map([v.id]).get(v.id)

    # Asset owner company
    asset_owner = None
    try:
        if getattr(v, "asset_owner_company_id", None):
            co = Company.query.get(v.asset_owner_company_id)
            asset_owner = getattr(co, "name", None) if co else None
    except Exception:
        pass

    # PT Pemakai
    pt_pemakai = None
    try:
        if getattr(v, "pt_pemakai_company_id", None):
            co = Company.query.get(v.pt_pemakai_company_id)
            pt_pemakai = getattr(co, "name", None) if co else None
    except Exception:
        pass
    if not pt_pemakai:
        pt_pemakai = getattr(v, "pt_pemakai", None) or getattr(v, "pt_pemakai_1", None)

    return {
        "no": idx,
        "pt": _export_scalar(getattr(v, "pt", None)),
        "asset_owner_company_name": _export_scalar(asset_owner or getattr(v, "asset_owner_company_name", None)),
        "active_name": _export_scalar(getattr(v, "active_name", None)),
        "name_as_asset_pt": _export_scalar(getattr(v, "name_as_asset_pt", None)),
        "new_asset_name": _export_scalar(getattr(v, "new_asset_name", None)),
        "merk": _export_scalar(getattr(v, "merk", None)),
        "type": _export_scalar(getattr(v, "type", None)),
        "jenis": _export_scalar(getattr(v, "jenis", None)),
        "plate_old": _export_scalar(getattr(v, "plate_old", None)),
        "plate_new": _export_scalar(getattr(v, "plate_new", None)),
        "year_of_use": _export_scalar(getattr(v, "year_of_use", None)),
        "user_old": _export_scalar(getattr(v, "user_old", None)),
        "user_new": _export_scalar(getattr(v, "user_new", None)),
        "status": _export_scalar(getattr(v, "status", None)),
        "pt_pemakai_company_name": _export_scalar(pt_pemakai),
        "kondisi_terkini": _export_scalar(getattr(v, "kondisi_terkini", None)),
        "lokasi": _export_scalar(getattr(v, "lokasi", None)),
        "tambahan_keterangan": _export_scalar(getattr(v, "tambahan_keterangan", None)),
        # KIR
        "kir_done_date": _export_date_value(kir_done),
        "kir_due_date": _export_date_value(kir_due),
        "kir_status": kir_status,
        "kir_result": _export_scalar(getattr(kir, "result", None) if kir else None),
        "kir_note": _export_scalar(getattr(kir, "note", None) if kir else None),
        # Servis
        "service_date": _export_date_value(svc_date),
        "service_next_due": _export_date_value(svc_due),
        "service_status": svc_status,
        "service_vendor": _export_scalar(getattr(svc_rows, "vendor", None) if svc_rows else None),
        "service_type": _export_scalar(getattr(svc_rows, "service_type", None) or getattr(svc_rows, "jenis_servis", None) if svc_rows else None),
        "service_odometer": _export_scalar(getattr(svc_rows, "odometer", None) if svc_rows else None),
        "service_cost": _export_scalar(getattr(svc_rows, "cost", None) if svc_rows else None),
        "service_note": _export_scalar(getattr(svc_rows, "note", None) if svc_rows else None),
        # Peminjaman
        "loan_borrower_name": _export_scalar(getattr(loan, "borrower_name", None) if loan else None),
        "loan_borrower_company": _export_scalar(getattr(loan, "borrower_company", None) if loan else None),

        "loan_date_out": _export_date_value(getattr(loan, "date_out", None) if loan else None),
        "loan_date_return_plan": _export_date_value(getattr(loan, "date_return_plan", None) if loan else None),
        "loan_date_return_actual": _export_date_value(getattr(loan, "date_return_actual", None) if loan else None),
        "loan_status": loan_status,
        "loan_note": _export_scalar(getattr(loan, "note", None) if loan else None),
        # Pajak Tahunan
        "tax_annual_due_date": _export_date_value(annual_row["due_date"] if annual_row else None),
        "tax_annual_paid_date": _export_date_value(annual_row["paid_date"] if annual_row else None),
        "tax_annual_status": ("Lunas" if (annual_row and annual_row.get("paid_date")) else ("Belum Bayar" if annual_row else "-")),
        "tax_annual_amount": _export_scalar(annual_row["amount"] if annual_row else None),
        # Pajak 5 Tahunan
        "tax_five_due_date": _export_date_value(five["due_date"] if five else None),
        "tax_five_paid_date": _export_date_value(five["paid_date"] if five else None),
        "tax_five_status": ("Lunas" if (five and five.get("paid_date")) else ("Belum Bayar" if five else "-")),
        "tax_five_amount": _export_scalar(five["amount"] if five else None),
        "tax_plate_before": _export_scalar(five["plate_before"] if five else None),
        "tax_plate_after": _export_scalar(five["plate_after"] if five else None),
        "tax_note": _export_scalar(five["note"] if five else None),
    }


def _service_history_value(r, col: str) -> str:
    """Ambil nilai kolom riwayat servis dari satu ServiceRecord."""
    if r is None:
        return "-"
    mapping = {
        "service_history_date": lambda: _export_date_value(getattr(r, "service_date", None)),
        "service_history_item": lambda: _export_scalar(getattr(r, "service_type", None) or getattr(r, "jenis_servis", None) or getattr(r, "item", None)),
        "service_history_odometer": lambda: _export_scalar(getattr(r, "odometer", None)),
        "service_history_cost": lambda: _export_scalar(getattr(r, "cost", None)),
        "service_history_note": lambda: _export_scalar(getattr(r, "note", None)),
    }
    return mapping.get(col, lambda: "-")()


def _kir_history_value(r, col: str) -> str:
    """Ambil nilai kolom riwayat KIR dari satu KirRecord."""
    if r is None:
        return "-"
    mapping = {
        "kir_history_done_date": lambda: _export_date_value(getattr(r, "done_date", None)),
        "kir_history_due_date": lambda: _export_date_value(getattr(r, "due_date", None)),
        "kir_history_result": lambda: _export_scalar(getattr(r, "result", None)),
        "kir_history_note": lambda: _export_scalar(getattr(r, "note", None)),
    }
    return mapping.get(col, lambda: "-")()


def _loan_history_value(r, col: str) -> str:
    """Ambil nilai kolom riwayat peminjaman dari satu LoanTransaction."""
    if r is None:
        return "-"
    loan_place_map = _loan_place_map([r.id]) if r else {}
    mapping = {
        "loan_history_date_out": lambda: _export_date_value(getattr(r, "date_out", None)),
        "loan_history_date_return_plan": lambda: _export_date_value(getattr(r, "date_return_plan", None)),
        "loan_history_date_return_actual": lambda: _export_date_value(getattr(r, "date_return_actual", None)),
        "loan_history_borrower": lambda: _export_scalar(getattr(r, "borrower_name", None)),

        "loan_history_note": lambda: _export_scalar(getattr(r, "note", None)),
    }
    return mapping.get(col, lambda: "-")()




def _vehicle_change_history_value(r, col: str) -> str:
    """Ambil nilai kolom Riwayat Kendaraan (snapshot lengkap per sesi edit)."""
    if r is None:
        return "-"

    def get_value(name, default=None):
        try:
            if hasattr(r, "get"):
                return r.get(name, default)
        except Exception:
            pass
        return getattr(r, name, default)

    snapshot = get_value("snapshot", {}) or {}
    dt = (
        _coerce_datetime_value(get_value("changed_at"))
        or _coerce_datetime_value(get_value("change_date"))
        or _coerce_datetime_value(get_value("created_at"))
        or _coerce_datetime_value(get_value("updated_at"))
    )

    def snap(key):
        return _export_scalar(snapshot.get(key))

    mapping = {
        "vehicle_history_no": lambda: _export_scalar(get_value("no")),
        "vehicle_history_changed_at": lambda: format_tgl_jam_id_full(dt) if dt else "-",
        "vehicle_history_changed_by": lambda: _export_scalar(get_value("changed_by")),
        "vehicle_history_pt": lambda: snap("pt"),
        "vehicle_history_asset_owner": lambda: snap("asset_owner_company_name"),
        "vehicle_history_active_name": lambda: snap("active_name"),
        "vehicle_history_name_as_asset_pt": lambda: snap("name_as_asset_pt"),
        "vehicle_history_new_asset_name": lambda: snap("new_asset_name"),
        "vehicle_history_merk": lambda: snap("merk"),
        "vehicle_history_type": lambda: snap("type"),
        "vehicle_history_jenis": lambda: snap("jenis"),
        "vehicle_history_plate_old": lambda: snap("plate_old"),
        "vehicle_history_plate_new": lambda: snap("plate_new"),
        "vehicle_history_year_of_use": lambda: snap("year_of_use"),
        "vehicle_history_user_old": lambda: snap("user_old"),
        "vehicle_history_user_new": lambda: snap("user_new"),
        "vehicle_history_status": lambda: snap("status"),
        "vehicle_history_pt_pemakai": lambda: snap("pt_pemakai_company_name"),
        "vehicle_history_kondisi": lambda: snap("kondisi_terkini"),
        "vehicle_history_lokasi": lambda: snap("lokasi"),
        "vehicle_history_keterangan": lambda: snap("tambahan_keterangan"),
        "vehicle_history_field": lambda: _export_scalar(get_value("field_label") or get_value("field_name")),
        "vehicle_history_old_value": lambda: _export_scalar(get_value("old_value")),
        "vehicle_history_new_value": lambda: _export_scalar(get_value("new_value")),
        "vehicle_history_note": lambda: _export_scalar(get_value("note")),
    }
    return mapping.get(col, lambda: "-")()


def _vehicle_history_sessions_for_export(vehicle_ids: list[int]) -> dict[int, list[dict]]:
    """Kelompokkan VehicleChangeHistory menjadi sesi Riwayat Kendaraan dengan snapshot lengkap."""
    result = {}
    if not vehicle_ids:
        return result
    rows = (
        VehicleChangeHistory.query
        .filter(VehicleChangeHistory.vehicle_id.in_(vehicle_ids))
        .order_by(VehicleChangeHistory.changed_at.desc(), VehicleChangeHistory.id.desc())
        .all()
    )
    if not rows:
        return result
    vehicle_map = {v.id: v for v in Vehicle.query.filter(Vehicle.id.in_(vehicle_ids)).all()}
    grouped = {}
    for r in rows:
        ts = r.changed_at.strftime('%Y-%m-%d %H:%M:%S') if getattr(r, 'changed_at', None) else 'unknown'
        key = (r.vehicle_id, ts, r.changed_by or 'system')
        item = grouped.setdefault(key, {'vehicle_id': r.vehicle_id, 'changed_at': getattr(r, 'changed_at', None), 'changed_by': r.changed_by or 'system', 'changed_fields': {}})
        item['changed_fields'][r.field_name or 'unknown'] = {'old': r.old_value or '', 'new': r.new_value or ''}
    sessions_by_vehicle = {}
    for key in sorted(grouped.keys(), key=lambda x: (x[1], x[0]), reverse=True):
        session_item = grouped[key]
        v = vehicle_map.get(session_item['vehicle_id'])
        if not v:
            continue
        snapshot = _vehicle_snapshot(v, include_custom=True)
        changed_at = session_item.get('changed_at')
        if changed_at:
            later_changes = [row for row in rows if row.vehicle_id == v.id and getattr(row, 'changed_at', None) and row.changed_at > changed_at]
            for lc in sorted(later_changes, key=lambda x: x.changed_at, reverse=True):
                if lc.field_name in snapshot:
                    snapshot[lc.field_name] = lc.old_value
        session_item['snapshot'] = snapshot
        sessions_by_vehicle.setdefault(v.id, []).append(session_item)
    for vid, items in sessions_by_vehicle.items():
        for idx, item in enumerate(items, 1):
            item['no'] = idx
        result[vid] = items
    return result


def _user_history_value(r, col: str) -> str:
    """Ambil nilai kolom riwayat user kendaraan. Aman untuk ORM object maupun row mapping SQL."""
    if r is None:
        return "-"
    def get_value(name, default=None):
        try:
            if hasattr(r, "get"):
                return r.get(name, default)
        except Exception:
            pass
        return getattr(r, name, default)
    dt = (_coerce_datetime_value(get_value("change_date")) or _coerce_datetime_value(get_value("changed_at")) or _coerce_datetime_value(get_value("created_at")) or _coerce_datetime_value(get_value("updated_at")))
    mapping = {
        "user_history_changed_at": lambda: format_tgl_jam_id_full(dt) if dt else "-",
        "user_history_user_old": lambda: _export_scalar(get_value("user_lama")),
        "user_history_user_new": lambda: _export_scalar(get_value("user_baru")),
        "user_history_pt_old": lambda: _export_scalar(get_value("pt_pemakai_lama")),
        "user_history_pt_new": lambda: _export_scalar(get_value("pt_pemakai_baru")),
        "user_history_note": lambda: _export_scalar(get_value("note")),
    }
    return mapping.get(col, lambda: "-")()

def _export_column_definitions(custom_cols=None):
    # Selalu ambil kolom custom terbaru supaya modal/export Excel ikut berubah
    # ketika user menambah kolom di Data Kendaraan.
    if custom_cols is None:
        custom_cols = _vehicle_custom_columns()

    defs = {
        "no": {"label": "No", "group": "Data Kendaraan", "section": "data"},
        "pt": {"label": "PT", "group": "Data Kendaraan", "section": "data"},
        "asset_owner_company_name": {"label": "PT Pemilik Aset", "group": "Data Kendaraan", "section": "data"},
        "active_name": {"label": "Nama Aktiva", "group": "Data Kendaraan", "section": "data"},
        "name_as_asset_pt": {"label": "Nama Asset PT", "group": "Data Kendaraan", "section": "data"},
        "new_asset_name": {"label": "Nama Aset Baru", "group": "Data Kendaraan", "section": "data"},
        "merk": {"label": "Merk", "group": "Data Kendaraan", "section": "data"},
        "type": {"label": "Type", "group": "Data Kendaraan", "section": "data"},
        "jenis": {"label": "Jenis", "group": "Data Kendaraan", "section": "data"},
        "plate_old": {"label": "No Polisi Lama", "group": "Data Kendaraan", "section": "data"},
        "plate_new": {"label": "No Polisi Baru", "group": "Data Kendaraan", "section": "data"},
        "year_of_use": {"label": "Tahun Pemakaian", "group": "Data Kendaraan", "section": "data"},
        "user_old": {"label": "User Lama", "group": "Data Kendaraan", "section": "data"},
        "user_new": {"label": "User Baru", "group": "Data Kendaraan", "section": "data"},
        "status": {"label": "Status", "group": "Data Kendaraan", "section": "data"},
        "pt_pemakai_company_name": {"label": "PT Pemakai", "group": "Data Kendaraan", "section": "data"},
        "kondisi_terkini": {"label": "Kondisi Terkini", "group": "Data Kendaraan", "section": "data"},
        "lokasi": {"label": "Lokasi", "group": "Data Kendaraan", "section": "data"},
        "tambahan_keterangan": {"label": "Tambahan Keterangan", "group": "Data Kendaraan", "section": "data"},

        # Data Terakhir (Summary)
        "kir_done_date": {"label": "Tanggal KIR Terakhir", "group": "Data KIR Terakhir", "section": "data"},
        "kir_due_date": {"label": "Jatuh Tempo KIR", "group": "Data KIR Terakhir", "section": "data"},
        "kir_status": {"label": "Status KIR", "group": "Data KIR Terakhir", "section": "data"},
        "kir_result": {"label": "Hasil KIR", "group": "Data KIR Terakhir", "section": "data"},
        "kir_note": {"label": "Catatan KIR", "group": "Data KIR Terakhir", "section": "data"},

        "service_date": {"label": "Tanggal Servis Terakhir", "group": "Data Servis Terakhir", "section": "data"},
        "service_next_due": {"label": "Jatuh Tempo Servis", "group": "Data Servis Terakhir", "section": "data"},
        "service_status": {"label": "Status Servis", "group": "Data Servis Terakhir", "section": "data"},
        "service_vendor": {"label": "Vendor Servis", "group": "Data Servis Terakhir", "section": "data"},
        "service_type": {"label": "Jenis Servis", "group": "Data Servis Terakhir", "section": "data"},
        "service_odometer": {"label": "Odometer Servis", "group": "Data Servis Terakhir", "section": "data"},
        "service_cost": {"label": "Biaya Servis", "group": "Data Servis Terakhir", "section": "data"},
        "service_note": {"label": "Catatan Servis", "group": "Data Servis Terakhir", "section": "data"},

        "loan_borrower_name": {"label": "Peminjam", "group": "Data Peminjaman Terakhir", "section": "data"},
        "loan_borrower_company": {"label": "Perusahaan Peminjam", "group": "Data Peminjaman Terakhir", "section": "data"},

        "loan_date_out": {"label": "Tanggal Pinjam", "group": "Data Peminjaman Terakhir", "section": "data"},
        "loan_date_return_plan": {"label": "Rencana Kembali", "group": "Data Peminjaman Terakhir", "section": "data"},
        "loan_date_return_actual": {"label": "Aktual Kembali", "group": "Data Peminjaman Terakhir", "section": "data"},
        "loan_status": {"label": "Status Peminjaman", "group": "Data Peminjaman Terakhir", "section": "data"},
        "loan_note": {"label": "Catatan Peminjaman", "group": "Data Peminjaman Terakhir", "section": "data"},

        # Riwayat (Multi-row)
        "service_history_date": {"label": "Riwayat Servis - Tanggal", "group": "Riwayat Servis", "section": "service_history"},
        "service_history_item": {"label": "Riwayat Servis - Rincian Item", "group": "Riwayat Servis", "section": "service_history"},
        "service_history_odometer": {"label": "Riwayat Servis - KM", "group": "Riwayat Servis", "section": "service_history"},
        "service_history_cost": {"label": "Riwayat Servis - Biaya", "group": "Riwayat Servis", "section": "service_history"},
        "service_history_note": {"label": "Riwayat Servis - Catatan", "group": "Riwayat Servis", "section": "service_history"},

        "kir_history_done_date": {"label": "Riwayat KIR - Tanggal KIR", "group": "Riwayat KIR", "section": "kir_history"},
        "kir_history_due_date": {"label": "Riwayat KIR - Jatuh Tempo", "group": "Riwayat KIR", "section": "kir_history"},
        "kir_history_result": {"label": "Riwayat KIR - Hasil", "group": "Riwayat KIR", "section": "kir_history"},
        "kir_history_note": {"label": "Riwayat KIR - Catatan", "group": "Riwayat KIR", "section": "kir_history"},

        "loan_history_date_out": {"label": "Riwayat Peminjaman - Tanggal Pinjam", "group": "Riwayat Peminjaman", "section": "loan_history"},
        "loan_history_date_return_plan": {"label": "Riwayat Peminjaman - Rencana Kembali", "group": "Riwayat Peminjaman", "section": "loan_history"},
        "loan_history_date_return_actual": {"label": "Riwayat Peminjaman - Aktual Kembali", "group": "Riwayat Peminjaman", "section": "loan_history"},
        "loan_history_borrower": {"label": "Riwayat Peminjaman - Peminjam", "group": "Riwayat Peminjaman", "section": "loan_history"},

        "loan_history_note": {"label": "Riwayat Peminjaman - Catatan", "group": "Riwayat Peminjaman", "section": "loan_history"},

        # Riwayat Kendaraan = snapshot lengkap kendaraan per sesi perubahan
        "vehicle_history_no": {"label": "No", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_changed_at": {"label": "Tanggal", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_changed_by": {"label": "Diubah Oleh", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_pt": {"label": "PT", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_asset_owner": {"label": "PT Pemilik Aset", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_active_name": {"label": "Nama Aktiva", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_name_as_asset_pt": {"label": "Nama Asset PT", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_new_asset_name": {"label": "Nama Aset Baru", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_merk": {"label": "Merk", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_type": {"label": "Type", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_jenis": {"label": "Jenis", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_plate_old": {"label": "No Polisi Lama", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_plate_new": {"label": "No Polisi Baru", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_year_of_use": {"label": "Tahun", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_user_old": {"label": "User Lama", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_user_new": {"label": "User Baru", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_status": {"label": "Status", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_pt_pemakai": {"label": "PT Pemakai", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_kondisi": {"label": "Kondisi", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_lokasi": {"label": "Lokasi", "group": "Riwayat Kendaraan", "section": "vehicle_history"},
        "vehicle_history_keterangan": {"label": "Keterangan", "group": "Riwayat Kendaraan", "section": "vehicle_history"},

        # Riwayat User tetap dipisah dari Riwayat Kendaraan
        "user_history_changed_at": {"label": "Riwayat User - Tanggal", "group": "Riwayat User", "section": "user_history"},
        "user_history_user_old": {"label": "Riwayat User - User Lama", "group": "Riwayat User", "section": "user_history"},
        "user_history_user_new": {"label": "Riwayat User - User Baru", "group": "Riwayat User", "section": "user_history"},
        "user_history_pt_old": {"label": "Riwayat User - PT Pemakai Lama", "group": "Riwayat User", "section": "user_history"},
        "user_history_pt_new": {"label": "Riwayat User - PT Pemakai Baru", "group": "Riwayat User", "section": "user_history"},
        "user_history_note": {"label": "Riwayat User - Catatan", "group": "Riwayat User", "section": "user_history"},

        # Pajak
        "tax_annual_due_date": {"label": "Exp Pajak Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_annual_paid_date": {"label": "Tanggal Bayar Pajak Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_annual_status": {"label": "Status Pajak Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_annual_amount": {"label": "Nominal Pajak Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_five_due_date": {"label": "Exp Pajak 5 Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_five_paid_date": {"label": "Tanggal Bayar Pajak 5 Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_five_status": {"label": "Status Pajak 5 Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_five_amount": {"label": "Nominal Pajak 5 Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_plate_before": {"label": "Plat Sebelum Pajak 5 Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_plate_after": {"label": "Plat Sesudah Pajak 5 Tahunan", "group": "Data Pajak", "section": "data"},
        "tax_note": {"label": "Catatan Pajak", "group": "Data Pajak", "section": "data"},
    }
    if custom_cols:
        custom_defs = {
            col['column_key']: {"label": col['column_label'], "group": "Data Kendaraan", "section": "data"}
            for col in custom_cols
        }
        if custom_defs:
            # Letakkan kolom custom tepat setelah kolom Tambahan Keterangan
            # agar urutannya mengikuti tabel Data Kendaraan dan bukan terselip
            # di bawah grup lain seperti KIR/Servis/Pajak.
            ordered_defs = {}
            inserted_custom = False
            for key, value in defs.items():
                ordered_defs[key] = value
                if key == "tambahan_keterangan":
                    ordered_defs.update(custom_defs)
                    inserted_custom = True
            if not inserted_custom:
                ordered_defs.update(custom_defs)
            defs = ordered_defs
    return defs

def _export_date_value(value):
    return format_tgl_id_full(value) if value else "-"


def _export_scalar(value, fallback="-"):
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def _autosize_worksheet_columns(ws, min_width=12, max_width=40):
    # Scan header (baris 1-2) + maks 50 baris sampel merata, tidak scan semua row.
    # Scan penuh bisa memakan beberapa detik untuk sheet besar karena openpyxl
    # mengakses tiap cell satu per satu via Python loop.
    total_rows = ws.max_row
    sample_rows = list(range(1, min(3, total_rows + 1)))
    if total_rows > 2:
        step = max(1, (total_rows - 2) // 48)
        sample_rows += list(range(3, total_rows + 1, step))[:48]
    sample_set = set(sample_rows)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in sample_set:
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                l = len(str(value))
                if l > max_length:
                    max_length = l
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)

def _loan_status_label(row):
    status_text = (_normalize_text(getattr(row, "status", None)) or "").lower()
    if status_text == "selesai" or getattr(row, "date_return_actual", None):
        return "Selesai"
    plan = getattr(row, "date_return_plan", None)
    if plan and plan < date.today():
        return "Terlambat"
    if status_text == "dipinjam":
        return "Dipinjam"
    return _export_scalar(getattr(row, "status", None))


@bp.get("/api/peminjaman/export/<int:vehicle_id>")
def api_peminjaman_export(vehicle_id: int):
    v = Vehicle.query.get_or_404(vehicle_id)

    export_type = (_normalize_text(request.args.get("export_type")) or "all").lower()
    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))

    if date_from and date_to and date_from > date_to:
        return jsonify({"success": False, "message": "Tanggal awal tidak boleh lebih besar dari tanggal akhir."}), 400

    query = (
        LoanTransaction.query
        .filter(LoanTransaction.vehicle_id == vehicle_id)
        .order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc())
    )
    loan_rows = query.all()

    filtered_rows = []
    for row in loan_rows:
        row_date = getattr(row, "date_out", None) or getattr(row, "date_return_plan", None) or getattr(row, "date_return_actual", None)
        if not _date_in_range(row_date, date_from, date_to):
            continue
        filtered_rows.append(row)

    loan_place_lookup = _loan_place_map([r.id for r in (filtered_rows or loan_rows)]) if (filtered_rows or loan_rows) else {}
    latest = filtered_rows[0] if filtered_rows else (loan_rows[0] if loan_rows else None)
    total_selesai = sum(1 for r in filtered_rows if (_normalize_text(getattr(r, "status", None)).lower() == "selesai" or getattr(r, "date_return_actual", None)))
    total_aktif = sum(1 for r in filtered_rows if not (_normalize_text(getattr(r, "status", None)).lower() == "selesai" or getattr(r, "date_return_actual", None)))

    vehicle_name = vehicle_display_name(v)
    no_polisi = vehicle_plate(v)
    pt_name = getattr(v, "pt", None) or "-"
    _nama_utama = (
        getattr(v, "active_name", None)
        or getattr(v, "new_asset_name", None)
        or getattr(v, "name_as_asset_pt", None)
        or vehicle_name
    )

    summary_data = [
        ("Nama Aset", _nama_utama or "-"),
        ("No Polisi", no_polisi or "-"),
        ("PT", pt_name),
        ("Merk", _export_scalar(getattr(v, "merk", None))),
        ("Tipe", _export_scalar(getattr(v, "type", None))),
        ("Jenis", _export_scalar(getattr(v, "jenis", None))),
        ("Tahun Pemakaian", _export_scalar(getattr(v, "year_of_use", None))),
        ("Tanggal Awal Filter", format_tgl_id_full(date_from) if date_from else "Semua"),
        ("Tanggal Akhir Filter", format_tgl_id_full(date_to) if date_to else "Semua"),
        ("Total Riwayat Peminjaman", len(filtered_rows)),
        ("Peminjaman Aktif", total_aktif),
        ("Peminjaman Selesai", total_selesai),
        ("Tanggal Pinjam Terakhir", _export_date_value(getattr(latest, "date_out", None) if latest else None)),
        ("Rencana Kembali Terakhir", _export_date_value(getattr(latest, "date_return_plan", None) if latest else None)),
        ("Aktual Kembali Terakhir", _export_date_value(getattr(latest, "date_return_actual", None) if latest else None)),
        ("Dari PT Terakhir", _export_scalar(getattr(latest, "borrower_company", None) if latest else None)),
        ("Ke PT Terakhir", _export_scalar(getattr(latest, "borrower_name", None) if latest else None)),

        ("Status Terakhir", _loan_status_label(latest) if latest else "-"),
        ("Catatan Terakhir", _export_scalar(getattr(latest, "note", None) if latest else None)),
    ]
    df_summary = pd.DataFrame(summary_data, columns=["Informasi", "Nilai"])

    history_rows = []
    for idx, row in enumerate(filtered_rows, 1):
        history_rows.append({
            "No": idx,
            "PT": pt_name,
            "Kendaraan": vehicle_name,
            "No Polisi": no_polisi,
            "Dari PT": _export_scalar(getattr(row, "borrower_company", None)),
            "Ke PT": _export_scalar(getattr(row, "borrower_name", None)),
            "Tanggal Pinjam": _export_date_value(getattr(row, "date_out", None)),
            "Rencana Kembali": _export_date_value(getattr(row, "date_return_plan", None)),
            "Aktual Kembali": _export_date_value(getattr(row, "date_return_actual", None)),
            "Status": _loan_status_label(row),
            "Catatan": _export_scalar(getattr(row, "note", None)),
        })

    if not history_rows:
        history_rows = [{
            "No": "-",
            "PT": pt_name,
            "Kendaraan": vehicle_name,
            "No Polisi": no_polisi,
            "Dari PT": "-",
            "Ke PT": "-",
            "Tempat": "-",
            "Tanggal Pinjam": "-",
            "Rencana Kembali": "-",
            "Aktual Kembali": "-",
            "Tujuan": "-",
            "Status": "-",
            "Catatan": "-",
        }]
    df_history = pd.DataFrame(history_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if export_type in {"all", "summary"}:
            df_summary.to_excel(writer, index=False, sheet_name="Ringkasan Peminjaman")
        if export_type in {"all", "history"}:
            df_history.to_excel(writer, index=False, sheet_name="Riwayat Peminjaman")

        for ws in writer.book.worksheets:
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True, color="FFD15F")
                    cell.fill = cell.fill.copy(fill_type="solid", fgColor="1E4080")
            _autosize_worksheet_columns(ws, min_width=14, max_width=38)

    output.seek(0)
    safe_name = _safe_export_filename_part(vehicle_name, "KENDARAAN")
    safe_plate = _safe_export_filename_part(no_polisi, "-")
    filename = f"Laporan_Peminjaman_{safe_name.replace(' ', '_')}_{safe_plate.replace(' ', '_')}_{date.today().strftime('%d-%m-%Y')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.post("/kendaraan/export")
@master_required
def kendaraan_export():
    try:
        data = request.get_json(silent=True) or {}
        selected_pt = data.get("pt") or []
        selected_columns = data.get("columns") or []
        date_from_str, date_to_str = data.get("date_from"), data.get("date_to")
        d_from, d_to = _parse_date(date_from_str), _parse_date(date_to_str)

        if d_from and d_to and d_from > d_to:
            return jsonify({"success": False, "message": "Tanggal awal tidak boleh lebih besar dari tanggal akhir."}), 400

        custom_cols = _vehicle_custom_columns()
        column_defs = _export_column_definitions(custom_cols)

        # Pastikan kolom tambahan yang dibuat user di Data Kendaraan ikut masuk
        # ke export Excel. Ini juga membuat export tetap aman jika halaman/modal
        # belum refresh setelah user menambah kolom baru.
        custom_column_keys = [col['column_key'] for col in custom_cols]
        data_columns_selected = any(
            col in column_defs and column_defs[col].get("section") == "data"
            for col in selected_columns
        )
        if data_columns_selected:
            for custom_key in custom_column_keys:
                if custom_key not in selected_columns:
                    selected_columns.append(custom_key)

        valid_columns = [col for col in selected_columns if col in column_defs]
        if not valid_columns:
            return jsonify({"success": False, "message": "Pilih minimal 1 kolom export."}), 400

        include_history = bool(data.get("include_history"))
        svc_hist_cols = [col for col in valid_columns if column_defs[col].get("section") == "service_history"]
        kir_hist_cols = [col for col in valid_columns if column_defs[col].get("section") == "kir_history"]
        loan_hist_cols = [col for col in valid_columns if column_defs[col].get("section") == "loan_history"]
        vehicle_hist_cols = [col for col in valid_columns if column_defs[col].get("section") == "vehicle_history"]
        user_hist_cols = [col for col in valid_columns if column_defs[col].get("section") == "user_history"]
        if include_history:
            # Riwayat Kendaraan selalu lengkap: snapshot kendaraan, bukan field/nilai lama/nilai baru.
            vehicle_hist_cols = [
                    "vehicle_history_no", "vehicle_history_changed_at", "vehicle_history_changed_by",
                    "vehicle_history_pt", "vehicle_history_asset_owner", "vehicle_history_active_name",
                    "vehicle_history_name_as_asset_pt", "vehicle_history_new_asset_name", "vehicle_history_merk",
                    "vehicle_history_type", "vehicle_history_jenis", "vehicle_history_plate_old", "vehicle_history_plate_new",
                    "vehicle_history_year_of_use", "vehicle_history_user_old", "vehicle_history_user_new",
                    "vehicle_history_status", "vehicle_history_pt_pemakai", "vehicle_history_kondisi",
                    "vehicle_history_lokasi", "vehicle_history_keterangan"
                ]
            if not user_hist_cols:
                user_hist_cols = ["user_history_changed_at", "user_history_user_old", "user_history_user_new", "user_history_note"]
        hist_cols = svc_hist_cols + kir_hist_cols + loan_hist_cols + vehicle_hist_cols + user_hist_cols
        base_columns = [col for col in valid_columns if col not in hist_cols]

        # Export Data Kendaraan hanya untuk kendaraan aktif/hidup.
        # Kendaraan dengan status TERJUAL tidak ikut pilihan/download ini;
        # data tersebut tetap khusus di Laporan Terjual.
        query = _apply_vehicle_scope(_active_vehicle_base_query())
        cleaned_pt = [str(x).strip() for x in selected_pt if str(x).strip()]
        # Untuk admin PT: filter hanya ke PT yang diizinkan, abaikan pilihan PT dari client yang di luar scope
        if _is_admin_pt_session():
            allowed_pts = _scope_pt_names()
            if cleaned_pt:
                cleaned_pt = [pt for pt in cleaned_pt if pt.lower() in {x.lower() for x in allowed_pts}]
            if not cleaned_pt:
                cleaned_pt = allowed_pts
        if cleaned_pt: query = query.filter(Vehicle.pt.in_(cleaned_pt))

        vehicles = query.order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).all()
        if not vehicles: return jsonify({"success": False, "message": "Tidak ada data kendaraan yang cocok."}), 404

        v_ids = [v.id for v in vehicles]
        latest_kir_map, five_map, annual_map, custom_values_map = _latest_kir_map(v_ids), _latest_five_tax_map(v_ids), _annual_payment_map(v_ids), _vehicle_custom_values_map(v_ids)

        svc_rows_by_v, loan_rows_by_v, kir_rows_by_v, vehicle_hist_by_v, user_hist_by_v = {}, {}, {}, {}, {}
        if v_ids:
            all_svc = ServiceRecord.query.filter(ServiceRecord.vehicle_id.in_(v_ids)).order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).all()
            for r in all_svc: svc_rows_by_v.setdefault(r.vehicle_id, []).append(r)
            all_loans = LoanTransaction.query.filter(LoanTransaction.vehicle_id.in_(v_ids)).order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).all()
            for r in all_loans: loan_rows_by_v.setdefault(r.vehicle_id, []).append(r)
            all_kirs = KirRecord.query.filter(KirRecord.vehicle_id.in_(v_ids)).order_by(KirRecord.done_date.desc(), KirRecord.id.desc()).all()
            for r in all_kirs: kir_rows_by_v.setdefault(r.vehicle_id, []).append(r)
            if vehicle_hist_cols:
                vehicle_hist_by_v = _vehicle_history_sessions_for_export(v_ids)
            if user_hist_cols:
                try:
                    sql_user_hist = text("""
                        SELECT id, vehicle_id, user_lama, user_baru, change_date, note
                        FROM user_histories
                        WHERE vehicle_id IN :ids
                        ORDER BY change_date DESC, id DESC
                    """).bindparams(bindparam('ids', expanding=True))
                    all_user_hist = db.session.execute(sql_user_hist, {'ids': v_ids}).mappings().all()
                    for r in all_user_hist: user_hist_by_v.setdefault(r.get('vehicle_id'), []).append(r)
                except Exception:
                    all_user_hist = UserHistory.query.filter(UserHistory.vehicle_id.in_(v_ids)).order_by(UserHistory.id.desc()).all()
                    for r in all_user_hist: user_hist_by_v.setdefault(r.vehicle_id, []).append(r)

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Data Kendaraan"

        sections = []
        if base_columns: sections.append(("DATA KENDARAAN", base_columns))
        if svc_hist_cols: sections.append(("RIWAYAT SERVIS", svc_hist_cols))
        if kir_hist_cols: sections.append(("RIWAYAT KIR", kir_hist_cols))
        if loan_hist_cols: sections.append(("RIWAYAT PEMINJAMAN", loan_hist_cols))
        if vehicle_hist_cols: sections.append(("RIWAYAT KENDARAAN", vehicle_hist_cols))
        if user_hist_cols: sections.append(("RIWAYAT USER", user_hist_cols))

        top_labels, second_labels = [], []
        for title, cols in sections:
            top_labels.extend([title] + [""] * (len(cols) - 1))
            second_labels.extend([column_defs[col]["label"] for col in cols])
        ws.append(top_labels)
        ws.append(second_labels)

        group_fill, sub_fill = PatternFill("solid", fgColor="17365D"), PatternFill("solid", fgColor="1E4080")
        group_font, sub_font, thin_side = Font(bold=True, color="FFFFFF"), Font(bold=True, color="FFD15F"), Side(style="thin", color="365F91")
        
        col_idx = 1
        for title, cols in sections:
            if len(cols) > 1: ws.merge_cells(start_row=1, end_row=1, start_column=col_idx, end_column=col_idx + len(cols) - 1)
            for i in range(len(cols)):
                cell = ws.cell(row=1, column=col_idx + i)
                cell.fill, cell.font, cell.alignment = group_fill, group_font, Alignment(horizontal="center", vertical="center")
            col_idx += len(cols)

        for c in range(1, len(second_labels) + 1):
            cell = ws.cell(row=2, column=c)
            cell.fill, cell.font, cell.alignment, cell.border = sub_fill, sub_font, Alignment(horizontal="center", vertical="center", wrap_text=True), Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

        # Buat object style sekali di luar loop, di-reuse tiap cell agar tidak
        # membuat ribuan object Alignment/Border baru untuk setiap kendaraan x history x kolom.
        _cell_align = Alignment(vertical="top", wrap_text=True)
        _cell_border = Border(left=thin_side, right=thin_side)
        _last_border = Border(bottom=thin_side, left=thin_side, right=thin_side)

        row_cursor = 3
        for idx, v in enumerate(vehicles, 1):
            row_base_data = _base_row_data(idx, v)
            row_base_data.update(custom_values_map.get(v.id, {}))
            
            svc_items = [r for r in svc_rows_by_v.get(v.id, []) if (not d_from or (r.service_date and r.service_date >= d_from)) and (not d_to or (r.service_date and r.service_date <= d_to))]
            kir_items = [r for r in kir_rows_by_v.get(v.id, []) if (not d_from or ((r.done_date or r.due_date) and (r.done_date or r.due_date) >= d_from)) and (not d_to or ((r.done_date or r.due_date) and (r.done_date or r.due_date) <= d_to))]
            loan_items = [r for r in loan_rows_by_v.get(v.id, []) if (not d_from or (r.date_out and r.date_out >= d_from)) and (not d_to or (r.date_out and r.date_out <= d_to))]
            vehicle_hist_items = [r for r in vehicle_hist_by_v.get(v.id, []) if _date_in_range((_history_timestamp_value(r, 'changed_at', 'change_date', 'created_at', 'updated_at') or datetime.min).date(), d_from, d_to)]
            user_hist_items = []
            for r in user_hist_by_v.get(v.id, []):
                def _uh_get(name):
                    try:
                        return r.get(name) if hasattr(r, 'get') else getattr(r, name, None)
                    except Exception:
                        return None
                uh_dt = _coerce_datetime_value(_uh_get('change_date')) or _coerce_datetime_value(_uh_get('changed_at')) or _coerce_datetime_value(_uh_get('created_at'))
                if _date_in_range(uh_dt.date() if hasattr(uh_dt, 'date') else uh_dt, d_from, d_to):
                    user_hist_items.append(r)

            history_count = max(
                len(svc_items) if svc_hist_cols else 0,
                len(kir_items) if kir_hist_cols else 0,
                len(loan_items) if loan_hist_cols else 0,
                len(vehicle_hist_items) if vehicle_hist_cols else 0,
                len(user_hist_items) if user_hist_cols else 0,
                1
            )
            
            start_row = row_cursor
            for h_idx in range(history_count):
                row_values = []
                for col in base_columns: row_values.append(row_base_data.get(col, "-") if h_idx == 0 else "")
                for col in svc_hist_cols: row_values.append(_service_history_value(svc_items[h_idx], col) if h_idx < len(svc_items) else ("-" if h_idx == 0 and not svc_items else ""))
                for col in kir_hist_cols: row_values.append(_kir_history_value(kir_items[h_idx], col) if h_idx < len(kir_items) else ("-" if h_idx == 0 and not kir_items else ""))
                for col in loan_hist_cols: row_values.append(_loan_history_value(loan_items[h_idx], col) if h_idx < len(loan_items) else ("-" if h_idx == 0 and not loan_items else ""))
                for col in vehicle_hist_cols: row_values.append(_vehicle_change_history_value(vehicle_hist_items[h_idx], col) if h_idx < len(vehicle_hist_items) else ("-" if h_idx == 0 and not vehicle_hist_items else ""))
                for col in user_hist_cols: row_values.append(_user_history_value(user_hist_items[h_idx], col) if h_idx < len(user_hist_items) else ("-" if h_idx == 0 and not user_hist_items else ""))
                ws.append(row_values)
                for c in range(1, len(row_values) + 1):
                    cell = ws.cell(row=row_cursor, column=c)
                    cell.alignment = _cell_align
                    cell.border = _cell_border
                row_cursor += 1
            
            if history_count > 1 and base_columns:
                for c in range(1, len(base_columns) + 1): ws.merge_cells(start_row=start_row, end_row=row_cursor - 1, start_column=c, end_column=c)
            for c in range(1, ws.max_column + 1): ws.cell(row=row_cursor-1, column=c).border = _last_border

        _autosize_worksheet_columns(ws, min_width=12, max_width=40)

        ws.freeze_panes = "A3"
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Data_Kendaraan_{date.today().strftime('%d-%m-%Y')}.xlsx")
    except Exception as exc:
        return jsonify({"success": False, "message": f"Export gagal: {exc}"}), 500
        
@bp.get("/perusahaan/ringkasan")
def perusahaan_ringkasan():
    company = (request.args.get("company") or "").strip()
    company = _sanitize_company_filter(company)
    if not company:
        flash("Pilih PT yang valid dulu.", "warning")
        return redirect(url_for("main.kendaraan"))
    payload = _company_detail_payload(company, _parse_date(request.args.get('date_from')), _parse_date(request.args.get('date_to')))
    return render_template(
        "perusahaan_ringkasan.html",
        payload=payload,
        company=company,
        title=f"Ringkasan {company}",
        back_url=_detail_back_url(),
    )

@bp.get('/api/perusahaan/detail')
def api_perusahaan_detail():
    company = _sanitize_company_filter((request.args.get('company') or '').strip())
    if not company:
        return jsonify({'success': False, 'message': 'Pilih PT yang valid dulu.'}), 400
    payload = _company_detail_payload(company, _parse_date(request.args.get('date_from')), _parse_date(request.args.get('date_to')))
    return jsonify({'success': True, **payload})

@bp.get("/kendaraan/upload")
@master_required
def kendaraan_upload_page():
    return render_template("kendaraan_upload.html",headers=UPLOAD_HEADERS,title="Upload Data Kendaraan - Sinar Group")

@bp.post("/kendaraan/upload")
@master_required
def kendaraan_upload_post():
    file=request.files.get("file")
    if not file or not file.filename:
        flash("Pilih file Excel terlebih dulu.","danger")
        return redirect(url_for("main.kendaraan_upload_page"))
    ext=Path(file.filename).suffix.lower()
    if ext not in {".xlsx",".xlsm"}:
        flash("File harus format .xlsx atau .xlsm.","danger")
        return redirect(url_for("main.kendaraan_upload_page"))
    actor_name=session.get("user_name") or "SYSTEM"
    allowed_pts = _scope_pt_names()  # kosong [] jika master, terisi jika admin PT
    try:
        created,updated,total,skipped_pts=_import_vehicle_rows(
            file,
            actor_name=actor_name,
            allowed_pts=allowed_pts if allowed_pts else None
        )
        try:audit("UPLOAD","Vehicle",0,note=f"Upload data kendaraan: total={total}, baru={created}, update={updated}")
        except Exception:pass
        msg=f"Upload berhasil. Total {total} baris diproses, {created} data baru, {updated} data diperbarui."
        if skipped_pts:
            msg+=f" ({len(skipped_pts)} PT di-skip karena di luar akses Anda: {', '.join(sorted(skipped_pts))})"
        flash(msg,"success")
    except Exception as exc:
        Vehicle.query.session.rollback()
        flash(f"Upload gagal: {exc}","danger")
    return redirect(url_for("main.kendaraan"))

@bp.post("/kendaraan/tambah")
@master_required
def kendaraan_tambah():
    """Tambah baris kendaraan baru langsung dari tabel (tanpa upload Excel)."""
    session_db=Vehicle.query.session
    actor_name=session.get("user_name") or "SYSTEM"
    payload=_request_payload()
    
    # Scope Validation: Admin PT hanya boleh menambah data untuk PT yang diizinkan
    pt_name = _normalize_text(payload.get("pt"))
    allowed_pts = _scope_pt_names()
    
    # Jika input PT kosong tapi user punya daftar PT yang diizinkan, gunakan PT pertama sebagai default
    if not pt_name and allowed_pts:
        pt_name = allowed_pts[0]
    
    if not _is_allowed_pt_name(pt_name):
        return jsonify({
            "success": False, 
            "message": f"Anda tidak memiliki akses untuk menambah data pada PT: {pt_name or '-'}. Akses Anda: {', '.join(allowed_pts) or 'Semua PT'}"
        }), 403

    try:
        now=_now_naive()
        v=Vehicle(created_at=now,updated_at=now,is_deleted=False)
        session_db.add(v)
        session_db.flush()

        v.pt=pt_name or None
        v.no=_parse_int(payload.get("no"))

        asset_owner = _get_or_create_company(payload.get("asset_owner_company_name"))
        pt_pemakai = _get_or_create_company(payload.get("pt_pemakai_company_name"))

        v.asset_owner_company_id = asset_owner.id if asset_owner else None
        v.pt_pemakai_company_id = pt_pemakai.id if pt_pemakai else None

        v.active_name         =_normalize_text(payload.get("active_name"))         or None
        v.name_as_asset_pt    =_normalize_text(payload.get("name_as_asset_pt"))    or None
        v.new_asset_name      =_normalize_text(payload.get("new_asset_name"))      or None
        v.merk                =_normalize_text(payload.get("merk"))                or None
        v.type                =_normalize_text(payload.get("type"))                or None
        v.jenis               =_normalize_text(payload.get("jenis"))               or None
        v.plate_old           =_normalize_text(payload.get("plate_old"))           or None
        v.plate_new           =_normalize_text(payload.get("plate_new"))           or None
        v.year_of_use         =_parse_year(payload.get("year_of_use"))
        v.user_old            =_normalize_text(payload.get("user_old"))            or None
        v.user_new            =_normalize_text(payload.get("user_new"))            or None
        v.status              =_normalize_text(payload.get("status"))              or None
        v.kondisi_terkini     =_normalize_text(payload.get("kondisi_terkini"))     or None
        v.lokasi              =_normalize_text(payload.get("lokasi"))              or None
        v.tambahan_keterangan =_normalize_text(payload.get("tambahan_keterangan")) or None

        custom_columns = _vehicle_custom_columns()
        for col in custom_columns:
            _set_vehicle_custom_value(v.id, col['column_key'], payload.get(col['column_key']) or '')

        session_db.flush()
        try:
            audit("CREATE","Vehicle",v.id,note=f"Tambah baris kendaraan baru via tabel inline oleh {actor_name}")
        except Exception:
            pass
        session_db.commit()

        custom_values = _vehicle_custom_values_map([v.id]).get(v.id, {})
        return jsonify({
            "success": True,
            "message": "Baris kendaraan berhasil ditambahkan.",
            "id": v.id,
            "vehicle_label": vehicle_display_name(v),
            "detail_url": url_for("main.vehicle_detail", vehicle_id=v.id),
            "delete_url": url_for("main.kendaraan_delete", vehicle_id=v.id),
            "edit_url": url_for("main.kendaraan_edit_page", vehicle_id=v.id),
            "field_update_url": url_for("main.kendaraan_update_field", vehicle_id=v.id),
            "pt": v.pt or "",
            "asset_owner_company_name": _company_name(getattr(v,"asset_owner_company",None)),
            "active_name": v.active_name or "",
            "name_as_asset_pt": v.name_as_asset_pt or "",
            "new_asset_name": v.new_asset_name or "",
            "merk": v.merk or "",
            "type": v.type or "",
            "jenis": v.jenis or "",
            "plate_old": v.plate_old or "",
            "plate_new": v.plate_new or "",
            "year_of_use": str(v.year_of_use) if v.year_of_use else "",
            "user_old": v.user_old or "",
            "user_new": v.user_new or "",
            "status": v.status or "",
            "pt_pemakai_company_name": _company_name(getattr(v,"pt_pemakai_company",None)),
            "kondisi_terkini": v.kondisi_terkini or "",
            "lokasi": v.lokasi or "",
            "tambahan_keterangan": v.tambahan_keterangan or "",
            "values": {col["column_key"]: custom_values.get(col["column_key"], "") for col in custom_columns},
        })
    except Exception as exc:
        session_db.rollback()
        return jsonify({"success":False,"message":str(exc)}),400

@bp.post('/kendaraan/columns')
@master_required
def kendaraan_column_create():
    payload = _request_payload()
    label = _normalize_text(payload.get('label'))
    raw_key = _normalize_text(payload.get('key'))
    key = (raw_key or label).lower().replace(' ', '_')
    insert_index = _parse_int(payload.get('insert_index'))
    if insert_index is None:
        insert_index = 999

    if not label:
        return jsonify({'success': False, 'message': 'Nama kolom wajib diisi.'}), 400

    key = ''.join(ch for ch in key if ch.isalnum() or ch == '_').strip('_')
    if not key:
        return jsonify({'success': False, 'message': 'Key kolom tidak valid.'}), 400

    if key in EDITABLE_VEHICLE_FIELDS:
        return jsonify({'success': False, 'message': 'Nama kolom bentrok dengan kolom bawaan.'}), 400

    _ensure_vehicle_custom_tables()
    exists = db.session.execute(text('SELECT 1 FROM vehicle_custom_columns WHERE column_key=:key'), {'key': key}).first()
    if exists:
        return jsonify({'success': False, 'message': 'Kolom dengan key ini sudah ada.'}), 400

    db.session.execute(
        text('INSERT INTO vehicle_custom_columns (column_key, column_label, position_index) VALUES (:key,:label,:pos)'),
        {'key':key,'label':label,'pos':insert_index}
    )
    db.session.commit()
    return jsonify({
        'success': True,
        'message': 'Kolom berhasil ditambahkan.',
        'column': {'key': key, 'label': label, 'insertIndex': insert_index}
    })

@bp.post('/kendaraan/columns/<string:column_key>/delete')
@master_required
def kendaraan_column_delete(column_key:str):
    # Hapus kolom tambahan yang dibuat user.
    # Nilai yang terkirim dari tampilan bisa berupa label (contoh: "Kolom Baru")
    # atau key database (contoh: "kolom_baru"), jadi lookup dibuat fleksibel.
    raw_value = _normalize_text(column_key)
    clean_key = raw_value.lower().replace(' ', '_')
    clean_key = ''.join(ch for ch in clean_key if ch.isalnum() or ch == '_').strip('_')
    if not clean_key and not raw_value:
        return jsonify({'success': False, 'message': 'Kolom tidak valid.'}), 400

    _ensure_vehicle_custom_tables()

    row = db.session.execute(
        text('''
            SELECT column_key
            FROM vehicle_custom_columns
            WHERE column_key = :clean_key
               OR LOWER(TRIM(column_label)) = LOWER(TRIM(:raw_value))
               OR LOWER(REPLACE(TRIM(column_label), ' ', '_')) = :clean_key
            LIMIT 1
        '''),
        {'clean_key': clean_key, 'raw_value': raw_value}
    ).mappings().first()

    if not row:
        return jsonify({'success': False, 'message': 'Kolom tambahan tidak ditemukan atau sudah terhapus.'}), 404

    target_key = row['column_key']
    db.session.execute(text('DELETE FROM vehicle_custom_values WHERE column_key=:key'), {'key': target_key})
    db.session.execute(text('DELETE FROM vehicle_custom_columns WHERE column_key=:key'), {'key': target_key})
    db.session.commit()
    return jsonify({'success': True, 'message': 'Kolom berhasil dihapus.'})

@bp.get("/kendaraan/<int:vehicle_id>")
@master_required
def vehicle_detail(vehicle_id:int):
    """Alias klik Nama Aktiva di tabel.

    Template kendaraan.html masih banyak link lama ke /kendaraan/<id>.
    Supaya klik Nama Aktiva langsung masuk halaman detail kendaraan yang benar,
    route ini diarahkan ke route detail_page yang sudah memakai template
    kendaraan_detail.html.
    """
    return redirect(url_for("main.vehicle_detail_page", vehicle_id=vehicle_id))

@bp.get("/kendaraan/<int:vehicle_id>/detail")
def kendaraan_detail_json(vehicle_id: int):
    """JSON endpoint untuk modal Detail Kendaraan di halaman daftar kendaraan."""
    v = Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)
    date_from = _parse_date(request.args.get('date_from'))
    date_to = _parse_date(request.args.get('date_to'))
    custom_values = _vehicle_custom_values_map([v.id]).get(v.id, {})

    # ── Data kendaraan ─────────────────────────────────────────────────────
    vehicle_data = {
        "id": v.id,
        "label": vehicle_display_name(v),
        "pt": v.pt or "",
        "asset_owner_company_name": _company_name(getattr(v, "asset_owner_company", None)),
        "pt_pemakai_1_company_name": _company_name(getattr(v, "pt_pemakai_1_company", None)),
        "pt_pemakai_2_company_name": _company_name(getattr(v, "pt_pemakai_2_company", None)),
        "active_name": v.active_name or "",
        "name_as_asset_pt": v.name_as_asset_pt or "",
        "new_asset_name": v.new_asset_name or "",
        "merk": v.merk or "",
        "type": v.type or "",
        "jenis": v.jenis or "",
        "plate_old": v.plate_old or "",
        "plate_new": v.plate_new or vehicle_plate(v),
        "year_of_use": str(v.year_of_use) if v.year_of_use else "",
        "user_old": v.user_old or "",
        "user_new": v.user_new or "",
        "status": v.status or "",
        "kondisi_terkini": v.kondisi_terkini or "",
        "lokasi": v.lokasi or "",
        "tambahan_keterangan": v.tambahan_keterangan or "",
        "updated_at": v.updated_at.isoformat() if getattr(v, "updated_at", None) else None,
        "created_at": v.created_at.isoformat() if getattr(v, "created_at", None) else None,
        "custom_values": custom_values,
    }

    # ── Peminjaman ──────────────────────────────────────────────────────────
    loan_rows = (
        LoanTransaction.query
        .filter_by(vehicle_id=v.id)
        .order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc())
        .all()
    )
    loan_place_lookup = _loan_place_map([r.id for r in loan_rows]) if loan_rows else {}
    loans = []
    for r in loan_rows:
        date_out = getattr(r, "date_out", None) or getattr(r, "loan_date_out", None) or getattr(r, "tanggal_pinjam", None)
        date_plan = getattr(r, "date_return_plan", None) or getattr(r, "tanggal_kembali_rencana", None)
        date_actual = getattr(r, "date_return_actual", None) or getattr(r, "tanggal_kembali_aktual", None)
        loan_filter_date = date_out or date_plan or date_actual
        if not _date_in_range(loan_filter_date, date_from, date_to):
            continue
        loans.append({
            "id": r.id,
            "date_out": date_out.isoformat() if date_out else None,
            "date_return_plan": date_plan.isoformat() if date_plan else None,
            "date_return_actual": date_actual.isoformat() if date_actual else None,
            "borrower": getattr(r, "borrower_name", None) or getattr(r, "borrower", None) or getattr(r, "peminjam", None) or "",
            "borrower_company": getattr(r, "borrower_company", None) or "",
            "place": loan_place_lookup.get(r.id, "") or "",
            "purpose": getattr(r, "purpose", None) or getattr(r, "keperluan", None) or "",
            "status": getattr(r, "status", None) or "",
            "note": getattr(r, "note", None) or getattr(r, "catatan", None) or "",
            "created_at": r.created_at.isoformat() if getattr(r, "created_at", None) else None,
        })

    # ── Servis ──────────────────────────────────────────────────────────────
    service_rows = (
        ServiceRecord.query
        .filter_by(vehicle_id=v.id)
        .order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc())
        .all()
    )
    services = []
    for r in service_rows:
        svc_date = getattr(r, "service_date", None)
        next_due = service_due_date(r)
        if not _date_in_range(svc_date, date_from, date_to):
            continue
        services.append({
            "id": r.id,
            "service_date": svc_date.isoformat() if svc_date else None,
            "service_type": getattr(r, "service_type", None) or getattr(r, "jenis_servis", None) or getattr(r, "jenis_service", None) or "",
            "vendor": getattr(r, "vendor", None) or getattr(r, "bengkel", None) or "",
            "cost": getattr(r, "cost", None) or getattr(r, "biaya", None) or 0,
            "odometer": getattr(r, "odometer", None) or 0,
            "next_due": next_due.isoformat() if next_due else None,
            "note": getattr(r, "note", None) or getattr(r, "catatan", None) or "",
        })

    # ── Pajak Tahunan ───────────────────────────────────────────────────────
    annual_taxes = []
    try:
        annual_rows = db.session.execute(
            text(
                "SELECT id, vehicle_id, tax_year, due_date, paid_date, amount, note "
                "FROM annual_tax_payments WHERE vehicle_id = :vid ORDER BY tax_year DESC, id DESC"
            ),
            {"vid": v.id}
        ).mappings().all()
        for r in annual_rows:
            row_date = r["paid_date"] or r["due_date"]
            if not _date_in_range(row_date, date_from, date_to):
                continue
            annual_taxes.append({
                "id": r["id"],
                "tax_year": r["tax_year"],
                "due_date": r["due_date"].isoformat() if r["due_date"] else None,
                "paid_date": r["paid_date"].isoformat() if r["paid_date"] else None,
                "amount": r["amount"] or 0,
                "note": r["note"] or "",
            })
    except Exception:
        pass

    # ── Pajak 5 Tahunan ─────────────────────────────────────────────────────
    five_year_taxes = []
    try:
        five_rows = db.session.execute(
            text(
                "SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note "
                "FROM five_year_tax_payments WHERE vehicle_id = :vid ORDER BY due_date DESC, id DESC"
            ),
            {"vid": v.id}
        ).mappings().all()
        for r in five_rows:
            row_date = r["paid_date"] or r["due_date"]
            if not _date_in_range(row_date, date_from, date_to):
                continue
            five_year_taxes.append({
                "id": r["id"],
                "due_date": r["due_date"].isoformat() if r["due_date"] else None,
                "paid_date": r["paid_date"].isoformat() if r["paid_date"] else None,
                "amount": r["amount"] or 0,
                "plate_before": r["plate_before"] or "",
                "plate_after": r["plate_after"] or "",
                "note": r["note"] or "",
            })
    except Exception:
        pass

    # ── KIR ─────────────────────────────────────────────────────────────────
    kir_rows = (
        KirRecord.query
        .filter_by(vehicle_id=v.id)
        .order_by(KirRecord.due_date.desc(), KirRecord.id.desc())
        .all()
    )
    kirs = []
    for r in kir_rows:
        done_date = getattr(r, "done_date", None) or getattr(r, "kir_date", None)
        due_date = getattr(r, "due_date", None)
        kir_filter_date = done_date or due_date
        if not _date_in_range(kir_filter_date, date_from, date_to):
            continue
        kirs.append({
            "id": r.id,
            "done_date": done_date.isoformat() if done_date else None,
            "due_date": due_date.isoformat() if due_date else None,
            "result": getattr(r, "result", None) or getattr(r, "hasil", None) or "",
            "note": getattr(r, "note", None) or getattr(r, "catatan", None) or "",
            "status": _kir_status(due_date) if due_date else "empty",
        })

    # ── History Kendaraan ───────────────────────────────────────────────────
    change_rows = (
        VehicleChangeHistory.query
        .filter_by(vehicle_id=v.id)
        .order_by(VehicleChangeHistory.changed_at.desc(), VehicleChangeHistory.id.desc())
        .limit(200).all()
    )
    vehicle_history = []
    for r in change_rows:
        vh_date = getattr(r, "changed_at", None)
        if not _date_in_range(vh_date.date() if hasattr(vh_date, "date") else vh_date, date_from, date_to):
            continue
        vehicle_history.append({
            "id": r.id,
            "changed_at": r.changed_at.isoformat() if getattr(r, "changed_at", None) else None,
            "field_label": getattr(r, "field_label", None) or getattr(r, "field_name", None) or "",
            "old_value": getattr(r, "old_value", None) or "",
            "new_value": getattr(r, "new_value", None) or "",
            "change_type": getattr(r, "change_type", None) or "",
            "changed_by": getattr(r, "changed_by", None) or "system",
            "note": getattr(r, "note", None) or "",
        })

    # ── History User ────────────────────────────────────────────────────────
    user_hist_rows = (
        UserHistory.query
        .filter_by(vehicle_id=v.id)
        .order_by(UserHistory.change_date.desc(), UserHistory.id.desc())
        .limit(100).all()
    )
    user_history = []
    for r in user_hist_rows:
        uh_date = _history_timestamp_value(r, "change_date", "changed_at", "created_at", "updated_at")
        if not _date_in_range(uh_date.date() if hasattr(uh_date, "date") else uh_date, date_from, date_to):
            continue
        user_history.append({
            "id": r.id,
            "changed_at": uh_date.isoformat() if uh_date else None,
            "user_lama": getattr(r, "user_lama", None) or "",
            "user_baru": getattr(r, "user_baru", None) or "",
            "changed_by": getattr(r, "changed_by", None) or "system",
            "note": getattr(r, "note", None) or "",
        })

    # ── History Peminjaman ──────────────────────────────────────────────────
    loan_hist_rows = (
        LoanHistory.query
        .filter_by(vehicle_id=v.id)
        .order_by(LoanHistory.changed_at.desc(), LoanHistory.id.desc())
        .limit(100).all()
    ) if hasattr(LoanHistory, "vehicle_id") else []
    loan_history = []
    for r in loan_hist_rows:
        lh_date = _history_timestamp_value(r, "changed_at", "created_at", "updated_at")
        if not _date_in_range(lh_date.date() if hasattr(lh_date, "date") else lh_date, date_from, date_to):
            continue
        loan_history.append({
            "id": r.id,
            "changed_at": lh_date.isoformat() if lh_date else None,
            "changed_by": getattr(r, "changed_by", None) or "system",
            "note": getattr(r, "note", None) or "",
            "description": getattr(r, "description", None) or "",
        })

    summary = _summarize_vehicle_detail(v, loans, services, annual_taxes, five_year_taxes, kirs, vehicle_history, user_history, loan_history)
    return jsonify({
        "success": True,
        "date_from": _iso(date_from),
        "date_to": _iso(date_to),
        "summary": summary,
        "vehicle": vehicle_data,
        "loans": loans,
        "services": services,
        "annual_taxes": annual_taxes,
        "five_year_taxes": five_year_taxes,
        "kirs": kirs,
        "vehicle_history": vehicle_history,
        "user_history": user_history,
        "loan_history": loan_history,
    })


@bp.get("/kendaraan/<int:vehicle_id>/edit")
@master_required
def kendaraan_edit_page(vehicle_id:int):
    v=Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)
    companies=Company.query.order_by(Company.name.asc()).all()
    custom_columns=_vehicle_custom_columns()
    custom_values=_vehicle_custom_values_map([v.id]).get(v.id,{})
    return render_template("kendaraan_edit.html",vehicle=v,companies=companies,custom_columns=custom_columns,custom_values=custom_values,title=f"Edit Kendaraan - {vehicle_plate(v)}")

@bp.post("/kendaraan/<int:vehicle_id>/edit")
@master_required
def kendaraan_edit_post(vehicle_id:int):
    v=Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)
    session_db=Vehicle.query.session
    before=_vehicle_snapshot(v)
    actor_name=session.get("user_name") or "SYSTEM"
    pending_plate_update=_pending_plate_update_row(v.id)
    payload=_request_payload()

    # Scope Validation: Admin PT hanya boleh mengedit data ke PT yang diizinkan
    pt_name = _normalize_text(payload.get("pt"))
    if not _is_allowed_pt_name(pt_name):
        flash(f"Gagal update: Anda tidak memiliki akses ke PT {pt_name or '-'}.", "danger")
        return redirect(url_for("main.kendaraan_edit_page", vehicle_id=v.id))

    try:
        asset_owner = _get_or_create_company(payload.get("pt_pemilik_aset"))
        pt_pemakai = _get_or_create_company(payload.get("pt_pemakai"))

        submitted_plate_old = _normalize_text(payload.get("plate_old")) or None
        submitted_plate_new = _normalize_text(payload.get("plate_new")) or None

        v.asset_owner_company_id = asset_owner.id if asset_owner else None
        v.pt_pemakai_company_id = pt_pemakai.id if pt_pemakai else None

        if pending_plate_update:
            plate_before=_normalize_text(pending_plate_update.get("plate_before"))
            if not submitted_plate_new or submitted_plate_new == plate_before:
                flash("Plat baru wajib diupdate di Data Kendaraan karena pajak 5 tahunan sudah lunas.","danger")
                return redirect(url_for("main.kendaraan_edit_page",vehicle_id=v.id))

        v.no = _parse_int(payload.get("no"))
        v.pt = pt_name or None
        v.asset_owner_company_id = asset_owner.id if asset_owner else None
        v.pt_pemakai_company_id = pt_pemakai.id if pt_pemakai else None
        v.active_name = _normalize_text(payload.get("active_name")) or None
        v.name_as_asset_pt = _normalize_text(payload.get("name_as_asset_pt")) or None
        v.new_asset_name = _normalize_text(payload.get("new_asset_name")) or None
        v.merk = _normalize_text(payload.get("merk")) or None
        v.type = _normalize_text(payload.get("type")) or None
        v.jenis = _normalize_text(payload.get("jenis")) or None
        v.plate_old = submitted_plate_old
        v.plate_new = submitted_plate_new
        v.year_of_use = _parse_year(payload.get("year_of_use"))
        v.user_old = _normalize_text(payload.get("user_old")) or None
        v.user_new = _normalize_text(payload.get("user_new")) or None
        v.status = _normalize_text(payload.get("status")) or None
        v.kondisi_terkini = _normalize_text(payload.get("kondisi_terkini")) or None
        v.lokasi = _normalize_text(payload.get("lokasi")) or None
        v.tambahan_keterangan = _normalize_text(payload.get("tambahan_keterangan")) or None
        v.updated_at = _now_naive()
        # Simpan kolom kustom dari form edit agar masuk history
        custom_columns=_vehicle_custom_columns()
        for col in custom_columns:
            col_key=col['column_key']
            if col_key in payload:
                _set_vehicle_custom_value(v.id, col_key, payload.get(col_key) or '')
        session_db.flush()
        _sync_vehicle_plate_update(v)
        after=_vehicle_snapshot(v)
        _log_vehicle_edit(v,before,after,changed_by=actor_name)
        session_db.commit()
        if pending_plate_update:
            flash("Data kendaraan berhasil diupdate. Plat baru sudah tercatat ke history pembayaran.","success")
        else:
            flash("Data kendaraan berhasil diupdate dan masuk ke history.","success")
        return redirect(url_for("main.vehicle_detail",vehicle_id=v.id))
    except Exception as exc:
        session_db.rollback()
        flash(f"Gagal update kendaraan: {exc}","danger")
        return redirect(url_for("main.kendaraan_edit_page",vehicle_id=v.id))

@bp.post("/kendaraan/<int:vehicle_id>/delete")
@master_required
def kendaraan_delete(vehicle_id: int):
    v = Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)
    v.is_deleted = True

    try:
        audit("DELETE", "Vehicle", v.id, note=f"Hapus kendaraan {vehicle_plate(v)}")
    except Exception:
        pass

    return _safe_commit(
        "Data kendaraan berhasil dihapus.",
        "Gagal hapus kendaraan",
        "main.kendaraan"
    )

def _replace_year_safe(base_date, year:int):
    if not base_date:
        return None
    try:
        return base_date.replace(year=year)
    except ValueError:
        if base_date.month == 2 and base_date.day == 29:
            return date(year, 2, 28)
        raise


def _shift_years_safe(base_date, years:int):
    if not base_date:
        return None
    return _replace_year_safe(base_date, base_date.year + years)


def _add_months_safe(base_date, months:int):
    """
    Tambah bulan kalender secara aman.
    Contoh:
    - 31-12-2025 + 6 bulan = 30-06-2026
    - 31-01-2026 + 6 bulan = 31-07-2026
    - 31-08-2026 + 6 bulan = 28-02-2027
    """
    if not base_date:
        return None

    import calendar

    month = base_date.month + months
    year = base_date.year + (month - 1) // 12
    month = (month - 1) % 12 + 1

    last_day = calendar.monthrange(year, month)[1]
    day = min(base_date.day, last_day)

    return date(year, month, day)

def get_annual_due(main_due_date, year:int|None=None):
    if not main_due_date:
        return None
    target_year = year or date.today().year
    return _replace_year_safe(main_due_date, target_year)


def annual_due_to_five_year_due(annual_due_date):
    # Siklus Indonesia: 4x pajak tahunan, lalu tahun ke-5 pajak 5 tahunan/ganti plat.
    # Jadi kalau exp tahunan aktif 2027, exp 5 tahunan jatuh di 2031.
    return _shift_years_safe(annual_due_date, 4)


def five_year_due_to_annual_due(five_year_due_date):
    # Kebalikan dari annual_due_to_five_year_due: start siklus tahunan = exp 5 tahunan - 4 tahun.
    return _shift_years_safe(five_year_due_date, -4)


def _is_paid_complete(paid_date, amount):
    return bool(paid_date and amount is not None and float(amount) > 0)


def _tax_payment_is_paid(row) -> bool:
    """True kalau row pembayaran sudah lunas."""
    if not row:
        return False
    return _is_paid_complete(row.get("paid_date"), row.get("amount"))


def _ensure_unpaid_tax_payment(row, label="pajak"):
    """Cegah bayar dua kali untuk jatuh tempo/record yang sudah lunas."""
    if _tax_payment_is_paid(row):
        raise ValueError(f"{label} ini sudah dibayar. Tidak bisa bayar lagi sampai jadwal berikutnya.")


def _create_next_annual_stub(vehicle_id:int, due_date, note=None, now=None):
    """Buat jadwal pajak tahunan berikutnya kalau belum ada."""
    if not vehicle_id or not due_date:
        return None
    now = now or _now_naive()
    tax_year = due_date.year
    existing = db.session.execute(
        text("""
            SELECT id
            FROM annual_tax_payments
            WHERE vehicle_id = :vehicle_id AND tax_year = :tax_year
            ORDER BY id DESC
            LIMIT 1
        """),
        {"vehicle_id": vehicle_id, "tax_year": tax_year},
    ).mappings().first()
    if existing:
        return existing["id"]
    db.session.execute(
        text("""
            INSERT INTO annual_tax_payments
            (vehicle_id, tax_year, due_date, paid_date, amount, note, created_at, updated_at)
            VALUES
            (:vehicle_id, :tax_year, :due_date, NULL, 0, :note, :now, :now)
        """),
        {"vehicle_id": vehicle_id, "tax_year": tax_year, "due_date": due_date, "note": note, "now": now},
    )
    return None


def _annual_due_years_for_cycle(main_due_date):
    if not main_due_date:
        return []
    # Siklus pajak tahunan adalah 4 tahun sebelum jatuh tempo 5 tahunan.
    # Contoh: Exp 5 Tahunan 2031, maka siklus tahunannya 2027, 2028, 2029, 2030.
    start_year = five_year_due_to_annual_due(main_due_date).year
    end_year = main_due_date.year - 1
    
    # Jika start_year masih di masa depan yang jauh, kita sesuaikan agar user bisa melihat tahun sekarang
    today_year = date.today().year
    if start_year > today_year + 1:
        # Jika selisihnya terlalu jauh, kemungkinan data anchor salah atau butuh penyesuaian
        pass

    if end_year < start_year:
        return []
    return list(range(start_year, end_year + 1))


def _annual_tax_cycle(vehicle_id:int, main_due_date, annual_map:dict):
    if not main_due_date:
        return {
            "current_due_date": None,
            "display_due_date": None,
            "display_paid_date": None,
            "display_amount": 0,
            "display_note": None,
            "display_status": "empty",
            "tax_year": None,
            "current_payment": None,
        }

    today = date.today()
    valid_years = _annual_due_years_for_cycle(main_due_date)
    if not valid_years:
        return {
            "current_due_date": None,
            "display_due_date": None,
            "display_paid_date": None,
            "display_amount": 0,
            "display_note": None,
            "display_status": "empty",
            "tax_year": None,
            "current_payment": None,
        }

    candidate_year = min(max(today.year, valid_years[0]), valid_years[-1])
    unpaid_year = None
    for year in valid_years:
        payment = annual_map.get((vehicle_id, year))
        if not _is_paid_complete(
            payment.get("paid_date") if payment else None,
            payment.get("amount") if payment else None,
        ):
            unpaid_year = year
            break

    display_year = unpaid_year if unpaid_year is not None else candidate_year
    display_due = get_annual_due(main_due_date, display_year)
    current_payment = annual_map.get((vehicle_id, display_year)) if display_due else None
    current_paid_complete = _is_paid_complete(
        current_payment.get("paid_date") if current_payment else None,
        current_payment.get("amount") if current_payment else None,
    )

    if current_paid_complete:
        if unpaid_year is None:
            return {
                "current_due_date": display_due,
                "display_due_date": display_due,
                "display_paid_date": current_payment.get("paid_date") if current_payment else None,
                "display_amount": (current_payment.get("amount") if current_payment else 0) or 0,
                "display_note": current_payment.get("note") if current_payment else None,
                "display_status": "paid",
                "tax_year": display_year,
                "current_payment": current_payment,
            }

        next_due = get_annual_due(main_due_date, unpaid_year)
        return {
            "current_due_date": next_due,
            "display_due_date": next_due,
            "display_paid_date": None,
            "display_amount": 0,
            "display_note": None,
            "display_status": _tax_status(next_due),
            "tax_year": unpaid_year,
            "current_payment": annual_map.get((vehicle_id, unpaid_year)),
        }

    return {
        "current_due_date": display_due,
        "display_due_date": display_due,
        "display_paid_date": current_payment.get("paid_date") if current_payment else None,
        "display_amount": (current_payment.get("amount") if current_payment else 0) or 0,
        "display_note": current_payment.get("note") if current_payment else None,
        "display_status": _tax_status(display_due),
        "tax_year": display_year,
        "current_payment": current_payment,
    }


def _tax_status(main_due_date):
    if not main_due_date:
        return "empty"
    days = (main_due_date - date.today()).days
    if days < 0:
        return "overdue"
    if days <= DEFAULT_REMINDER_DAYS:
        return "soon"
    return "safe"


def _ensure_annual_tax_table():
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS annual_tax_payments (
            id INT NOT NULL AUTO_INCREMENT,
            vehicle_id INT NOT NULL,
            tax_year INT NOT NULL,
            due_date DATE DEFAULT NULL,
            paid_date DATETIME DEFAULT NULL,
            amount DECIMAL(18,2) DEFAULT 0,
            note TEXT,
            created_at DATETIME NOT NULL,
            updated_at DATETIME NOT NULL,
            PRIMARY KEY (id),
            KEY idx_annual_tax_vehicle_id (vehicle_id),
            KEY idx_annual_tax_vehicle_year (vehicle_id, tax_year),
            CONSTRAINT annual_tax_payments_ibfk_1
                FOREIGN KEY (vehicle_id) REFERENCES vehicles (id)
                ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
    """))

    try:
        db.session.execute(text("""
            ALTER TABLE annual_tax_payments
            ADD COLUMN tax_year INT NOT NULL DEFAULT 0 AFTER vehicle_id
        """))
    except Exception:
        pass

    try:
        db.session.execute(text("""
            ALTER TABLE annual_tax_payments
            MODIFY COLUMN due_date DATE NULL,
            MODIFY COLUMN paid_date DATETIME NULL,
            MODIFY COLUMN amount DECIMAL(18,2) DEFAULT 0,
            MODIFY COLUMN note TEXT NULL
        """))
    except Exception:
        pass

    try:
        db.session.execute(text("""
            CREATE INDEX idx_annual_tax_vehicle_year
            ON annual_tax_payments (vehicle_id, tax_year)
        """))
    except Exception:
        pass

    # Upgrade paid_date ke DATETIME agar jam tersimpan real-time
    try:
        db.session.execute(text("""
            ALTER TABLE annual_tax_payments
            MODIFY COLUMN paid_date DATETIME NULL
        """))
    except Exception:
        pass

    try:
        db.session.execute(text("""
            ALTER TABLE five_year_tax_payments
            MODIFY COLUMN paid_date DATETIME NULL
        """))
    except Exception:
        pass

    db.session.commit()


def _latest_five_tax_map(vehicle_ids:list[int]):
    rows={}
    if not vehicle_ids:return rows
    five_rows=(
        db.session.execute(
            text("""
                SELECT id, vehicle_id, due_date, paid_date, amount, status, plate_before, plate_after, note
                FROM five_year_tax_payments
                WHERE vehicle_id IN :ids
                ORDER BY vehicle_id ASC,
                         CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                         due_date DESC,
                         id DESC
            """).bindparams(bindparam("ids", expanding=True)),
            {"ids": vehicle_ids},
        )
        .mappings()
        .all()
    )
    for r in five_rows:
        if r["vehicle_id"] not in rows:rows[r["vehicle_id"]]=r
    return rows


def _annual_paid_count(vehicle_id:int)->int:
    _ensure_annual_tax_table()
    if not vehicle_id:
        return 0
    row = db.session.execute(
        text("""
            SELECT COUNT(*) AS total
            FROM annual_tax_payments
            WHERE vehicle_id = :vehicle_id
              AND paid_date IS NOT NULL
              AND amount IS NOT NULL
              AND amount > 0
        """),
        {"vehicle_id": vehicle_id},
    ).mappings().first()
    return int((row or {}).get("total") or 0)


def _annual_edit_locked(vehicle_id:int)->bool:
    return _annual_paid_count(vehicle_id) >= 1


def _latest_annual_row(vehicle_id:int):
    _ensure_annual_tax_table()
    if not vehicle_id:
        return None
    return db.session.execute(
        text("""
            SELECT id, vehicle_id, tax_year, due_date, paid_date, amount, note
            FROM annual_tax_payments
            WHERE vehicle_id = :vehicle_id
            ORDER BY COALESCE(paid_date, due_date) DESC, id DESC
            LIMIT 1
        """),
        {"vehicle_id": vehicle_id},
    ).mappings().first()


def _annual_payment_map(vehicle_ids:list[int]):
    _ensure_annual_tax_table()
    rows={}
    if not vehicle_ids:return rows
    annual_rows=(
        db.session.execute(
            text("""
                SELECT id, vehicle_id, tax_year, due_date, paid_date, amount, note
                FROM annual_tax_payments
                WHERE vehicle_id IN :ids
                ORDER BY vehicle_id ASC, tax_year DESC, id DESC
            """).bindparams(bindparam("ids", expanding=True)),
            {"ids": vehicle_ids},
        )
        .mappings()
        .all()
    )
    for r in annual_rows:
        rows[(r["vehicle_id"],r["tax_year"])]=r
    return rows


def _latest_paid_five_year_row(vehicle_id:int):
    if not vehicle_id:
        return None
    return db.session.execute(
        text("""
            SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note
            FROM five_year_tax_payments
            WHERE vehicle_id = :vehicle_id
              AND paid_date IS NOT NULL
              AND amount IS NOT NULL
              AND amount > 0
            ORDER BY paid_date DESC, id DESC
            LIMIT 1
        """),
        {"vehicle_id": vehicle_id},
    ).mappings().first()


def _pending_plate_update_row(vehicle_id:int):
    row = _latest_paid_five_year_row(vehicle_id)
    if not row:
        return None
    plate_after = _normalize_text(row.get("plate_after"))
    if plate_after:
        return None
    return row


def _sync_vehicle_plate_update(vehicle:Vehicle):
    pending = _pending_plate_update_row(vehicle.id)
    if not pending:
        return False
    new_plate = _normalize_text(getattr(vehicle, "plate_new", None))
    plate_before = _normalize_text(pending.get("plate_before"))
    if not new_plate or (plate_before and new_plate == plate_before):
        return False
    db.session.execute(
        text("""
            UPDATE five_year_tax_payments
            SET plate_after = :plate_after,
                updated_at = :now
            WHERE id = :id
        """),
        {"id": pending["id"], "plate_after": new_plate, "now": _now_naive()},
    )
    return True


def _tax_alert_items(q:str="",company:str=""):
    vehicles=(
        _vehicle_query_filtered(q,company)
        .order_by(Vehicle.updated_at.desc(),Vehicle.id.desc())
        .limit(1000)
        .all()
    )
    vehicle_ids=[v.id for v in vehicles]
    five_map=_latest_five_tax_map(vehicle_ids)
    annual_map=_annual_payment_map(vehicle_ids)
    items=[]
    for v in vehicles:
        five=five_map.get(v.id)
        main_due=five["due_date"] if five else None
        if not main_due:
            continue

        annual_info=_annual_tax_cycle(v.id, main_due, annual_map)
        annual_due=annual_info["current_due_date"]
        annual_paid_complete = _is_paid_complete(
            annual_info["current_payment"].get("paid_date") if annual_info["current_payment"] else None,
            annual_info["current_payment"].get("amount") if annual_info["current_payment"] else None,
        )
        annual_status=_tax_status(annual_due) if annual_due and not annual_paid_complete else None
        if annual_due and not annual_paid_complete and annual_status in {"soon","overdue"}:
            items.append({
                "kind":"annual","kind_label":"Pajak Tahunan","vehicle":v,"vehicle_id":v.id,
                "due_date":annual_due,"paid_date":None,"amount":0,"note":None,
                "status_tax":annual_status,"plate_before":getattr(v,"plate_old",None),
                "plate_after":getattr(v,"plate_new",None),"tax_year":annual_due.year,
            })

        five_status=_tax_status(main_due) if five and not five.get("paid_date") else None
        if five and not five.get("paid_date") and five_status in {"soon","overdue"}:
            items.append({
                "kind":"five","kind_label":"Pajak 5 Tahunan","vehicle":v,"vehicle_id":v.id,
                "record_id":five["id"],"due_date":main_due,"paid_date":five["paid_date"],
                "amount":five["amount"] or 0,"note":five["note"],"status_tax":five_status,
                "plate_before":five.get("plate_before") or getattr(v,"plate_old",None),
                "plate_after":five.get("plate_after") or getattr(v,"plate_new",None),
            })
    items.sort(key=lambda x:(x["due_date"] or date.max,x["kind"]!="annual",vehicle_plate(x["vehicle"])))
    return items

@bp.get("/api/vehicles/dropdown")
def api_vehicles_dropdown():
    """API endpoint untuk dropdown kendaraan dengan info detail + history."""
    q = request.args.get('q', '').strip()
    limit = min(max(_parse_int(request.args.get('limit')) or 50, 1), 100)

    query = _apply_vehicle_scope(Vehicle.query.filter_by(is_deleted=False))

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Vehicle.pt.ilike(like),
            Vehicle.merk.ilike(like),
            Vehicle.type.ilike(like),
            Vehicle.jenis.ilike(like),
            Vehicle.plate_new.ilike(like),
            Vehicle.plate_old.ilike(like),
            Vehicle.active_name.ilike(like),
            Vehicle.name_as_asset_pt.ilike(like),
            Vehicle.new_asset_name.ilike(like),
            Vehicle.user_old.ilike(like),
            Vehicle.user_new.ilike(like),
            Vehicle.status.ilike(like),
            Vehicle.lokasi.ilike(like),
        ))

    vehicles = query.order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).limit(limit).all()

    result = []
    for v in vehicles:
        vehicle_history_count = VehicleChangeHistory.query.filter_by(vehicle_id=v.id).count()
        user_history_count = UserHistory.query.filter_by(vehicle_id=v.id).count()
        loan_history_count = 0
        if hasattr(LoanHistory, "vehicle_id"):
            loan_history_count = LoanHistory.query.filter_by(vehicle_id=v.id).count()

        result.append({
            'id': v.id,
            'label': vehicle_display_name(v),
            'pt': v.pt or '-',
            'asset_owner_company_name': _company_name(getattr(v, 'asset_owner_company', None)) or '-',
            'pt_pemakai_1_company_name': _company_name(getattr(v, 'pt_pemakai_1_company', None)) or '-',
            'pt_pemakai_2_company_name': _company_name(getattr(v, 'pt_pemakai_2_company', None)) or '-',
            'merk': v.merk or '-',
            'type': v.type or '-',
            'plate': vehicle_plate(v),
            'jenis': v.jenis or '-',
            'status': v.status or '-',
            'user': v.user_new or v.user_old or '-',
            'kondisi': v.kondisi_terkini or '-',
            'lokasi': v.lokasi or '-',
            'history_count': vehicle_history_count + user_history_count + loan_history_count,
            'updated_at': v.updated_at.isoformat() if v.updated_at else None,
        })

    return jsonify({'success': True, 'data': result})


@bp.get("/api/companies/dropdown")
def api_companies_dropdown():
    """API endpoint untuk dropdown PT berdasarkan field Vehicle.pt, bukan master company relasi."""
    q = request.args.get('q', '').strip()
    limit = min(max(_parse_int(request.args.get('limit')) or 50, 1), 100)

    query = _apply_vehicle_scope(Vehicle.query.filter(Vehicle.is_deleted == False, Vehicle.pt.isnot(None), func.trim(Vehicle.pt) != ""))
    if q:
        like = f"%{q}%"
        query = query.filter(Vehicle.pt.ilike(like))

    pt_rows = (
        query.with_entities(func.trim(Vehicle.pt).label('name'))
        .distinct()
        .order_by(func.trim(Vehicle.pt).asc())
        .limit(limit)
        .all()
    )

    result = []
    for row in pt_rows:
        pt_name = (row[0] or '').strip()
        if not pt_name:
            continue
        vehicles = Vehicle.query.filter(
            Vehicle.is_deleted == False,
            func.lower(func.trim(Vehicle.pt)) == pt_name.lower()
        ).order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).all()
        vehicle_ids = [v.id for v in vehicles]
        total_history = 0
        if vehicle_ids:
            total_history += VehicleChangeHistory.query.filter(VehicleChangeHistory.vehicle_id.in_(vehicle_ids)).count()
            total_history += UserHistory.query.filter(UserHistory.vehicle_id.in_(vehicle_ids)).count()
            if hasattr(LoanHistory, "vehicle_id"):
                total_history += LoanHistory.query.filter(LoanHistory.vehicle_id.in_(vehicle_ids)).count()
        last_updated = vehicles[0].updated_at.isoformat() if vehicles and getattr(vehicles[0], 'updated_at', None) else None
        result.append({
            'id': pt_name,
            'name': pt_name,
            'vehicle_count': len(vehicles),
            'history_count': total_history,
            'active_vehicle_count': sum(1 for v in vehicles if (v.status or '').strip()),
            'last_updated': last_updated,
        })

    return jsonify({'success': True, 'data': result})

@bp.get("/data/pajak", endpoint="go_data_pajak")
def go_data_pajak():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip().lower()
    company = _sanitize_company_filter((request.args.get("company") or "").strip())
    companies = _scoped_company_choices()

    vehicles = (
        _vehicle_query_filtered(q, company)
        .order_by(Vehicle.updated_at.desc(), Vehicle.id.desc())
        .limit(1000)
        .all()
    )
    vehicle_ids = [v.id for v in vehicles]
    five_map = _latest_five_tax_map(vehicle_ids)
    annual_map = _annual_payment_map(vehicle_ids)
    today = date.today()

    def kendaraan_nama(v):
        parts = [
            getattr(v, "new_asset_name", None) or getattr(v, "active_name", None) or getattr(v, "name_as_asset_pt", None),
            getattr(v, "merk", None),
            getattr(v, "type", None),
            getattr(v, "jenis", None),
        ]
        txt = " / ".join([str(x).strip() for x in parts if x and str(x).strip()])
        return txt or vehicle_display_name(v)

    items = []
    for v in vehicles:
        f = five_map.get(v.id)
        raw_main_due = f["due_date"] if f else None
        raw_five_paid_date = f.get("paid_date") if f else None
        raw_five_is_paid = _is_paid_complete(raw_five_paid_date, (f.get("amount") if f else None))

        # Jika record 5 tahunan terakhir sudah lunas, tabel utama tidak boleh berhenti di
        # status "Sudah dibayar". Data Pajak harus menampilkan siklus aktif berikutnya:
        # 5 tahunan +5 tahun, lalu tahunan mulai lagi dari tahun pertama setelah 5 tahunan lama.
        if raw_main_due and raw_five_is_paid:
            main_due = _shift_years_safe(raw_main_due, 5)
            five_paid_date = None
            five_is_paid = False
            five_amount = 0
            five_note = None
            plate_before_value = getattr(v, "plate_new", None) or getattr(v, "plate_old", None)
            plate_after_value = getattr(v, "plate_new", None)
        else:
            main_due = raw_main_due
            five_paid_date = raw_five_paid_date
            five_is_paid = raw_five_is_paid
            five_amount = (f.get("amount") if f else 0) or 0
            five_note = f.get("note") if f else None
            plate_before_value = (f.get("plate_before") if f else None) or getattr(v, "plate_old", None)
            plate_after_value = (f.get("plate_after") if f else None) or getattr(v, "plate_new", None)

        annual_info = _annual_tax_cycle(v.id, main_due, annual_map)
        annual_due_date = annual_info["display_due_date"]
        annual_anchor_due_date = five_year_due_to_annual_due(main_due) if main_due else None
        annual_status = annual_info["display_status"] if annual_due_date else "empty"

        # Aksi aktif tidak menunggu H-berapa.
        # Kalau tanggal pajak tahunan aktif ada dan belum lunas, tombol Bayar aktif.
        # Setelah tahunan dibayar, _annual_tax_cycle otomatis geser ke tahun berikutnya.
        # Setelah semua tahunan dalam siklus selesai, aksi pindah ke pajak 5 tahunan.
        if not main_due and not annual_due_date:
            action_kind = "setup"
        elif annual_due_date and annual_status != "paid":
            action_kind = "annual"
        elif main_due and not five_is_paid:
            action_kind = "five"
        else:
            action_kind = "setup"

        # Tombol Bayar tetap tampil. Disabled hanya kalau tidak ada tanggal aktif.
        active_due_for_pay = annual_due_date if action_kind in {"annual", "setup"} else main_due
        can_pay = bool(active_due_for_pay)

        items.append({
            "no": 0,
            "record_id": f["id"] if f else None,
            "vehicle_id": v.id,
            "vehicle": v,
            "pt": v.pt or "-",
            "kendaraan": kendaraan_nama(v),
            "no_polisi": vehicle_plate(v),
            "annual_due_date": annual_due_date,
            "annual_paid_date": annual_info["display_paid_date"],
            "annual_anchor_due_date": annual_anchor_due_date,
            "annual_status": annual_status,
            "annual_amount": annual_info["display_amount"],
            "annual_note": annual_info["display_note"],
            "annual_tax_year": annual_info["tax_year"],
            "five_due_date": main_due,
            "five_paid_date": five_paid_date,
            "five_amount": five_amount,
            "plate_before": plate_before_value,
            "plate_after": plate_after_value,
            "plate_warning": (
                main_due is not None and five_is_paid and
                not plate_after_value and
                not getattr(v, "plate_new", None)
            ),
            "note": (annual_info["display_note"] if action_kind == "annual" and annual_info["display_note"] else five_note),
            "five_status": "empty" if not main_due else ("paid" if five_is_paid else _tax_status(main_due)),
            "is_five_paid": five_is_paid,
            "action_kind": action_kind,
            "can_pay": can_pay,
        })

    if status:
        items = [x for x in items if status in {str(x.get("annual_status") or "").lower(), str(x.get("five_status") or "").lower()}]

    items.sort(
        key=lambda x: ((x["five_due_date"] or x["annual_due_date"] or date.min), x["vehicle_id"]),
        reverse=True
    )
    for i, r in enumerate(items, 1):
        r["no"] = i

    highlight_id = _parse_int(request.args.get("highlight"))
    return render_template(
        "data_pajak.html",
        items=items,
        q=q,
        status=status,
        company=company,
        companies=companies,
        highlight_id=highlight_id,
        title="Data Pajak / Pembayaran - Sinar Group",
    )


@bp.post("/data/pajak/tambah")
@master_required
def data_pajak_tambah():
    _ensure_annual_tax_table()
    vehicle_id = _parse_int(request.form.get("vehicle_id"))
    v = Vehicle.query.get_or_404(vehicle_id)

    payment_kind = (_normalize_text(request.form.get("payment_kind")) or "setup").lower()
    annual_due_date = _parse_date(request.form.get("annual_due_date"))
    # Pajak 5 tahunan (due_date) harus mengikuti pajak tahunan (annual_due_date)
    due_date = annual_due_to_five_year_due(annual_due_date)
    paid_date = _parse_date_as_datetime(request.form.get("paid_date"))
    amount = _parse_float(request.form.get("amount"))
    plate_before = _normalize_text(request.form.get("plate_before")) or None
    plate_after = _normalize_text(request.form.get("plate_after")) or None
    note = _normalize_text(request.form.get("note")) or None

    if not annual_due_date:
        flash("Exp Pajak Tahunan wajib diisi.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    if (paid_date and amount is None) or (amount is not None and not paid_date):
        flash("Isi tanggal bayar dan nominal bayar supaya status menjadi lunas.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    if amount is not None and amount <= 0:
        flash("Nominal bayar harus lebih dari 0.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    existing = db.session.execute(
        text("""
            SELECT id, due_date, paid_date, amount
            FROM five_year_tax_payments
            WHERE vehicle_id = :vehicle_id
            ORDER BY CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                     due_date DESC,
                     id DESC
            LIMIT 1
        """),
        {"vehicle_id": vehicle_id},
    ).mappings().first()

    if existing:
        flash("Data pajak untuk kendaraan ini sudah ada. Silakan gunakan tombol aksi pada baris kendaraan.", "warning")
        return redirect(url_for("main.go_data_pajak"))

    db.session.execute(
        text("""
            INSERT INTO five_year_tax_payments
            (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
            VALUES
            (:vehicle_id, :due_date, NULL, 0, :plate_before, :plate_after, :note, :now, :now)
        """),
        {
            "vehicle_id": vehicle_id,
            "due_date": due_date,
            "plate_before": plate_before,
            "plate_after": plate_after,
            "note": note,
            "now": _now_naive(),
        },
    )

    if payment_kind in {"annual", "setup"} and paid_date and amount is not None:
        tax_year = annual_due_date.year
        db.session.execute(
            text("""
                INSERT INTO annual_tax_payments
                (vehicle_id, tax_year, due_date, paid_date, amount, note, created_at, updated_at)
                VALUES
                (:vehicle_id, :tax_year, :due_date, :paid_date, :amount, :note, :now, :now)
            """),
            {
                "vehicle_id": vehicle_id,
                "tax_year": tax_year,
                "due_date": annual_due_date,
                "paid_date": paid_date,
                "amount": amount,
                "note": note,
                "now": _now_naive(),
            },
        )

    try:
        audit("CREATE", "TaxPayment", 0, note=f"Tambah pajak kendaraan {vehicle_plate(v)}")
    except Exception:
        pass

    try:
        db.session.commit()
        flash("Exp date berhasil disimpan. Jenis pembayaran akan terdeteksi otomatis.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Gagal tambah pajak: {exc}", "danger")

    return redirect(url_for("main.go_data_pajak"))


@bp.post("/data/pajak/<int:record_id>/edit")
@master_required
def data_pajak_edit(record_id: int):
    _ensure_annual_tax_table()
    payment_kind = (_normalize_text(request.form.get("payment_kind")) or "annual").lower()
    annual_due_date = _parse_date(request.form.get("annual_due_date"))
    # Pajak 5 tahunan (due_date) harus mengikuti pajak tahunan (annual_due_date)
    due_date = annual_due_to_five_year_due(annual_due_date)
    paid_date = _parse_date_as_datetime(request.form.get("paid_date"))
    amount = _parse_float(request.form.get("amount"))
    plate_before = _normalize_text(request.form.get("plate_before")) or None
    plate_after = _normalize_text(request.form.get("plate_after")) or None
    note = _normalize_text(request.form.get("note")) or None

    if not annual_due_date:
        flash("Exp Pajak Tahunan wajib diisi.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    if (paid_date and amount is None) or (amount is not None and not paid_date):
        flash("Isi tanggal bayar dan nominal bayar supaya status menjadi lunas.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    if amount is not None and amount <= 0:
        flash("Nominal bayar harus lebih dari 0.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    current_five = db.session.execute(
        text("""
            SELECT id, vehicle_id, due_date, paid_date, amount, note, plate_before, plate_after
            FROM five_year_tax_payments
            WHERE id = :id
            LIMIT 1
        """),
        {"id": record_id},
    ).mappings().first()
    if not current_five:
        flash("Data pajak tidak ditemukan.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    now = _now_naive()
    annual_locked = payment_kind == "annual" and _annual_edit_locked(current_five["vehicle_id"])

    # Kalau sudah pernah dibayar minimal 1x, tanggal tidak boleh diubah lagi.
    # Yang boleh hanya catatan.
    if _annual_paid_count(current_five["vehicle_id"]) >= 1:
        # Cek apakah user mencoba mengubah tanggal
        date_changed = False
        if payment_kind == "annual":
            latest_annual = _latest_annual_row(current_five["vehicle_id"])
            # Bandingkan tanggal, pastikan menangani objek date vs string jika perlu
            curr_due = latest_annual.get("due_date") if latest_annual else None
            if curr_due and annual_due_date and curr_due != annual_due_date:
                date_changed = True
        elif payment_kind == "five":
            curr_five_due = current_five.get("due_date")
            if curr_five_due and due_date and curr_five_due != due_date:
                date_changed = True

        if date_changed:
            # Update hanya catatan saja
            db.session.execute(
                text("""
                    UPDATE five_year_tax_payments
                    SET note = :note,
                        updated_at = :now
                    WHERE id = :id
                """),
                {"id": record_id, "note": note, "now": now},
            )
            latest_annual = _latest_annual_row(current_five["vehicle_id"])
            if latest_annual:
                db.session.execute(
                    text("""
                        UPDATE annual_tax_payments
                        SET note = :note,
                            updated_at = :now
                        WHERE id = :id
                    """),
                    {"id": latest_annual["id"], "note": note, "now": now},
                )
            
            try:
                audit("UPDATE", "TaxPayment", record_id, note=f"Update catatan pajak (tanggal dikunci karena sudah ada pembayaran) untuk {vehicle_plate(Vehicle.query.get(current_five['vehicle_id']))}")
            except Exception: pass
            
            db.session.commit()
            flash("Sudah ada pembayaran, tanggal tidak bisa diubah. Hanya catatan yang diperbarui.", "warning")
            return redirect(url_for("main.go_data_pajak"))

    # Selalu update due_date di five_year_tax_payments agar sinkron dengan annual_due_date
    db.session.execute(
        text("""
            UPDATE five_year_tax_payments
            SET due_date = :due_date,
                plate_before = :plate_before,
                plate_after = :plate_after,
                note = :note,
                updated_at = :now
            WHERE id = :id
        """),
        {
            "id": record_id,
            "due_date": due_date,
            "plate_before": plate_before,
            "plate_after": plate_after,
            "note": note,
            "now": now,
        },
    )

    if payment_kind == "annual":
        # Selalu update atau insert record di annual_tax_payments agar sinkron
        tax_year = annual_due_date.year
        annual_row = db.session.execute(
            text("""
                SELECT id FROM annual_tax_payments
                WHERE vehicle_id = :vehicle_id AND tax_year = :tax_year
                ORDER BY id DESC
                LIMIT 1
            """),
            {"vehicle_id": current_five["vehicle_id"], "tax_year": tax_year},
        ).mappings().first()
        
        if annual_row:
            db.session.execute(
                text("""
                    UPDATE annual_tax_payments
                    SET due_date = :due_date,
                        paid_date = :paid_date,
                        amount = :amount,
                        note = :note,
                        updated_at = :now
                    WHERE id = :id
                """),
                {
                    "id": annual_row["id"],
                    "due_date": annual_due_date,
                    "paid_date": paid_date, # Bisa None jika belum bayar
                    "amount": amount or 0,
                    "note": note,
                    "now": now,
                },
            )
        else:
            db.session.execute(
                text("""
                    INSERT INTO annual_tax_payments
                    (vehicle_id, tax_year, due_date, paid_date, amount, note, created_at, updated_at)
                    VALUES
                    (:vehicle_id, :tax_year, :due_date, :paid_date, :amount, :note, :now, :now)
                """),
                {
                    "vehicle_id": current_five["vehicle_id"],
                    "tax_year": tax_year,
                    "due_date": annual_due_date,
                    "paid_date": paid_date,
                    "amount": amount or 0,
                    "note": note,
                    "now": now,
                },
            )
    elif payment_kind == "five":
        db.session.execute(
            text("""
                UPDATE five_year_tax_payments
                SET due_date = :due_date,
                    paid_date = :paid_date,
                    amount = :amount,
                    plate_before = :plate_before,
                    plate_after = :plate_after,
                    note = :note,
                    updated_at = :now
                WHERE id = :id
            """),
            {
                "id": record_id,
                "due_date": due_date,
                "paid_date": paid_date,
                "amount": amount or 0,
                "plate_before": plate_before,
                "plate_after": plate_after,
                "note": note,
                "now": now,
            },
        )

        if _is_paid_complete(paid_date, amount):
            next_due_date = _shift_years_safe(due_date, 5)
            next_row = db.session.execute(
                text("""
                    SELECT id
                    FROM five_year_tax_payments
                    WHERE vehicle_id = :vehicle_id
                      AND due_date = :due_date
                    ORDER BY id DESC
                    LIMIT 1
                """),
                {"vehicle_id": current_five["vehicle_id"], "due_date": next_due_date},
            ).mappings().first()
            if not next_row:
                db.session.execute(
                    text("""
                        INSERT INTO five_year_tax_payments
                        (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                        VALUES
                        (:vehicle_id, :due_date, NULL, 0, :plate_before, :plate_after, NULL, :now, :now)
                    """),
                    {
                        "vehicle_id": current_five["vehicle_id"],
                        "due_date": next_due_date,
                        "plate_before": plate_before or current_five.get("plate_before"),
                        "plate_after": plate_after or current_five.get("plate_after"),
                        "now": now,
                    },
                )

    try:
        audit("UPDATE", "TaxPayment", record_id, note=f"Edit pajak id={record_id}")
    except Exception:
        pass

    try:
        db.session.commit()
        flash("Data pajak berhasil diupdate.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Gagal update pajak: {exc}", "danger")

    return redirect(url_for("main.go_data_pajak"))

def _ensure_payment_history_table():
    """Memastikan tabel payment_history sudah ada di database."""
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS payment_history (
            id INT NOT NULL AUTO_INCREMENT,
            vehicle_id INT NOT NULL,
            payment_type ENUM('annual', 'five_year') NOT NULL,
            due_date DATE DEFAULT NULL,
            paid_date DATETIME DEFAULT NULL,
            amount DECIMAL(18,2) DEFAULT 0,
            plate_before VARCHAR(20) DEFAULT NULL,
            plate_after VARCHAR(20) DEFAULT NULL,
            note TEXT,
            created_at DATETIME NOT NULL,
            updated_at DATETIME NOT NULL,
            PRIMARY KEY (id),
            KEY idx_payment_vehicle_id (vehicle_id),
            KEY idx_payment_type (payment_type),
            KEY idx_payment_date (paid_date),
            CONSTRAINT payment_history_ibfk_1
                FOREIGN KEY (vehicle_id) REFERENCES vehicles (id)
                ON DELETE CASCADE
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_0900_ai_ci
    """))
    
    try:
        db.session.execute(text("""
            ALTER TABLE payment_history
            ADD COLUMN plate_before VARCHAR(20) DEFAULT NULL AFTER amount
        """))
    except Exception:
        pass
    
    try:
        db.session.execute(text("""
            ALTER TABLE payment_history
            ADD COLUMN plate_after VARCHAR(20) DEFAULT NULL AFTER plate_before
        """))
    except Exception:
        pass
    
    db.session.commit()

@bp.post("/api/pajak/claim")
@master_required
def api_pajak_claim():
    """
    API endpoint AJAX untuk pembayaran pajak.

    Logika pajak:
    - Pajak tahunan dibayar untuk tahun ke-1 sampai ke-4 dalam satu siklus.
    - Pajak 5 tahunan dibayar di tahun ke-5, lalu due 5 tahunan berikutnya rolling +5 tahun.
    - Saat bayar pajak 5 tahunan, plat lama otomatis diambil dari plat aktif kendaraan
      (NO POLISI BARU kalau ada, fallback NO POLISI LAMA). Plat baru diambil dari input
      Data Pajak, lalu langsung disinkronkan ke Data Kendaraan:
        Vehicle.plate_old = plat lama
        Vehicle.plate_new = plat baru
    """
    _ensure_annual_tax_table()
    _ensure_payment_history_table()

    vehicle_id = _parse_int(request.form.get("vehicle_id"))
    kind = (_normalize_text(request.form.get("kind")) or "annual").lower()
    due_date = _parse_date(request.form.get("due_date"))
    paid_date = _parse_date_as_datetime(request.form.get("paid_date")) or _today_as_datetime()
    amount = _parse_float(request.form.get("amount"))
    note = _normalize_text(request.form.get("note")) or None

    if not vehicle_id:
        return jsonify({"success": False, "error": "vehicle_id wajib diisi"}), 400
    if amount is None or amount <= 0:
        return jsonify({"success": False, "error": "Nominal bayar wajib diisi dan harus lebih dari 0"}), 400
    if not due_date:
        return jsonify({"success": False, "error": "Tanggal exp pajak wajib diisi"}), 400
    if not paid_date:
        return jsonify({"success": False, "error": "Tanggal bayar wajib diisi"}), 400

    v = Vehicle.query.get(vehicle_id)
    if not v:
        return jsonify({"success": False, "error": "Kendaraan tidak ditemukan"}), 404

    now = _now_naive()

    try:
        payment_type = "annual"
        plate_before_for_history = None
        plate_after_for_history = None

        if kind == "five":
            payment_type = "five_year"
            record_id = _parse_int(request.form.get("record_id"))
            if not record_id:
                return jsonify({"success": False, "error": "record_id wajib diisi untuk pajak 5 tahunan"}), 400

            current_five = db.session.execute(
                text("""
                    SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after
                    FROM five_year_tax_payments
                    WHERE id = :id
                    LIMIT 1
                """),
                {"id": record_id},
            ).mappings().first()
            if not current_five:
                return jsonify({"success": False, "error": "Data pajak 5 tahunan tidak ditemukan"}), 404

            # Jika baris yang dikirim dari tabel adalah record 5 tahunan lama yang sudah lunas,
            # tetapi due_date dari form sudah menunjuk siklus berikutnya, buat/pakai record baru.
            # Ini membuat data lama tetap jadi history, sementara pembayaran baru masuk siklus aktif.
            if _is_paid_complete(current_five.get("paid_date"), current_five.get("amount")) and due_date and current_five.get("due_date") != due_date:
                next_five = db.session.execute(
                    text("""
                        SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after
                        FROM five_year_tax_payments
                        WHERE vehicle_id = :vehicle_id AND due_date = :due_date
                        ORDER BY id DESC
                        LIMIT 1
                    """),
                    {"vehicle_id": vehicle_id, "due_date": due_date},
                ).mappings().first()
                if not next_five:
                    active_plate = _normalize_plate_value(getattr(v, "plate_new", None)) or _normalize_plate_value(getattr(v, "plate_old", None)) or None
                    db.session.execute(
                        text("""
                            INSERT INTO five_year_tax_payments
                            (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                            VALUES (:vehicle_id, :due_date, NULL, 0, :plate_before, NULL, NULL, :now, :now)
                        """),
                        {"vehicle_id": vehicle_id, "due_date": due_date, "plate_before": active_plate, "now": now},
                    )
                    db.session.flush()
                    next_five = db.session.execute(
                        text("""
                            SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after
                            FROM five_year_tax_payments
                            WHERE vehicle_id = :vehicle_id AND due_date = :due_date
                            ORDER BY id DESC
                            LIMIT 1
                        """),
                        {"vehicle_id": vehicle_id, "due_date": due_date},
                    ).mappings().first()
                current_five = next_five
                record_id = current_five["id"]

            _ensure_unpaid_tax_payment(current_five, "Pajak 5 tahunan")

            target_due_date = due_date or current_five["due_date"]

            stored_old_plate = _normalize_plate_value(getattr(v, "plate_old", None))
            stored_new_plate = _normalize_plate_value(getattr(v, "plate_new", None))
            current_active_plate = stored_new_plate or stored_old_plate

            submitted_before = _normalize_plate_value(request.form.get("plate_before"))
            submitted_after = _normalize_plate_value(request.form.get("plate_after"))

            # Jika form masih mengirim NO POLISI LAMA lama dari Data Kendaraan,
            # sementara kendaraan sudah punya NO POLISI BARU sebagai plat aktif,
            # maka plat lama pembayaran 5 tahunan harus mengikuti plat aktif.
            if stored_new_plate and (not submitted_before or submitted_before == stored_old_plate):
                plate_before_for_history = stored_new_plate
            else:
                plate_before_for_history = (
                    submitted_before
                    or current_active_plate
                    or _normalize_plate_value(current_five.get("plate_before"))
                    or None
                )

            plate_after_for_history = (
                submitted_after
                or _normalize_plate_value(current_five.get("plate_after"))
                or None
            )

            if not plate_before_for_history or not plate_after_for_history:
                return jsonify({"success": False, "error": "Plat lama dan plat baru wajib diisi untuk pajak 5 tahunan"}), 400

            db.session.execute(
                text("""
                    UPDATE five_year_tax_payments
                    SET due_date = :due_date,
                        paid_date = :paid_date,
                        amount = :amount,
                        plate_before = :plate_before,
                        plate_after = :plate_after,
                        note = :note,
                        updated_at = :now
                    WHERE id = :id
                """),
                {
                    "id": record_id,
                    "due_date": target_due_date,
                    "paid_date": paid_date,
                    "amount": amount,
                    "plate_before": plate_before_for_history,
                    "plate_after": plate_after_for_history,
                    "note": note,
                    "now": now,
                },
            )

            if _is_paid_complete(paid_date, amount):
                # Sinkronkan ganti plat dari Data Pajak ke Data Kendaraan.
                if plate_after_for_history:
                    before_snapshot = _vehicle_snapshot(v)
                    v.plate_old = plate_before_for_history or current_active_plate or v.plate_old
                    v.plate_new = plate_after_for_history
                    v.updated_at = now
                    db.session.flush()
                    after_snapshot = _vehicle_snapshot(v)
                    _log_vehicle_edit(v, before_snapshot, after_snapshot, changed_by=session.get("user_name") or "SYSTEM")

                # Rolling pajak 5 tahunan: setelah tahun ke-5 lunas, buat jadwal berikutnya +5 tahun.
                next_due = _shift_years_safe(target_due_date, 5)
                next_row = db.session.execute(
                    text("""
                        SELECT id
                        FROM five_year_tax_payments
                        WHERE vehicle_id = :vehicle_id AND due_date = :due_date
                        ORDER BY id DESC
                        LIMIT 1
                    """),
                    {"vehicle_id": vehicle_id, "due_date": next_due},
                ).mappings().first()
                if not next_row:
                    db.session.execute(
                        text("""
                            INSERT INTO five_year_tax_payments
                            (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                            VALUES
                            (:vehicle_id, :due_date, NULL, 0, :plate_before, NULL, NULL, :now, :now)
                        """),
                        {
                            "vehicle_id": vehicle_id,
                            "due_date": next_due,
                            "plate_before": plate_after_for_history or plate_before_for_history,
                            "now": now,
                        },
                    )

                # Setelah 5 tahunan lunas, langsung siapkan tahunan pertama siklus berikutnya.
                _create_next_annual_stub(vehicle_id, _shift_years_safe(target_due_date, 1), note=note, now=now)

            try:
                audit("PAY", "FiveYearTax", record_id, note=f"Tandai bayar pajak 5 tahunan {vehicle_plate(v)}")
            except Exception:
                pass

        elif kind == "annual":
            tax_year = due_date.year
            annual_row = db.session.execute(
                text("""
                    SELECT id, paid_date, amount
                    FROM annual_tax_payments
                    WHERE vehicle_id = :vehicle_id AND tax_year = :tax_year
                    ORDER BY id DESC
                    LIMIT 1
                """),
                {"vehicle_id": vehicle_id, "tax_year": tax_year},
            ).mappings().first()

            if annual_row:
                _ensure_unpaid_tax_payment(annual_row, "Pajak tahunan")
                db.session.execute(
                    text("""
                        UPDATE annual_tax_payments
                        SET due_date = :due_date,
                            paid_date = :paid_date,
                            amount = :amount,
                            note = :note,
                            updated_at = :now
                        WHERE id = :id
                    """),
                    {
                        "id": annual_row["id"],
                        "due_date": due_date,
                        "paid_date": paid_date,
                        "amount": amount,
                        "note": note,
                        "now": now,
                    },
                )
            else:
                db.session.execute(
                    text("""
                        INSERT INTO annual_tax_payments
                        (vehicle_id, tax_year, due_date, paid_date, amount, note, created_at, updated_at)
                        VALUES
                        (:vehicle_id, :tax_year, :due_date, :paid_date, :amount, :note, :now, :now)
                    """),
                    {
                        "vehicle_id": vehicle_id,
                        "tax_year": tax_year,
                        "due_date": due_date,
                        "paid_date": paid_date,
                        "amount": amount,
                        "note": note,
                        "now": now,
                    },
                )



            # Rolling tahunan: setelah pajak tahun ini dibayar, siapkan tahun berikutnya
            # selama masih dalam siklus sebelum pajak 5 tahunan.
            active_five = db.session.execute(
                text("""
                    SELECT due_date
                    FROM five_year_tax_payments
                    WHERE vehicle_id = :vehicle_id
                    ORDER BY CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                             due_date DESC,
                             id DESC
                    LIMIT 1
                """),
                {"vehicle_id": vehicle_id},
            ).mappings().first()
            active_five_due = active_five.get("due_date") if active_five else None
            next_annual_due = _shift_years_safe(due_date, 1)
            if active_five_due and next_annual_due and next_annual_due.year < active_five_due.year:
                _create_next_annual_stub(vehicle_id, next_annual_due, note=note, now=now)
            try:
                audit("PAY", "AnnualTax", vehicle_id, note=f"Tandai bayar pajak tahunan {vehicle_plate(v)}")
            except Exception:
                pass
        else:
            return jsonify({"success": False, "error": "Jenis pajak tidak valid"}), 400

        # Simpan ke history pembayaran agar modal detail/export membaca riwayat lengkap.
        db.session.execute(
            text("""
                INSERT INTO payment_history
                (vehicle_id, payment_type, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                VALUES
                (:vehicle_id, :payment_type, :due_date, :paid_date, :amount, :plate_before, :plate_after, :note, :now, :now)
            """),
            {
                "vehicle_id": vehicle_id,
                "payment_type": payment_type,
                "due_date": due_date,
                "paid_date": paid_date,
                "amount": amount,
                "plate_before": plate_before_for_history,
                "plate_after": plate_after_for_history,
                "note": note,
                "now": now,
            },
        )

        db.session.commit()
        return jsonify({"success": True, "message": "Pembayaran berhasil dicatat!"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@bp.post("/data/pajak/claim", endpoint="data_pajak_claim")
@master_required
def data_pajak_claim():
    _ensure_annual_tax_table()
    vehicle_id = _parse_int(request.form.get("vehicle_id"))
    kind = (_normalize_text(request.form.get("kind")) or "annual").lower()
    due_date = _parse_date(request.form.get("due_date"))
    paid_date = _parse_date_as_datetime(request.form.get("paid_date")) or _today_as_datetime()
    amount = _parse_float(request.form.get("amount"))
    note = _normalize_text(request.form.get("note")) or None
    back_to = (_normalize_text(request.form.get("back_to")) or "data_pajak").lower()

    if not vehicle_id:
        flash("Kendaraan tidak ditemukan.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    if amount is None or amount <= 0:
        flash("Nominal bayar wajib diisi dan harus lebih dari 0.", "danger")
        return redirect(url_for("main.go_data_pajak"))

    v = Vehicle.query.get_or_404(vehicle_id)
    now = _now_naive()

    try:
        if kind == "five":
            record_id = _parse_int(request.form.get("record_id"))
            if not record_id:
                flash("Data pajak 5 tahunan tidak ditemukan.", "danger")
                return redirect(url_for("main.go_data_pajak"))

            current_five = db.session.execute(
                text("""
                    SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after
                    FROM five_year_tax_payments
                    WHERE id = :id
                    LIMIT 1
                """),
                {"id": record_id},
            ).mappings().first()
            if not current_five:
                flash("Data pajak 5 tahunan tidak ditemukan.", "danger")
                return redirect(url_for("main.go_data_pajak"))
            _ensure_unpaid_tax_payment(current_five, "Pajak 5 tahunan")

            target_due_date = due_date or current_five["due_date"]

            stored_old_plate = _normalize_plate_value(getattr(v, "plate_old", None))
            stored_new_plate = _normalize_plate_value(getattr(v, "plate_new", None))
            current_active_plate = stored_new_plate or stored_old_plate
            submitted_before = _normalize_plate_value(request.form.get("plate_before"))
            submitted_after = _normalize_plate_value(request.form.get("plate_after"))

            if stored_new_plate and (not submitted_before or submitted_before == stored_old_plate):
                plate_before_for_history = stored_new_plate
            else:
                plate_before_for_history = (
                    submitted_before
                    or current_active_plate
                    or _normalize_plate_value(current_five.get("plate_before"))
                    or None
                )
            plate_after_for_history = (
                submitted_after
                or _normalize_plate_value(current_five.get("plate_after"))
                or None
            )

            if not plate_before_for_history or not plate_after_for_history:
                flash("Plat lama dan plat baru wajib diisi untuk pajak 5 tahunan.", "danger")
                return redirect(url_for("main.go_data_pajak"))

            db.session.execute(
                text("""
                    UPDATE five_year_tax_payments
                    SET due_date = :due_date,
                        paid_date = :paid_date,
                        amount = :amount,
                        note = :note,
                        plate_before = :plate_before,
                        plate_after = :plate_after,
                        updated_at = :now
                    WHERE id = :id
                """),
                {
                    "id": record_id,
                    "due_date": target_due_date,
                    "paid_date": paid_date,
                    "amount": amount,
                    "note": note,
                    "plate_before": plate_before_for_history,
                    "plate_after": plate_after_for_history,
                    "now": now,
                },
            )

            if _is_paid_complete(paid_date, amount):
                if plate_after_for_history:
                    before_snapshot = _vehicle_snapshot(v)
                    v.plate_old = plate_before_for_history or current_active_plate or v.plate_old
                    v.plate_new = plate_after_for_history
                    v.updated_at = now
                    db.session.flush()
                    after_snapshot = _vehicle_snapshot(v)
                    _log_vehicle_edit(v, before_snapshot, after_snapshot, changed_by=session.get("user_name") or "SYSTEM")

                next_due = _shift_years_safe(target_due_date, 5)
                next_row = db.session.execute(
                    text("""
                        SELECT id
                        FROM five_year_tax_payments
                        WHERE vehicle_id = :vehicle_id AND due_date = :due_date
                        ORDER BY id DESC
                        LIMIT 1
                    """),
                    {"vehicle_id": current_five["vehicle_id"], "due_date": next_due},
                ).mappings().first()
                if not next_row:
                    db.session.execute(
                        text("""
                            INSERT INTO five_year_tax_payments
                            (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                            VALUES
                            (:vehicle_id, :due_date, NULL, 0, :plate_before, NULL, NULL, :now, :now)
                        """),
                        {
                            "vehicle_id": current_five["vehicle_id"],
                            "due_date": next_due,
                            "plate_before": plate_after_for_history or plate_before_for_history,
                            "now": now,
                        },
                    )

                # Setelah 5 tahunan lunas, langsung siapkan tahunan pertama siklus berikutnya.
                _create_next_annual_stub(current_five["vehicle_id"], _shift_years_safe(target_due_date, 1), note=note, now=now)
            try:
                audit("PAY", "FiveYearTax", record_id, note=f"Tandai lunas pajak 5 tahunan {vehicle_plate(v)}")
            except Exception:
                pass
        else:
            if not due_date:
                flash("Tanggal exp pajak tahunan tidak ditemukan.", "danger")
                return redirect(url_for("main.go_data_pajak"))

            annual_due_date = due_date
            five_due_date = annual_due_to_five_year_due(annual_due_date)
            tax_year = annual_due_date.year

            existing_five = db.session.execute(
                text("""
                    SELECT id, vehicle_id, due_date, plate_before, plate_after
                    FROM five_year_tax_payments
                    WHERE vehicle_id = :vehicle_id
                    ORDER BY CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                             due_date DESC,
                             id DESC
                    LIMIT 1
                """),
                {"vehicle_id": vehicle_id},
            ).mappings().first()

            if existing_five:
                db.session.execute(
                    text("""
                        UPDATE five_year_tax_payments
                        SET due_date = :due_date,
                            updated_at = :now
                        WHERE id = :id
                    """),
                    {"id": existing_five["id"], "due_date": five_due_date, "now": now},
                )
            else:
                db.session.execute(
                    text("""
                        INSERT INTO five_year_tax_payments
                        (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                        VALUES
                        (:vehicle_id, :due_date, NULL, 0, :plate_before, :plate_after, NULL, :now, :now)
                    """),
                    {
                        "vehicle_id": vehicle_id,
                        "due_date": five_due_date,
                        "plate_before": getattr(v, "plate_old", None),
                        "plate_after": getattr(v, "plate_new", None),
                        "now": now,
                    },
                )

            annual_row = db.session.execute(
                text("""
                    SELECT id, paid_date, amount
                    FROM annual_tax_payments
                    WHERE vehicle_id = :vehicle_id AND tax_year = :tax_year
                    ORDER BY id DESC
                    LIMIT 1
                """),
                {"vehicle_id": vehicle_id, "tax_year": tax_year},
            ).mappings().first()
            if annual_row:
                _ensure_unpaid_tax_payment(annual_row, "Pajak tahunan")
                db.session.execute(
                    text("""
                        UPDATE annual_tax_payments
                        SET due_date = :due_date,
                            paid_date = :paid_date,
                            amount = :amount,
                            note = :note,
                            updated_at = :now
                        WHERE id = :id
                    """),
                    {
                        "id": annual_row["id"],
                        "due_date": annual_due_date,
                        "paid_date": paid_date,
                        "amount": amount,
                        "note": note,
                        "now": now,
                    },
                )
            else:
                db.session.execute(
                    text("""
                        INSERT INTO annual_tax_payments
                        (vehicle_id, tax_year, due_date, paid_date, amount, note, created_at, updated_at)
                        VALUES
                        (:vehicle_id, :tax_year, :due_date, :paid_date, :amount, :note, :now, :now)
                    """),
                    {
                        "vehicle_id": vehicle_id,
                        "tax_year": tax_year,
                        "due_date": annual_due_date,
                        "paid_date": paid_date,
                        "amount": amount,
                        "note": note,
                        "now": now,
                    },
                )


            # Rolling tahunan: setelah pajak tahun ini dibayar, siapkan tahun berikutnya
            # selama masih dalam siklus sebelum pajak 5 tahunan.
            active_five = db.session.execute(
                text("""
                    SELECT due_date
                    FROM five_year_tax_payments
                    WHERE vehicle_id = :vehicle_id
                    ORDER BY CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                             due_date DESC,
                             id DESC
                    LIMIT 1
                """),
                {"vehicle_id": vehicle_id},
            ).mappings().first()
            active_five_due = active_five.get("due_date") if active_five else None
            next_annual_due = _shift_years_safe(due_date, 1)
            if active_five_due and next_annual_due and next_annual_due.year < active_five_due.year:
                _create_next_annual_stub(vehicle_id, next_annual_due, note=note, now=now)
            try:
                audit("PAY", "AnnualTax", vehicle_id, note=f"Tandai lunas pajak tahunan {vehicle_plate(v)}")
            except Exception:
                pass

        db.session.commit()
        flash("Pembayaran selesai. Data otomatis masuk ke history pembayaran.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Gagal tandai lunas: {exc}", "danger")

    return redirect(url_for("main.history_pembayaran"))


@bp.post("/data/pajak/<int:record_id>/delete")
@master_required
def data_pajak_delete(record_id: int):
    db.session.execute(
        text("DELETE FROM five_year_tax_payments WHERE id = :id"),
        {"id": record_id},
    )

    try:
        audit("DELETE", "TaxPayment", record_id, note=f"Hapus pajak id={record_id}")
    except Exception:
        pass

    try:
        db.session.commit()
        flash("Data pajak berhasil dihapus.", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Gagal hapus pajak: {exc}", "danger")

    return redirect(url_for("main.go_data_pajak"))

@bp.post("/api/pajak/update-exp-note")
def api_pajak_update_exp_note():
    """
    Body JSON:
        { vehicle_id, record_id (opsional), annual_due, note }

    Logika:
    - annual_due diinput manual dari icon pensil.
    - five_due disimpan otomatis ke five_year_tax_payments.due_date
      sebagai annual_due + 5 tahun.
    - note disimpan ke five_year_tax_payments.note.
    - Jika annual_due dikosongkan → due_date ikut dikosongkan.
    - Jika belum ada record dan semua input kosong → tidak membuat record baru.
    """
    _ensure_annual_tax_table()
    data = request.get_json(force=True) or {}
    vehicle_id = _parse_int(data.get("vehicle_id"))
    record_id = _parse_int(data.get("record_id"))
    annual_due_str = (data.get("annual_due") or "").strip()
    note = _normalize_text(data.get("note")) or None

    if not vehicle_id:
        return jsonify({"success": False, "error": "vehicle_id wajib diisi"})

    v = Vehicle.query.get(vehicle_id)
    if not v:
        return jsonify({"success": False, "error": "Kendaraan tidak ditemukan"})

    annual_due = _parse_date(annual_due_str)
    if annual_due_str and not annual_due:
        return jsonify({"success": False, "error": "Format exp pajak tahunan tidak valid"}), 400

    five_due = annual_due_to_five_year_due(annual_due) if annual_due else None
    now = _now_naive()

    try:
        # Logika validasi ketat dihapus agar user bisa mengedit tanggal meskipun sudah ada pembayaran.
        # "Namanya manusia kurang teliti" - User request.
        pass

        target_id = record_id
        if target_id:
            existing = db.session.execute(
                text("SELECT id FROM five_year_tax_payments WHERE id = :id LIMIT 1"),
                {"id": target_id},
            ).mappings().first()
            if not existing:
                target_id = None

        if not target_id:
            existing = db.session.execute(
                text("""
                    SELECT id FROM five_year_tax_payments
                    WHERE vehicle_id = :vid
                    ORDER BY CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                             due_date DESC, id DESC
                    LIMIT 1
                """),
                {"vid": vehicle_id},
            ).mappings().first()
            target_id = existing["id"] if existing else None

        if target_id:
            db.session.execute(
                text("""
                    UPDATE five_year_tax_payments
                    SET due_date = :due_date,
                        note = :note,
                        updated_at = :now
                    WHERE id = :id
                """),
                {"id": target_id, "due_date": five_due, "note": note, "now": now},
            )
        elif annual_due or note:
            db.session.execute(
                text("""
                    INSERT INTO five_year_tax_payments
                    (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at)
                    VALUES
                    (:vid, :due_date, NULL, 0, :plate_before, NULL, :note, :now, :now)
                """),
                {
                    "vid": vehicle_id,
                    "due_date": five_due,
                    "plate_before": getattr(v, "plate_old", None) or getattr(v, "plate_new", None),
                    "note": note,
                    "now": now,
                },
            )
        else:
            return jsonify({"success": True, "annual_due": None, "five_due": None})

        # Reset semua annual_tax_payments saat exp diubah
        # karena tanggal/bulan/tahun bisa berubah total → siklus lama tidak valid lagi
        # History pembayaran tetap aman di tabel payment_history
        if annual_due:
            db.session.execute(
                text("DELETE FROM annual_tax_payments WHERE vehicle_id = :vid"),
                {"vid": vehicle_id},
            )
            db.session.execute(
                text("""
                    INSERT INTO annual_tax_payments
                    (vehicle_id, tax_year, due_date, paid_date, amount, note, created_at, updated_at)
                    VALUES (:vid, :tax_year, :due_date, NULL, 0, :note, :now, :now)
                """),
                {"vid": vehicle_id, "tax_year": annual_due.year, "due_date": annual_due, "note": note, "now": now},
            )
        elif annual_due_str == "":
            # Dikosongkan - hapus semua
            db.session.execute(
                text("DELETE FROM annual_tax_payments WHERE vehicle_id = :vid"),
                {"vid": vehicle_id},
            )

        db.session.commit()
        try:
            audit(
                "UPDATE",
                "TaxPayment",
                vehicle_id,
                note=f"Update exp pajak tahunan: {annual_due_str or 'kosong'} | exp 5 tahunan auto: {(five_due.isoformat() if five_due else 'kosong')} | note: {note or 'kosong'}",
            )
        except Exception:
            pass
        return jsonify({
            "success": True,
            "annual_due": annual_due.isoformat() if annual_due else None,
            "five_due": five_due.isoformat() if five_due else None,
        })
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success": False, "error": str(exc)})


def _safe_export_filename_part(value:str|None, fallback:str="-KENDARAAN-") -> str:
    raw = _normalize_text(value) or fallback
    cleaned = "".join(ch if ch.isalnum() or ch in " ()-_" else "_" for ch in raw)
    cleaned = " ".join(cleaned.split())
    return cleaned.strip(" ._") or fallback


# ── API: export Excel resume pajak per kendaraan ─────────────
@bp.get("/api/pajak/history/<int:vehicle_id>")
def api_pajak_history(vehicle_id: int):
    """
    API untuk mendapatkan history pembayaran pajak dari tabel payment_history.
    Menampilkan riwayat pembayaran lengkap.
    - Pajak Tahunan: tanpa informasi plat
    - Pajak 5 Tahunan: dengan informasi ganti plat
    """
    _ensure_annual_tax_table()
    _ensure_payment_history_table()
    
    v = Vehicle.query.get_or_404(vehicle_id)
    vehicle_name = vehicle_display_name(v)
    plate = vehicle_plate(v)
    
    # Ambil dari tabel payment_history (history pembayaran)
    history_rows = db.session.execute(
        text("""
            SELECT payment_type, due_date, paid_date, amount, plate_before, plate_after, note
            FROM payment_history
            WHERE vehicle_id = :vid
            ORDER BY COALESCE(paid_date, due_date) DESC, id DESC
        """),
        {"vid": vehicle_id},
    ).mappings().all()
    
    def fd(d):
        return format_tgl_id_full(d) if d else "-"
    
    history = []
    for r in history_rows:
        payment_type = r.get("payment_type") or "unknown"
        type_label = "Pajak Tahunan" if payment_type == "annual" else "Pajak 5 Tahunan"
        
        if payment_type == "five_year":
            # Pajak 5 Tahunan - tampilkan informasi ganti plat
            text_content = (
                f"{type_label} • "
                f"Jatuh tempo: {fd(r.get('due_date'))} • "
                f"Bayar: {fd(r.get('paid_date'))} • "
                f"Nominal: {rupiah(r.get('amount')) if r.get('amount') else '-'} • "
                f"Plat lama: {r.get('plate_before') or '-'} • "
                f"Plat baru: {r.get('plate_after') or '-'} • "
                f"Catatan: {r.get('note') or '-'}"
            )
            type_order = 1
        else:
            # Pajak Tahunan - tanpa informasi plat
            text_content = (
                f"{type_label} • "
                f"Jatuh tempo: {fd(r.get('due_date'))} • "
                f"Bayar: {fd(r.get('paid_date'))} • "
                f"Nominal: {rupiah(r.get('amount')) if r.get('amount') else '-'} • "
                f"Catatan: {r.get('note') or '-'}"
            )
            type_order = 0
        
        history.append({
            "sort_date": r.get("paid_date") or r.get("due_date") or date.min,
            "type_order": type_order,
            "date": fd(r.get("paid_date") or r.get("due_date")),
            "text": text_content
        })
    
    history.sort(key=lambda x: (x["sort_date"], x["type_order"]), reverse=True)
    
    return jsonify({
        "success": True,
        "vehicle": vehicle_name,
        "plate": plate,
        "history": [{"date": x["date"], "text": x["text"]} for x in history],
    })

# ── API: export Excel resume pajak per kendaraan ─────────────
@bp.get("/api/pajak/export/<int:vehicle_id>")
def api_pajak_export(vehicle_id):
    """
    Mirip export servis:
    - support filter date_from/date_to
    - support export_type = all | summary | history_annual | history_five | history
    - hasil resume ikut menampilkan periode filter
    """
    _ensure_annual_tax_table()
    v = Vehicle.query.get_or_404(vehicle_id)

    export_type = (_normalize_text(request.args.get("export_type")) or "all").lower()
    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))

    if date_from and date_to and date_from > date_to:
        return jsonify({"success": False, "message": "Tanggal awal tidak boleh lebih besar dari tanggal akhir."}), 400

    nama = vehicle_display_name(v)
    plat = vehicle_plate(v)
    pt = getattr(v, "pt", None) or "-"

    five_row = db.session.execute(
        text("""
            SELECT id, due_date, paid_date, amount, note, plate_before, plate_after
            FROM five_year_tax_payments
            WHERE vehicle_id = :vid
            ORDER BY CASE WHEN paid_date IS NULL OR amount IS NULL OR amount <= 0 THEN 0 ELSE 1 END ASC,
                     due_date DESC, id DESC
            LIMIT 1
        """),
        {"vid": vehicle_id}
    ).mappings().first()

    main_due = five_row["due_date"] if five_row else None
    annual_map = _annual_payment_map([vehicle_id])
    annual_info = _annual_tax_cycle(vehicle_id, main_due, annual_map)

    five_is_paid = _is_paid_complete(
        five_row["paid_date"] if five_row else None,
        five_row["amount"] if five_row else None
    ) if five_row else False

    annual_status_label = {
        "paid": "Sudah dibayar", "overdue": "Lewat jatuh tempo",
        "soon": "Jatuh tempo dekat", "empty": "Belum diisi", "safe": "Aman"
    }.get(annual_info["display_status"], "Aman")
    five_status = "empty" if not main_due else ("paid" if five_is_paid else _tax_status(main_due))
    five_status_label = {
        "paid": "Sudah dibayar", "overdue": "Lewat jatuh tempo",
        "soon": "Jatuh tempo dekat", "empty": "Belum diisi", "safe": "Aman"
    }.get(five_status, "Aman")

    annual_query = db.session.execute(
        text("""
            SELECT tax_year, due_date, paid_date, amount, note
            FROM annual_tax_payments
            WHERE vehicle_id = :vid
            ORDER BY COALESCE(paid_date, due_date) DESC, id DESC
        """),
        {"vid": vehicle_id}
    ).mappings().all()

    five_query = db.session.execute(
        text("""
            SELECT due_date, paid_date, amount, plate_before, plate_after, note
            FROM five_year_tax_payments
            WHERE vehicle_id = :vid
            ORDER BY COALESCE(paid_date, due_date) DESC, id DESC
        """),
        {"vid": vehicle_id}
    ).mappings().all()

    def _pajak_export_as_date(value):
        """Samakan nilai tanggal dari DB/form menjadi datetime.date.
        Ini mencegah TypeError: can't compare datetime.datetime to datetime.date
        saat filter export membandingkan paid_date/due_date dengan date_from/date_to.
        """
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        dt_value = _coerce_datetime_value(value)
        return dt_value.date() if dt_value else None

    def in_range(row_date):
        row_date = _pajak_export_as_date(row_date)
        if not row_date:
            return False if (date_from or date_to) else True
        if date_from and row_date < date_from:
            return False
        if date_to and row_date > date_to:
            return False
        return True

    annual_history = [
        r for r in annual_query
        if in_range(r.get("paid_date") or r.get("due_date"))
    ]
    five_history = [
        r for r in five_query
        if in_range(r.get("paid_date") or r.get("due_date"))
    ]

    def fd(d):
        return format_tgl_id_full(d) if d else "-"

    def rp(val):
        return rupiah(val) if val not in (None, "") else "-"

    total_annual = sum(float(r.get("amount") or 0) for r in annual_history)
    total_five = sum(float(r.get("amount") or 0) for r in five_history)
    total_rows = len(annual_history) + len(five_history)
    latest_history_date = None
    if annual_history or five_history:
        latest_history_date = max([
            _pajak_export_as_date(r.get("paid_date") or r.get("due_date")) or date.min
            for r in annual_history + five_history
        ])

    _nama_utama = (
        getattr(v, "active_name", None)
        or getattr(v, "new_asset_name", None)
        or getattr(v, "name_as_asset_pt", None)
        or nama
    )
    summary_data = [
        ("Nama Aset", _nama_utama or "-"),
        ("No Polisi", plat or "-"),
        ("PT", pt),
        ("Merk", _export_scalar(getattr(v, "merk", None))),
        ("Tipe", _export_scalar(getattr(v, "type", None))),
        ("Jenis", _export_scalar(getattr(v, "jenis", None))),
        ("Tahun Pemakaian", _export_scalar(getattr(v, "year_of_use", None))),
        ("Tanggal Awal Filter", fd(date_from) if date_from else "Semua"),
        ("Tanggal Akhir Filter", fd(date_to) if date_to else "Semua"),
        ("Total Riwayat Pajak", total_rows),
        ("Total Riwayat Tahunan", len(annual_history)),
        ("Total Riwayat 5 Tahunan", len(five_history)),
        ("Total Nominal Tahunan", rupiah(total_annual)),
        ("Total Nominal 5 Tahunan", rupiah(total_five)),
        ("Total Nominal Pajak", rupiah(total_annual + total_five)),
        ("Exp Pajak Tahunan (Aktif)", fd(annual_info.get("display_due_date"))),
        ("Tahun Pajak Aktif", str(annual_info.get("tax_year") or "-")),
        ("Status Tahunan", annual_status_label),
        ("Tanggal Bayar Tahunan Aktif", fd(annual_info.get("display_paid_date"))),
        ("Nominal Tahunan Aktif", rp(annual_info.get("display_amount"))),
        ("Exp Pajak 5 Tahunan", fd(main_due)),
        ("Status 5 Tahunan", five_status_label),
        ("Tanggal Bayar 5 Tahunan Terakhir", fd(five_row["paid_date"] if five_row else None)),
        ("Nominal 5 Tahunan Terakhir", rp(five_row["amount"] if five_row else None)),
        ("Plat Lama (5 Tahunan)", five_row.get("plate_before") if five_row and five_row.get("plate_before") else "-"),
        ("Plat Baru (5 Tahunan)", five_row.get("plate_after") if five_row and five_row.get("plate_after") else "-"),
        ("Catatan Aktif", (annual_info.get("display_note") or (five_row.get("note") if five_row else None) or "-")),
        ("Tanggal Riwayat Terakhir Dalam Filter", fd(latest_history_date)),
    ]
    df_summary = pd.DataFrame(summary_data, columns=["Informasi", "Nilai"])

    if annual_history:
        df_annual = pd.DataFrame([{
            "Tanggal Referensi": fd(r.get("paid_date") or r.get("due_date")),
            "Tahun Pajak": r.get("tax_year") or "-",
            "Exp / Jatuh Tempo": fd(r.get("due_date")),
            "Tanggal Bayar": fd(r.get("paid_date")),
            "Nominal": rp(r.get("amount")),
            "Catatan": r.get("note") or "-",
        } for r in annual_history])
    else:
        df_annual = pd.DataFrame(columns=["Tanggal Referensi","Tahun Pajak","Exp / Jatuh Tempo","Tanggal Bayar","Nominal","Catatan"])

    if five_history:
        df_five = pd.DataFrame([{
            "Tanggal Referensi": fd(r.get("paid_date") or r.get("due_date")),
            "Exp / Jatuh Tempo": fd(r.get("due_date")),
            "Tanggal Bayar": fd(r.get("paid_date")),
            "Nominal": rp(r.get("amount")),
            "Plat Lama": r.get("plate_before") or "-",
            "Plat Baru": r.get("plate_after") or "-",
            "Catatan": r.get("note") or "-",
        } for r in five_history])
    else:
        df_five = pd.DataFrame(columns=["Tanggal Referensi","Exp / Jatuh Tempo","Tanggal Bayar","Nominal","Plat Lama","Plat Baru","Catatan"])

    history_combined = []
    for r in annual_history:
        history_combined.append({
            "Tanggal Referensi": fd(r.get("paid_date") or r.get("due_date")),
            "Jenis": "Pajak Tahunan",
            "Tahun Pajak": r.get("tax_year") or "-",
            "Exp / Jatuh Tempo": fd(r.get("due_date")),
            "Tanggal Bayar": fd(r.get("paid_date")),
            "Nominal": rp(r.get("amount")),
            "Plat Lama": "-",
            "Plat Baru": "-",
            "Catatan": r.get("note") or "-",
        })
    for r in five_history:
        history_combined.append({
            "Tanggal Referensi": fd(r.get("paid_date") or r.get("due_date")),
            "Jenis": "Pajak 5 Tahunan",
            "Tahun Pajak": "-",
            "Exp / Jatuh Tempo": fd(r.get("due_date")),
            "Tanggal Bayar": fd(r.get("paid_date")),
            "Nominal": rp(r.get("amount")),
            "Plat Lama": r.get("plate_before") or "-",
            "Plat Baru": r.get("plate_after") or "-",
            "Catatan": r.get("note") or "-",
        })
    if history_combined:
        def _sort_key(item):
            return item.get("Tanggal Referensi") or ""
        df_history = pd.DataFrame(history_combined)
    else:
        df_history = pd.DataFrame(columns=["Tanggal Referensi","Jenis","Tahun Pajak","Exp / Jatuh Tempo","Tanggal Bayar","Nominal","Plat Lama","Plat Baru","Catatan"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if export_type in {"all", "summary"}:
            df_summary.to_excel(writer, sheet_name="Ringkasan Pajak", index=False)
        if export_type in {"all", "history_annual"}:
            df_annual.to_excel(writer, sheet_name="Riwayat Tahunan", index=False)
        if export_type in {"all", "history_five"}:
            df_five.to_excel(writer, sheet_name="Riwayat 5 Tahunan", index=False)
        if export_type in {"all", "history"}:
            df_history.to_excel(writer, sheet_name="Riwayat Gabungan", index=False)

        for ws in writer.book.worksheets:
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True, color="FFD15F")
                    cell.fill = cell.fill.copy(fill_type="solid", fgColor="1E4080")
            _autosize_worksheet_columns(ws, min_width=14, max_width=38)

    output.seek(0)
    safe_name = _safe_export_filename_part(nama, "KENDARAAN")
    safe_plate = _safe_export_filename_part(plat, "-")
    period_suffix = []
    if date_from:
        period_suffix.append(date_from.strftime("%Y-%m-%d"))
    if date_to:
        period_suffix.append(date_to.strftime("%Y-%m-%d"))
    period_text = f"_{'_to_'.join(period_suffix)}" if period_suffix else "_semua-periode"
    filename = f"PAJAK_{safe_name.replace(' ', '_')}_{safe_plate.replace(' ', '_')}{period_text}_{date.today().strftime('%d-%m-%Y')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.post("/alert/pajak/claim")
@master_required
def alert_pajak_claim():
    """Endpoint AJAX bayar pajak — return JSON, tidak redirect."""
    _ensure_annual_tax_table()
    vehicle_id=_parse_int(request.form.get("vehicle_id"))
    kind=_normalize_text(request.form.get("kind")).lower()
    due_date=_parse_date(request.form.get("due_date"))
    paid_date=_parse_date_as_datetime(request.form.get("paid_date")) or _today_as_datetime()
    amount=_parse_float(request.form.get("amount"))
    note=_normalize_text(request.form.get("note")) or None

    if not vehicle_id:
        return jsonify({"success":False,"message":"Kendaraan tidak ditemukan."}),400
    if amount is None or amount<=0:
        return jsonify({"success":False,"message":"Nominal bayar wajib diisi dan harus lebih dari 0."}),400

    v=Vehicle.query.get_or_404(vehicle_id)

    try:
        if kind=="annual":
            if not due_date:
                return jsonify({"success":False,"message":"Due date pajak tahunan tidak ditemukan."}),400
            tax_year=due_date.year
            row=db.session.execute(text("SELECT id FROM annual_tax_payments WHERE vehicle_id=:vehicle_id AND tax_year=:tax_year ORDER BY id DESC LIMIT 1"),{"vehicle_id":vehicle_id,"tax_year":tax_year}).mappings().first()
            if row:
                db.session.execute(text("UPDATE annual_tax_payments SET due_date=:due_date, paid_date=:paid_date, amount=:amount, note=:note, updated_at=:now WHERE id=:id"),{"id":row["id"],"due_date":due_date,"paid_date":paid_date,"amount":amount,"note":note,"now":_now_naive()})
            else:
                db.session.execute(text("INSERT INTO annual_tax_payments (vehicle_id,tax_year,due_date,paid_date,amount,note,created_at,updated_at) VALUES (:vehicle_id,:tax_year,:due_date,:paid_date,:amount,:note,:now,:now)"),{"vehicle_id":vehicle_id,"tax_year":tax_year,"due_date":due_date,"paid_date":paid_date,"amount":amount,"note":note,"now":_now_naive()})
            try:audit("PAY","AnnualTax",vehicle_id,note=f"Tandai bayar pajak tahunan {vehicle_plate(v)}")
            except Exception:pass
        elif kind=="five":
            record_id=_parse_int(request.form.get("record_id"))
            if not record_id:
                return jsonify({"success":False,"message":"Record pajak 5 tahunan tidak ditemukan."}),400
            current_five=db.session.execute(text("SELECT vehicle_id, due_date, plate_before, plate_after FROM five_year_tax_payments WHERE id=:id LIMIT 1"),{"id":record_id}).mappings().first()
            if not current_five:
                return jsonify({"success":False,"message":"Data pajak 5 tahunan tidak ditemukan."}),400
            current_plate=_normalize_text(getattr(v,"plate_new",None) or getattr(v,"plate_old",None))
            db.session.execute(text("UPDATE five_year_tax_payments SET paid_date=:paid_date, amount=:amount, note=:note, plate_before=:plate_before, plate_after=NULL, updated_at=:now WHERE id=:id"),{"id":record_id,"paid_date":paid_date,"amount":amount,"note":note,"plate_before":current_plate or (current_five.get("plate_before") if current_five else None),"now":_now_naive()})
            if _is_paid_complete(paid_date, amount):
                next_due=_shift_years_safe(current_five["due_date"],5)
                next_row=db.session.execute(text("SELECT id FROM five_year_tax_payments WHERE vehicle_id=:vehicle_id AND due_date=:due_date ORDER BY id DESC LIMIT 1"),{"vehicle_id":current_five["vehicle_id"],"due_date":next_due}).mappings().first()
                if not next_row:
                    db.session.execute(text("INSERT INTO five_year_tax_payments (vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note, created_at, updated_at) VALUES (:vehicle_id,:due_date,NULL,0,:plate_before,NULL,NULL,:now,:now)"),{"vehicle_id":current_five["vehicle_id"],"due_date":next_due,"plate_before":current_plate or current_five.get("plate_before"),"now":_now_naive()})
            try:audit("PAY","FiveYearTax",record_id,note=f"Tandai bayar pajak 5 tahunan {vehicle_plate(v)}")
            except Exception:pass
        else:
            return jsonify({"success":False,"message":"Jenis pajak tidak valid."}),400
        db.session.commit()
        return jsonify({"success":True,"message":"Pembayaran berhasil dicatat!"})
    except Exception as exc:
        db.session.rollback()
        return jsonify({"success":False,"message":f"Gagal simpan pembayaran: {exc}"}),500



def _service_anchor_note(clean_note: str | None = None) -> str:
    clean_note = _normalize_text(clean_note)
    parts = ["[JADWAL_SERVIS]"]
    if clean_note:
        parts.append("Catatan: " + clean_note)
    return " || ".join(parts)


def _latest_service_anchor(vehicle_id: int):
    """Ambil jatuh tempo servis terakhir (anchor jadwal)."""
    return (
        ServiceRecord.query
        .filter(ServiceRecord.vehicle_id == vehicle_id)
        .filter(func.lower(ServiceRecord.note).like("%jadwal_servis%"))
        .order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc())
        .first()
    )


def _latest_processed_service(vehicle_id: int):
    """Ambil transaksi servis terakhir, exclude record anchor [JADWAL_SERVIS]."""
    return (
        ServiceRecord.query
        .filter(ServiceRecord.vehicle_id == vehicle_id)
        .filter(or_(ServiceRecord.note.is_(None), ~func.lower(ServiceRecord.note).like("%jadwal_servis%")))
        .order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc())
        .first()
    )


def _create_service_anchor(vehicle_id: int, due_date, note: str | None = None):
    """Simpan anchor jatuh tempo servis baru sebagai record khusus jadwal."""
    if not due_date:
        return None
    target = ServiceRecord(
        vehicle_id=vehicle_id,
        service_date=due_date,
        note=_service_anchor_note(note),
        vendor=None,
        cost=0,
        odometer_km=None,
    )
    for attr in ("service_type", "jenis_servis", "jenis_service"):
        if hasattr(target, attr):
            setattr(target, attr, "Jadwal Servis")
            break
    db.session.add(target)
    db.session.flush()
    _service_upsert_meta(target.id, _now_naive())
    return target

@bp.get("/data/servis", endpoint="go_data_servis")
def go_data_servis():
    _ensure_service_support_tables()
    q = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status") or "").strip().lower()
    company = _sanitize_company_filter((request.args.get("company") or "").strip())
    companies = _scoped_company_choices()

    # PENTING: Data Servis harus ikut master Data Kendaraan.
    # Sumber utama halaman ini adalah Vehicle aktif, bukan ServiceRecord.
    # Jadi kendaraan baru tetap muncul walaupun belum punya riwayat/jadwal servis.
    vehicles = (
        _vehicle_query_filtered(q, company)
        .order_by(Vehicle.updated_at.desc(), Vehicle.id.desc())
        .limit(2000)
        .all()
    )

    rows = []
    prepared = []
    latest_service_ids = []
    for i, v in enumerate(vehicles, 1):
        # Transaksi servis terakhir (history) dipisah dari anchor jadwal.
        s_any = _latest_processed_service(v.id)

        # Anchor jadwal servis: tanggal jatuh tempo yang disimpan lewat tombol Edit.
        s_anchor = _latest_service_anchor(v.id)

        if s_any and getattr(s_any, "id", None):
            latest_service_ids.append(s_any.id)
        prepared.append((i, v, s_any, s_anchor))

    items_map = _service_items_map(latest_service_ids)
    processed_map = _service_processed_map(latest_service_ids)

    for i, v, s_any, s_anchor in prepared:
        # Anchor adalah JATUH TEMPO servis saat ini.
        # Tanggal proses servis tidak menjadi patokan jadwal, sama seperti tanggal bayar pajak.
        current_due = getattr(s_anchor, "service_date", None) if s_anchor else None
        next_due_preview = _add_6_months(current_due) if current_due else None
        row_status = _service_status(current_due) if current_due else "empty"
        nama_utama = (
            getattr(v, "active_name", None)
            or getattr(v, "new_asset_name", None)
            or getattr(v, "name_as_asset_pt", None)
            or vehicle_display_name(v)
        )
        no_polisi = vehicle_plate(v)
        rows.append({
            "no": i,
            "vehicle_id": v.id,
            "pt": v.pt or "-",
            "nama_utama": nama_utama,
            "kendaraan": nama_utama,
            "no_polisi": no_polisi,
            "tgl_servis": current_due,
            "next_servis": next_due_preview,
            "status_servis": row_status,
            "vendor": getattr(s_any, "vendor", None) if s_any else None,
            "jenis_servis": (
                getattr(s_any, "service_type", None)
                or getattr(s_any, "jenis_servis", None)
                or getattr(s_any, "jenis_service", None)
            ) if s_any else None,
            "odo_km": (
                getattr(s_any, "odometer_km", None)
                or getattr(s_any, "odometer", None)
                or getattr(s_any, "km", None)
            ) if s_any else None,
            "biaya": (
                getattr(s_any, "cost", None)
                or getattr(s_any, "amount", None)
                or getattr(s_any, "biaya", None)
                or 0
            ) if s_any else 0,
            "note": getattr(s_any, "note", None) if s_any else None,
            "note_display": _clean_service_note(getattr(s_any, "note", None)) if s_any else None,
            "items_list": items_map.get(getattr(s_any, "id", None), []) if s_any else [],
            "items_summary": _service_items_summary(s_any) if s_any else "",
            "processed_at": processed_map.get(getattr(s_any, "id", None)) if s_any else None,
            "service_id": getattr(s_anchor, "id", None) if s_anchor else None,
            "process_service_id": getattr(s_any, "id", None) if s_any else None,
            "can_process": bool(s_anchor and getattr(s_anchor, "service_date", None)),
            "vehicle": v,
        })

    # Filter status dari URL harus memakai variabel khusus, jangan ditimpa oleh status per baris.
    if status_filter:
        rows = [r for r in rows if str(r.get("status_servis") or "").lower() == status_filter]
        for idx, r in enumerate(rows, 1):
            r["no"] = idx

    # Filter tambahan di sisi row dibuat seluas Data Kendaraan, supaya pencarian PT/merk/type/user/lokasi
    # tidak hilang setelah kendaraan berhasil ditemukan dari _vehicle_query_filtered().
    if q:
        ql = q.lower()
        rows = [
            r for r in rows
            if ql in " ".join([
                _normalize_text(r.get("pt")),
                _normalize_text(r.get("nama_utama")),
                _normalize_text(r.get("no_polisi")),
                _normalize_text(r.get("vendor")),
                _normalize_text(r.get("jenis_servis")),
                _normalize_text(r.get("note")),
                _normalize_text(getattr(r.get("vehicle"), "active_name", None)),
                _normalize_text(getattr(r.get("vehicle"), "new_asset_name", None)),
                _normalize_text(getattr(r.get("vehicle"), "name_as_asset_pt", None)),
                _normalize_text(getattr(r.get("vehicle"), "merk", None)),
                _normalize_text(getattr(r.get("vehicle"), "type", None)),
                _normalize_text(getattr(r.get("vehicle"), "jenis", None)),
                _normalize_text(getattr(r.get("vehicle"), "plate_old", None)),
                _normalize_text(getattr(r.get("vehicle"), "plate_new", None)),
                _normalize_text(getattr(r.get("vehicle"), "user_old", None)),
                _normalize_text(getattr(r.get("vehicle"), "user_new", None)),
                _normalize_text(getattr(r.get("vehicle"), "status", None)),
                _normalize_text(getattr(r.get("vehicle"), "kondisi_terkini", None)),
                _normalize_text(getattr(r.get("vehicle"), "lokasi", None)),
                _normalize_text(getattr(r.get("vehicle"), "tambahan_keterangan", None)),
            ]).lower()
        ]
        for idx, r in enumerate(rows, 1):
            r["no"] = idx

    highlight_id = _parse_int(request.args.get("highlight"))
    return render_template(
        "data_servis.html",
        service_rows=rows,
        q=q,
        status=status_filter,
        company=company,
        companies=companies,
        highlight_id=highlight_id,
        title="Data Servis - Sinar Group",
    )

@bp.post("/data/servis/tambah")
@master_required
def data_servis_tambah():
    _ensure_service_support_tables()
    vehicle_id = _parse_int(request.form.get("vehicle_id"))
    v = Vehicle.query.get_or_404(vehicle_id)

    anchor = _latest_service_anchor(v.id)
    if not anchor or not getattr(anchor, "service_date", None):
        flash("Tidak bisa proses servis sebelum isi tanggal servis terakhir lewat tombol Edit.", "danger")
        return redirect(url_for("main.go_data_servis"))

    service_date_raw = request.form.get("service_date")
    service_date = _parse_date_as_datetime(service_date_raw)
    if not service_date:
        flash("Tanggal servis wajib diisi.", "danger")
        return redirect(url_for("main.go_data_servis"))

    items = []
    try:
        items = json.loads(request.form.get("items_json") or "[]")
    except Exception:
        pass

    total_cost = sum(float(i.get("biaya") or 0) for i in items)
    if total_cost <= 0:
        total_cost = _parse_float(request.form.get("cost")) or 0

    kategori = _normalize_text(request.form.get("kategori_servis")) or "rutin"
    jenis_pembayaran = _normalize_text(request.form.get("jenis_pembayaran")) or "reguler"
    service_type = f"Servis {kategori.capitalize()}"
    first_vendor = next((i.get("vendor","") for i in items if i.get("vendor","")), None)
    vendor = _normalize_text(request.form.get("vendor") or first_vendor or "") or None

    note_parts = []
    if jenis_pembayaran == "asuransi": note_parts.append("[Asuransi]")
    user_note = _normalize_text(request.form.get("note") or "")
    if user_note: note_parts.append("Catatan: " + user_note)
    final_note = " || ".join(note_parts) if note_parts else None

    current_due = getattr(anchor, "service_date", None)

    row = ServiceRecord(
        vehicle_id=v.id,
        service_date=service_date,
        odometer_km=_parse_int(request.form.get("odometer_km")),
        vendor=vendor,
        cost=total_cost,
        note=final_note,
    )
    for attr in ("service_type", "jenis_servis", "jenis_service"):
        if hasattr(row, attr):
            setattr(row, attr, service_type)
            break
    Vehicle.query.session.add(row)
    Vehicle.query.session.flush()
    _service_replace_items(row.id, items)
    _service_upsert_meta(row.id, _now_naive())

    # Jika servis rutin, jatuh tempo berikutnya maju 6 bulan dari JATUH TEMPO saat ini,
    # bukan dari tanggal proses. Servis berat tidak mengubah jadwal rutin.
    if "berat" not in kategori.lower():
        _create_service_anchor(v.id, _add_6_months(current_due))

    try:
        audit("CREATE", "ServiceRecord", 0, note=f"Tambah servis kendaraan {vehicle_plate(v)}")
    except Exception:
        pass

    return _safe_commit(
        "Servis selesai dicatat dan otomatis masuk ke history servis.",
        "Gagal tambah servis",
        "main.go_data_servis"
    )


@bp.post("/data/servis/<int:service_id>/edit")
@master_required
def data_servis_edit(service_id: int):
    _ensure_service_support_tables()
    row = ServiceRecord.query.get_or_404(service_id)

    anchor = _latest_service_anchor(row.vehicle_id)
    if not anchor or not getattr(anchor, "service_date", None):
        flash("Tidak bisa proses servis sebelum isi tanggal servis terakhir lewat tombol Edit.", "danger")
        return redirect(url_for("main.go_data_servis"))

    service_date_raw = request.form.get("service_date")
    service_date = _parse_date_as_datetime(service_date_raw)
    if not service_date:
        flash("Tanggal servis wajib diisi.", "danger")
        return redirect(url_for("main.go_data_servis"))

    items = []
    try:
        items = json.loads(request.form.get("items_json") or "[]")
    except Exception:
        pass

    total_cost = sum(float(i.get("biaya") or 0) for i in items)
    if total_cost <= 0:
        total_cost = _parse_float(request.form.get("cost")) or 0
    if total_cost <= 0:
        flash("Biaya servis wajib diisi dan harus lebih dari 0.", "danger")
        return redirect(url_for("main.go_data_servis"))

    kategori = _normalize_text(request.form.get("kategori_servis")) or "rutin"
    jenis_pembayaran = _normalize_text(request.form.get("jenis_pembayaran")) or "reguler"
    service_type = f"Servis {kategori.capitalize()}"
    first_vendor = next((i.get("vendor","") for i in items if i.get("vendor","")), None)
    vendor = _normalize_text(request.form.get("vendor") or first_vendor or "") or None

    note_parts = []
    if jenis_pembayaran == "asuransi":
        note_parts.append("[Asuransi]")
    user_note = _normalize_text(request.form.get("note") or "")
    if user_note:
        note_parts.append("Catatan: " + user_note)
    final_note = " || ".join(note_parts) if note_parts else None

    current_due = getattr(anchor, "service_date", None)

    new_row = _service_clone_record(
        row,
        service_date=service_date,
        odometer_km=_parse_int(request.form.get("odometer_km")),
        vendor=vendor,
        cost=total_cost,
        note=final_note,
        service_type=service_type,
        items=items,
        processed_at=_now_naive(),
    )

    if "berat" not in kategori.lower():
        _create_service_anchor(row.vehicle_id, _add_6_months(current_due))

    try:
        audit("UPDATE", "ServiceRecord", new_row.id, note=f"Edit servis buat history baru dari id={row.id} ke id={new_row.id}")
    except Exception:
        pass

    return _safe_commit(
        "Data servis berhasil diupdate dan tercetak sebagai history baru.",
        "Gagal update servis",
        "main.go_data_servis"
    )


@bp.post("/data/servis/<int:service_id>/delete")
@master_required
def data_servis_delete(service_id: int):
    row = ServiceRecord.query.get_or_404(service_id)
    Vehicle.query.session.delete(row)

    try:
        audit("DELETE", "ServiceRecord", service_id, note=f"Hapus servis id={service_id}")
    except Exception:
        pass

    return _safe_commit(
        "Data servis berhasil dihapus.",
        "Gagal hapus servis",
        "main.go_data_servis"
    )

@bp.post("/api/servis/update-tgl-note")
@master_required
def api_servis_update_tgl_note():
    _ensure_service_support_tables()
    data = request.get_json() or {}
    vehicle_id = _parse_int(data.get("vehicle_id"))
    tgl_servis_str = (data.get("tgl_servis") or "").strip()
    note = data.get("note")

    if not vehicle_id:
        return jsonify({"success": False, "error": "Vehicle ID is required"}), 400

    latest_anchor = _latest_service_anchor(vehicle_id)

    clean_note = _normalize_text(note) or None
    parsed_service_date = _parse_date_as_datetime(tgl_servis_str) if tgl_servis_str else None
    if tgl_servis_str and not parsed_service_date:
        return jsonify({"success": False, "error": "Format tanggal servis tidak valid"}), 400

    if not parsed_service_date and not clean_note:
        return jsonify({"success": True})
    if not parsed_service_date and not latest_anchor:
        return jsonify({"success": False, "error": "Tanggal servis wajib diisi untuk membuat jadwal servis awal"}), 400

    v = Vehicle.query.get_or_404(vehicle_id)
    target = _create_service_anchor(
        v.id,
        parsed_service_date if parsed_service_date else getattr(latest_anchor, 'service_date', None),
        clean_note,
    )

    try:
        audit(
            "UPDATE",
            "ServiceRecord",
            getattr(target, "id", 0),
            note=f"Update tanggal/catatan servis kendaraan {vehicle_id} menjadi history baru"
        )
    except Exception:
        pass

    try:
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


# =========================
# EXPORT SERVIS EXCEL
# =========================

def _service_record_type_label(record):
    raw = (
        getattr(record, "service_type", None)
        or getattr(record, "jenis_servis", None)
        or getattr(record, "jenis_service", None)
        or ""
    ).strip().lower()

    if "berat" in raw:
        return "Berat / Lainnya"
    if "rutin" in raw:
        return "Rutin"
    return "-"


def _service_payment_label(record):
    note = (getattr(record, "note", None) or "").lower()
    if "asuransi" in note:
        return "Asuransi"
    return "Reguler"


def _service_total_cost(record):
    raw = (
        getattr(record, "cost", None)
        or getattr(record, "amount", None)
        or getattr(record, "biaya", None)
        or 0
    )
    try:
        return float(raw or 0)
    except Exception:
        return 0.0


def _service_odometer_value(record):
    return (
        getattr(record, "odometer_km", None)
        or getattr(record, "odometer", None)
        or getattr(record, "km", None)
        or None
    )

_service_tables_ensured = False

def _ensure_service_support_tables():
    global _service_tables_ensured
    if _service_tables_ensured:
        return
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS service_record_items (
            id INT NOT NULL AUTO_INCREMENT,
            service_record_id INT NOT NULL,
            item_name VARCHAR(255) NULL,
            vendor VARCHAR(255) NULL,
            cost DECIMAL(18,2) NOT NULL DEFAULT 0,
            position_index INT NOT NULL DEFAULT 0,
            created_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            KEY idx_service_record_items_record (service_record_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """))
    db.session.execute(text("""
        CREATE TABLE IF NOT EXISTS service_record_meta (
            service_record_id INT NOT NULL,
            processed_at DATETIME NULL,
            updated_at TIMESTAMP NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (service_record_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
    """))
    db.session.commit()
    _service_tables_ensured = True


def _clean_service_items(items):
    cleaned = []
    for idx, it in enumerate(items or []):
        if not isinstance(it, dict):
            continue
        nama = _normalize_text(it.get("nama") or it.get("item_name"))
        vendor = _normalize_text(it.get("vendor"))
        try:
            biaya = float(it.get("biaya") or it.get("cost") or 0)
        except Exception:
            biaya = 0.0
        if not nama and not vendor and biaya <= 0:
            continue
        cleaned.append({"nama": nama, "vendor": vendor, "biaya": biaya, "position_index": idx})
    return cleaned


def _service_replace_items(service_record_id:int, items):
    _ensure_service_support_tables()
    db.session.execute(text("DELETE FROM service_record_items WHERE service_record_id=:service_record_id"), {"service_record_id": service_record_id})
    for idx, it in enumerate(_clean_service_items(items)):
        db.session.execute(
            text("""
                INSERT INTO service_record_items (service_record_id, item_name, vendor, cost, position_index)
                VALUES (:service_record_id, :item_name, :vendor, :cost, :position_index)
            """),
            {
                "service_record_id": service_record_id,
                "item_name": it.get("nama") or None,
                "vendor": it.get("vendor") or None,
                "cost": float(it.get("biaya") or 0),
                "position_index": idx,
            }
        )


def _service_items_map(service_record_ids:list[int]):
    result = {}
    ids = [int(x) for x in (service_record_ids or []) if x]
    if not ids:
        return result
    try:
        _ensure_service_support_tables()
        rows = db.session.execute(
            text("""
                SELECT service_record_id, item_name, vendor, cost, position_index
                FROM service_record_items
                WHERE service_record_id IN :ids
                ORDER BY service_record_id ASC, position_index ASC, id ASC
            """).bindparams(bindparam('ids', expanding=True)),
            {'ids': ids}
        ).mappings().all()
        for r in rows:
            result.setdefault(r['service_record_id'], []).append({
                'nama': (r['item_name'] or '').strip(),
                'vendor': (r['vendor'] or '').strip(),
                'biaya': float(r['cost'] or 0),
            })
    except Exception:
        return {}
    return result


def _service_upsert_meta(service_record_id:int, processed_at=None):
    _ensure_service_support_tables()
    processed_at = processed_at or _now_naive()
    exists = db.session.execute(text("SELECT service_record_id FROM service_record_meta WHERE service_record_id=:service_record_id"), {"service_record_id": service_record_id}).first()
    if exists:
        db.session.execute(text("UPDATE service_record_meta SET processed_at=:processed_at WHERE service_record_id=:service_record_id"), {"service_record_id": service_record_id, "processed_at": processed_at})
    else:
        db.session.execute(text("INSERT INTO service_record_meta (service_record_id, processed_at) VALUES (:service_record_id, :processed_at)"), {"service_record_id": service_record_id, "processed_at": processed_at})


def _service_clone_record(source_record, service_date=None, note=None, odometer_km=None, vendor=None, cost=None, service_type=None, items=None, processed_at=None):
    """Buat record servis baru dari record lama supaya history tidak ketimpa."""
    if not source_record:
        raise ValueError("Source service record tidak ditemukan")
    new_row = ServiceRecord(
        vehicle_id=source_record.vehicle_id,
        service_date=service_date if service_date is not None else getattr(source_record, 'service_date', None),
        odometer_km=odometer_km if odometer_km is not None else getattr(source_record, 'odometer_km', None),
        vendor=vendor if vendor is not None else getattr(source_record, 'vendor', None),
        cost=cost if cost is not None else getattr(source_record, 'cost', None),
        note=note if note is not None else getattr(source_record, 'note', None),
    )
    chosen_type = service_type
    if chosen_type is None:
        chosen_type = (
            getattr(source_record, 'service_type', None)
            or getattr(source_record, 'jenis_servis', None)
            or getattr(source_record, 'jenis_service', None)
            or 'Servis Rutin'
        )
    for attr in ('service_type', 'jenis_servis', 'jenis_service'):
        if hasattr(new_row, attr):
            setattr(new_row, attr, chosen_type)
            break
    db.session.add(new_row)
    db.session.flush()
    source_items = items
    if source_items is None:
        source_items = _service_items_for_record(source_record)
    _service_replace_items(new_row.id, source_items)
    _service_upsert_meta(new_row.id, processed_at or _now_naive())
    return new_row


def _service_processed_map(service_record_ids:list[int]):
    result = {}
    ids = [int(x) for x in (service_record_ids or []) if x]
    if not ids:
        return result
    try:
        _ensure_service_support_tables()
        rows = db.session.execute(
            text("SELECT service_record_id, processed_at FROM service_record_meta WHERE service_record_id IN :ids").bindparams(bindparam('ids', expanding=True)),
            {'ids': ids}
        ).mappings().all()
        for r in rows:
            result[r['service_record_id']] = r['processed_at']
    except Exception:
        return {}
    return result


def _service_processed_at(record):
    if not record:
        return None
    if getattr(record, 'id', None):
        processed = _service_processed_map([record.id]).get(record.id)
        if processed:
            return processed
    raw = getattr(record, 'updated_at', None) or getattr(record, 'created_at', None) or getattr(record, 'service_date', None)
    if isinstance(raw, str):
        try:
            return datetime.fromisoformat(raw)
        except Exception:
            return _parse_date(raw)
    return raw


def _parse_service_items_from_note(note_text: str):
    """
    Parse note format lama:
    [Asuransi] || Catatan: xxx || Rincian: Ganti Oli - Rp 500.000 | Filter Oli - Rp 100.000
    """
    note_text = (note_text or "").strip()
    if not note_text:
        return []

    rincian_raw = ""
    for part in note_text.split("||"):
        part = part.strip()
        if part.lower().startswith("rincian:"):
            rincian_raw = part.split(":", 1)[1].strip()
            break

    if not rincian_raw:
        return []

    items = []
    for chunk in rincian_raw.split("|"):
        chunk = chunk.strip()
        if not chunk:
            continue

        nama = chunk
        biaya = 0.0

        if " - " in chunk:
            left, right = chunk.rsplit(" - ", 1)
            nama = left.strip()
            biaya_text = right.strip().lower().replace("rp", "").replace(".", "").replace(",", ".")
            try:
                biaya = float(biaya_text)
            except Exception:
                biaya = 0.0

        items.append({
            "nama": nama,
            "vendor": "",
            "biaya": biaya,
        })
    return items


def _service_items_for_record(record):
    """
    Prioritas:
    1. items_list dari query/template helper kalau ada
    2. tabel service_record_items
    3. fallback parse dari note lama
    """
    items = []

    maybe_items = getattr(record, "items_list", None)
    if maybe_items and isinstance(maybe_items, list):
        items = _clean_service_items(maybe_items)
    if items:
        return items

    if record and getattr(record, 'id', None):
        mapped = _service_items_map([record.id]).get(record.id) or []
        items = _clean_service_items(mapped)
    if items:
        return items

    note = getattr(record, "note", None) or ""
    return _parse_service_items_from_note(note)


def _service_items_summary(record):
    parts = []
    for it in _service_items_for_record(record):
        nama = it.get('nama') or '-'
        biaya = float(it.get('biaya') or 0)
        if biaya > 0:
            parts.append(f"{nama} ({rupiah(biaya)})")
        else:
            parts.append(nama)
    return " | ".join(parts) if parts else ""


def _clean_service_note(note_text: str):
    """
    Buang prefix teknis kayak [Asuransi], Catatan:, Rincian:
    Biar di excel lebih bersih.
    """
    note_text = (note_text or "").strip()
    if not note_text:
        return "-"

    cleaned_parts = []
    for part in note_text.split("||"):
        part = part.strip()
        if not part:
            continue
        if part.startswith("[") and part.endswith("]"):
            continue
        if part.lower().startswith("rincian:"):
            continue
        if part.lower().startswith("catatan:"):
            cleaned_parts.append(part.split(":", 1)[1].strip())
        else:
            cleaned_parts.append(part)

    cleaned = " | ".join([p for p in cleaned_parts if p])
    return cleaned or "-"


def _service_export_filename(vehicle_name: str):
    safe = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in (vehicle_name or "Kendaraan"))
    safe = "_".join(safe.split())
    return f"SERVIS_{safe}_{_get_now_jkt().strftime('%Y%m%d_%H%M%S')}.xlsx"


@bp.get("/api/servis/export/<int:vehicle_id>")
def api_servis_export_excel(vehicle_id: int):
    v = Vehicle.query.get_or_404(vehicle_id)

    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))

    if date_from and date_to and date_from > date_to:
        return jsonify({"success": False, "message": "Tanggal awal tidak boleh lebih besar dari tanggal akhir."}), 400

    query = (
        ServiceRecord.query
        .filter(ServiceRecord.vehicle_id == vehicle_id)
        .order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc())
    )

    if date_from:
        query = query.filter(ServiceRecord.service_date >= date_from)
    if date_to:
        query = query.filter(ServiceRecord.service_date <= date_to)

    service_rows = query.all()

    vehicle_name = vehicle_display_name(v)
    no_polisi = vehicle_plate(v)
    pt_name = getattr(v, "pt", None) or "-"

    # Mencari servis rutin terakhir (tanpa filter tanggal) untuk info Ringkasan
    s_rutin_last = ServiceRecord.query.filter(ServiceRecord.vehicle_id == vehicle_id)
    filters_rutin = []
    for attr in ["service_type", "jenis_servis", "jenis_service"]:
        if hasattr(ServiceRecord, attr):
            filters_rutin.append(func.lower(getattr(ServiceRecord, attr)).like("%rutin%"))
    if filters_rutin:
        s_rutin_last = s_rutin_last.filter(or_(*filters_rutin))
    else:
        s_rutin_last = s_rutin_last.filter(func.lower(ServiceRecord.note).like("%rutin%"))
    s_rutin_last = s_rutin_last.order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).first()

    latest = service_rows[0] if service_rows else None
    # Next Servis dihitung dari servis rutin terakhir
    latest_due = _add_6_months(getattr(s_rutin_last, "service_date", None)) if s_rutin_last else None
    latest_status_label = "-"
    if latest_due:
        st_code = _service_status(latest_due)
        latest_status_label = {"overdue": "Terlambat", "soon": "H-30", "safe": "Aman"}.get(st_code, "-")

    total_cost_all = sum(_service_total_cost(r) for r in service_rows)
    total_rutin = sum(1 for r in service_rows if _service_record_type_label(r).lower().startswith("rutin"))
    total_berat = sum(1 for r in service_rows if _service_record_type_label(r).lower().startswith("berat"))

    # SHEET 1: RINGKASAN SERVIS (Format Vertikal)
    _nama_utama = (
        getattr(v, "active_name", None)
        or getattr(v, "new_asset_name", None)
        or getattr(v, "name_as_asset_pt", None)
        or vehicle_name
    )
    summary_data = [
        ("Nama Aset", _nama_utama or "-"),
        ("No Polisi", no_polisi or "-"),
        ("PT", pt_name),
        ("Merk", getattr(v, "merk", None) or "-"),
        ("Tipe", getattr(v, "type", None) or "-"),
        ("Jenis", getattr(v, "jenis", None) or "-"),
        ("Tahun Pemakaian", getattr(v, "year_of_use", None) or "-"),
        ("Tanggal Awal Filter", format_tgl_id_full(date_from) if date_from else "Semua"),
        ("Tanggal Akhir Filter", format_tgl_id_full(date_to) if date_to else "Semua"),
        ("Total Riwayat Servis", len(service_rows)),
        ("Total Biaya Servis", rupiah(total_cost_all)),
        ("Jumlah Servis Rutin", total_rutin),
        ("Jumlah Servis Berat", total_berat),
        ("Tanggal Servis Terakhir", format_tgl_id_full(getattr(latest, "service_date", None)) if latest else "-"),
        ("Next Servis", format_tgl_id_full(latest_due) if latest_due else "-"),
        ("Status Servis", latest_status_label),
        ("Vendor Terakhir", getattr(latest, "vendor", None) or "-"),
        ("ODO Terakhir", _service_odometer_value(latest) or "-"),
        ("Kategori Terakhir", _service_record_type_label(latest) if latest else "-"),
        ("Pembayaran Terakhir", _service_payment_label(latest) if latest else "-"),
        ("Total Biaya Terakhir", rupiah(_service_total_cost(latest)) if latest else "Rp 0"),
        ("Catatan Terakhir", _clean_service_note(getattr(latest, "note", None)) if latest else "-"),
    ]
    df_summary = pd.DataFrame(summary_data, columns=["Informasi", "Nilai"])

    # SHEET 2: RIWAYAT SERVIS
    history_rows = []
    for idx, r in enumerate(service_rows, 1):
        history_rows.append({
            "No": idx,
            "PT": pt_name,
            "Kendaraan": vehicle_name,
            "No Polisi": no_polisi,
            "Tanggal Servis": format_tgl_id_full(getattr(r, "service_date", None)) or "-",
            "Tanggal Proses": format_tgl_id_full(_service_processed_at(r)) or "-",
            "Jam Proses": _service_processed_at(r).strftime("%H:%M") if hasattr(_service_processed_at(r), "strftime") else "-",
            "Kategori Servis": _service_record_type_label(r),
            "Pembayaran": _service_payment_label(r),
            "Vendor": getattr(r, "vendor", None) or "-",
            "ODO": _service_odometer_value(r) or "-",
            "Total Biaya": rupiah(_service_total_cost(r)),
            "Catatan": _clean_service_note(getattr(r, "note", None)) or "-",
        })
    df_history = pd.DataFrame(history_rows)

    # SHEET 3: RINCIAN ITEM SERVIS
    item_rows = []
    item_no = 1
    for r in service_rows:
        items = _service_items_for_record(r)
        if items:
            for it in items:
                qty = it.get("qty") or 1
                harga = it.get("biaya") or 0
                item_rows.append({
                    "No": item_no,
                    "Tanggal Servis": format_tgl_id_full(getattr(r, "service_date", None)) or "-",
                    "Kendaraan": vehicle_name,
                    "No Polisi": no_polisi,
                    "Kategori Servis": _service_record_type_label(r),
                    "Pembayaran": _service_payment_label(r),
                    "Vendor": getattr(r, "vendor", None) or "-",
                    "Nama Item": it.get("nama") or "-",
                    "Qty": qty,
                    "Harga Satuan": rupiah(harga),
                    "Subtotal": rupiah(float(qty) * float(harga)),
                    "Catatan Servis": _clean_service_note(getattr(r, "note", None)) or "-",
                })
                item_no += 1
        else:
            item_rows.append({
                "No": item_no,
                "Tanggal Servis": format_tgl_id_full(getattr(r, "service_date", None)) or "-",
                "Kendaraan": vehicle_name,
                "No Polisi": no_polisi,
                "Kategori Servis": _service_record_type_label(r),
                "Pembayaran": _service_payment_label(r),
                "Vendor": getattr(r, "vendor", None) or "-",
                "Nama Item": "-",
                "Qty": 1,
                "Harga Satuan": rupiah(0),
                "Subtotal": rupiah(0),
                "Catatan Servis": _clean_service_note(getattr(r, "note", None)) or "-",
            })
            item_no += 1
    df_items = pd.DataFrame(item_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Ringkasan Servis")
        df_history.to_excel(writer, index=False, sheet_name="Riwayat Servis")
        df_items.to_excel(writer, index=False, sheet_name="Rincian Item Servis")

        # Formatting
        for ws in writer.book.worksheets:
            # Header Styling
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True, color="FFD15F")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="1E4080")
            
            # Auto Width
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            val_len = len(str(cell.value))
                            if val_len > max_length: max_length = val_len
                    except: pass
                ws.column_dimensions[column_letter].width = min(max(max_length + 2, 15), 50)

    output.seek(0)
    
    # Filename Formatting
    pt_clean = (pt_name or "Semua_PT").replace(" ", "_")
    export_date = date.today().strftime("%d-%m-%Y")
    filename = f"Laporan_Servis_{pt_clean}_{export_date}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.get("/data/kir",endpoint="go_data_kir")
def go_data_kir():
    q=(request.args.get("q") or "").strip()
    status=(request.args.get("status") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()

    vehicles=(
        _vehicle_query_filtered(q,company)
        .order_by(Vehicle.updated_at.desc(),Vehicle.id.desc())
        .limit(1000)
        .all()
    )

    latest_kir_map=_latest_kir_map([v.id for v in vehicles])
    rows=[]
    for i,v in enumerate(vehicles,1):
        r=latest_kir_map.get(v.id)
        due=getattr(r,"due_date",None) if r else None
        current_status=_kir_status(due)
        rows.append({
            "no":i,
            "vehicle_id":v.id,
            "kir_id":getattr(r,"id",None) if r else None,
            "pt":v.pt or "-",
            "nama_utama":getattr(v,"active_name",None) or getattr(v,"new_asset_name",None) or getattr(v,"name_as_asset_pt",None) or vehicle_display_name(v),
            "kendaraan":vehicle_display_name(v),
            "no_polisi":vehicle_plate(v),
            "tgl_kir":getattr(r,"done_date",None) if r else None,
            "next_kir":due,
            "status_kir":current_status,
            "hasil":_kir_display_result(r),
            "note":getattr(r,"note",None) if r else None,
            "vehicle":v,
        })

    if status:
        rows=[r for r in rows if r["status_kir"]==status]
        for idx,r in enumerate(rows,1):
            r["no"]=idx

    highlight_id = _parse_int(request.args.get("highlight"))
    return render_template(
        "data_kir.html",
        kir_rows=rows,
        q=q,
        status=status,
        company=company,
        companies=companies,
        highlight_id=highlight_id,
        title="Data KIR - Sinar Group",
    )

@bp.post("/data/kir/tambah")
@master_required
def data_kir_tambah():
    vehicle_id=_parse_int(request.form.get("vehicle_id"))
    v=Vehicle.query.get_or_404(vehicle_id)
    # done_date = tgl stempel KIR (patokan next KIR)
    # pay_date  = tgl bayar administrasi (boleh beda, tidak mempengaruhi next KIR)
    done_date=_parse_date_as_datetime(request.form.get("done_date"))
    pay_date=_parse_date(request.form.get("pay_date")) or None

    if not done_date:
        flash("Tanggal KIR selesai wajib diisi.", "danger")
        return redirect(url_for("main.go_data_kir"))

    # Next KIR selalu dari done_date (tgl KIR selesai), bukan pay_date
    due_date = _add_months_safe(done_date, 6)

    payload=_request_payload()
    result=_normalize_text(request.form.get("result")) or None
    status_val=_normalize_text(payload.get("status")) or None
    note=_normalize_text(request.form.get("note")) or None

    row=KirRecord(
        vehicle_id=v.id,
        due_date=due_date,
        done_date=done_date,
        pay_date=pay_date,
        result=result,
        status=status_val,
        note=note,
        created_at=_now_naive()
    )
    _assign_model_timestamp(row, "created_at", ("updated_at", "processed_at",))
    Vehicle.query.session.add(row)
    try:audit("CREATE","KirRecord",0,note=f"Tambah KIR kendaraan {vehicle_plate(v)}")
    except Exception:pass
    return _safe_commit("Proses KIR berhasil disimpan dan masuk ke history KIR.","Gagal simpan KIR","main.go_data_kir")

@bp.post("/data/kir/<int:kir_id>/delete")
@master_required
def data_kir_delete(kir_id:int):
    row=KirRecord.query.get_or_404(kir_id)
    try:audit("DELETE","KirRecord",kir_id,note=f"Hapus KIR kendaraan {vehicle_plate(getattr(row,'vehicle',None))}")
    except Exception:pass
    Vehicle.query.session.delete(row)
    return _safe_commit("Data KIR berhasil dihapus.","Gagal hapus KIR","main.go_data_kir")

@bp.post("/api/kir/update-tgl-note")
@master_required
def api_kir_update_tgl_note():
    data = request.get_json() or {}
    vehicle_id = _parse_int(data.get("vehicle_id"))
    tgl_kir_str = (data.get("tgl_kir") or "").strip()
    note = _normalize_text(data.get("note")) or None

    if not vehicle_id:
        return jsonify({"success": False, "error": "Vehicle ID is required"}), 400

    # Cari data KIR terakhir berdasarkan ID (id.desc()) agar kita mengedit data yang tampil di tabel utama, bukan menambah baris baru di history.
    latest = KirRecord.query.filter_by(vehicle_id=vehicle_id).order_by(KirRecord.id.desc()).first()
    
    done_date = _parse_date_as_datetime(tgl_kir_str) if tgl_kir_str else None
    if tgl_kir_str and not done_date:
        return jsonify({"success": False, "error": "Format tanggal KIR tidak valid"}), 400

    if not latest:
        # Jika benar-benar belum ada data KIR sama sekali untuk kendaraan ini, baru kita buat satu.
        if not done_date and not note:
            return jsonify({"success": True})

        latest = KirRecord(
            vehicle_id=vehicle_id,
            done_date=done_date,
            due_date=_add_months_safe(done_date, 6) if done_date else None,
            note=note,
            result=None,
            status=None,
            created_at=_now_naive()
        )
        _assign_model_timestamp(latest, "created_at", ("updated_at", "processed_at",))
        db.session.add(latest)
    else:
        # EDIT DATA YANG ADA: Kita ganti data yang sudah ada di tabel utama, bukan tambah ke history.
        # Dan tetap hitung ulang jatuh tempo berikutnya (rolling) agar otomatis.
        latest.done_date = done_date
        latest.due_date = _add_months_safe(done_date, 6) if done_date else latest.due_date
        latest.note = note
        _assign_model_timestamp(latest, "updated_at", ("processed_at",))
        # Edit tanggal/catatan bukan proses KIR, jadi jangan otomatis isi Hasil/Status.
        # Bersihkan data legacy yang dulu pernah otomatis menjadi Lulus/Selesai
        # saat hanya edit tanggal awal.
        if (
            _normalize_text(getattr(latest, "result", None)).lower() == "lulus"
            and _normalize_text(getattr(latest, "status", None)).lower() == "selesai"
            and not getattr(latest, "pay_date", None)
            and not note
        ):
            latest.result = None
            latest.status = None

    try:
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@bp.get("/api/kir/history/<int:vehicle_id>")
def api_kir_history(vehicle_id:int):
    v = Vehicle.query.get_or_404(vehicle_id)
    vehicle_name = vehicle_display_name(v)
    plate = vehicle_plate(v)

    rows = (
        KirRecord.query
        .filter(KirRecord.vehicle_id == vehicle_id)
        .order_by(
            KirRecord.done_date.is_(None),
            KirRecord.done_date.desc(),
            KirRecord.due_date.is_(None),
            KirRecord.due_date.desc(),
            KirRecord.id.desc(),
        )
        .all()
    )

    history = []
    lulus = 0
    gagal = 0

    for r in rows:
        result_text = _normalize_text(getattr(r, "result", None))
        status_text = _normalize_text(getattr(r, "status", None))
        note_text = _normalize_text(getattr(r, "note", None))

        if result_text.lower() == "lulus":
            lulus += 1
        elif result_text.lower() in {"tidak lulus", "gagal"}:
            gagal += 1

        history.append({
            "date": format_tgl_id_full(getattr(r, "done_date", None) or getattr(r, "due_date", None)) or "-",
            "text": (
                f"Tanggal KIR: {format_tgl_id_full(getattr(r, 'done_date', None)) or '-'} • "
                f"Jatuh tempo: {format_tgl_id_full(getattr(r, 'due_date', None)) or '-'} • "
                f"Hasil: {result_text or '-'} • "
                f"Status: {status_text or '-'} • "
                f"Catatan: {note_text or '-'}"
            ),
        })

    latest = rows[0] if rows else None
    latest_due = getattr(latest, "due_date", None) if latest else None

    return jsonify({
        "success": True,
        "vehicle": vehicle_name,
        "plate": plate,
        "summary": {
            "total": len(rows),
            "lulus": lulus,
            "gagal": gagal,
            "tglTerakhir": format_tgl_id_full(getattr(latest, "done_date", None)) if latest else "-",
            "jatuhTempo": format_tgl_id_full(latest_due) if latest_due else "-",
            "status": _kir_status(latest_due) if latest_due else "empty",
        },
        "history": history,
    })


@bp.get("/api/kir/export/<int:vehicle_id>")
def api_kir_export(vehicle_id: int):
    """
    Export Excel data KIR per kendaraan
    Support:
    - date_from / date_to (optional, default kosong = semua data)
    - export_type:
        - all
        - summary
        - history
    Format filename disamakan dengan Pajak.
    """
    v = Vehicle.query.get_or_404(vehicle_id)

    export_type = (_normalize_text(request.args.get("export_type")) or "all").lower()
    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))

    if date_from and date_to and date_from > date_to:
        return jsonify({
            "success": False,
            "message": "Tanggal awal tidak boleh lebih besar dari tanggal akhir."
        }), 400

    query = (
        KirRecord.query
        .filter(KirRecord.vehicle_id == vehicle_id)
        .order_by(
            KirRecord.done_date.is_(None),
            KirRecord.done_date.desc(),
            KirRecord.due_date.is_(None),
            KirRecord.due_date.desc(),
            KirRecord.id.desc(),
        )
    )

    if date_from:
        query = query.filter(
            or_(
                KirRecord.done_date >= date_from,
                db.and_(KirRecord.done_date.is_(None), KirRecord.due_date >= date_from),
            )
        )
    if date_to:
        query = query.filter(
            or_(
                KirRecord.done_date <= date_to,
                db.and_(KirRecord.done_date.is_(None), KirRecord.due_date <= date_to),
            )
        )

    kir_rows = query.all()

    latest = kir_rows[0] if kir_rows else None
    latest_due = getattr(latest, "due_date", None) if latest else None
    latest_done = getattr(latest, "done_date", None) if latest else None

    total_lulus = 0
    total_gagal = 0
    history_rows = []

    for idx, r in enumerate(kir_rows, 1):
        due_date = getattr(r, "due_date", None)
        done_date = getattr(r, "done_date", None)
        result_text = _normalize_text(getattr(r, "result", None))
        status_text = _normalize_text(getattr(r, "status", None))
        note_text = _normalize_text(getattr(r, "note", None))

        if result_text.lower() == "lulus":
            total_lulus += 1
        elif result_text.lower() in {"tidak lulus", "gagal"}:
            total_gagal += 1

        history_rows.append({
            "No": idx,
            "PT": _export_scalar(v.pt),
            "Nama Kendaraan": vehicle_display_name(v),
            "No Polisi": vehicle_plate(v),
            "Tanggal KIR": _export_date_value(done_date),
            "Jatuh Tempo KIR": _export_date_value(due_date),
            "Status KIR": _export_scalar(_kir_status(due_date)),
            "Hasil KIR": _export_scalar(result_text),
            "Status Proses": _export_scalar(status_text),
            "Catatan KIR": _export_scalar(note_text),
        })

    _nama_utama = (
        getattr(v, "active_name", None)
        or getattr(v, "new_asset_name", None)
        or getattr(v, "name_as_asset_pt", None)
        or vehicle_display_name(v)
    )
    summary_data = [
        ("Nama Aset", _nama_utama or "-"),
        ("No Polisi", vehicle_plate(v) or "-"),
        ("PT", _export_scalar(v.pt)),
        ("Merk", _export_scalar(getattr(v, "merk", None))),
        ("Tipe", _export_scalar(getattr(v, "type", None))),
        ("Jenis", _export_scalar(getattr(v, "jenis", None))),
        ("Tahun Pemakaian", _export_scalar(getattr(v, "year_of_use", None))),
        ("Tanggal Awal Filter", format_tgl_id_full(date_from) if date_from else "Semua"),
        ("Tanggal Akhir Filter", format_tgl_id_full(date_to) if date_to else "Semua"),
        ("Total Riwayat KIR", len(kir_rows)),
        ("Total KIR Lulus", total_lulus),
        ("Total KIR Gagal", total_gagal),
        ("Tanggal KIR Terakhir", _export_date_value(latest_done)),
        ("Jatuh Tempo KIR Aktif", _export_date_value(latest_due)),
        ("Status KIR Aktif", _export_scalar(_kir_status(latest_due) if latest_due else "empty")),
        ("Hasil KIR Terakhir", _export_scalar(getattr(latest, "result", None) if latest else None)),
        ("Catatan KIR Terakhir", _export_scalar(getattr(latest, "note", None) if latest else None)),
    ]
    df_summary = pd.DataFrame(summary_data, columns=["Informasi", "Nilai"])

    if not history_rows:
        history_rows = [{
            "No": "-",
            "PT": _export_scalar(v.pt),
            "Nama Kendaraan": vehicle_display_name(v),
            "No Polisi": vehicle_plate(v),
            "Tanggal KIR": "-",
            "Jatuh Tempo KIR": "-",
            "Status KIR": "empty",
            "Hasil KIR": "-",
            "Status Proses": "-",
            "Catatan KIR": "-",
        }]

    df_history = pd.DataFrame(history_rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if export_type in {"all", "summary"}:
            df_summary.to_excel(writer, index=False, sheet_name="Ringkasan KIR")
        if export_type in {"all", "history"}:
            df_history.to_excel(writer, index=False, sheet_name="Riwayat KIR")

        for ws in writer.book.worksheets:
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True, color="FFD15F")
                    cell.fill = cell.fill.copy(fill_type="solid", fgColor="1E4080")
            _autosize_worksheet_columns(ws, min_width=14, max_width=38)

    output.seek(0)
    safe_name = _safe_export_filename_part(vehicle_display_name(v), "KENDARAAN")
    safe_plate = _safe_export_filename_part(vehicle_plate(v), "-")
    period_suffix = []
    if date_from:
        period_suffix.append(date_from.strftime("%Y-%m-%d"))
    if date_to:
        period_suffix.append(date_to.strftime("%Y-%m-%d"))
    period_text = f"_{'_to_'.join(period_suffix)}" if period_suffix else "_semua-periode"
    filename = f"KIR_{safe_name.replace(' ', '_')}_{safe_plate.replace(' ', '_')}{period_text}_{date.today().strftime('%d-%m-%Y')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _loan_owner_pt(vehicle:Vehicle|None)->str:
    """PT pemilik/asal peminjaman selalu diambil dari data kendaraan.
    Ini menjaga agar mobil milik PT A tidak tercatat sebagai mobil dari PT lain."""
    return _normalize_text(getattr(vehicle, "pt", None)) if vehicle else ""


def _active_loan_for_vehicle(vehicle_id:int, exclude_loan_id:int|None=None):
    query = LoanTransaction.query.filter(
        LoanTransaction.vehicle_id == vehicle_id,
        LoanTransaction.date_return_actual.is_(None),
        func.lower(func.coalesce(LoanTransaction.status, "")) != "selesai",
    )
    if exclude_loan_id:
        query = query.filter(LoanTransaction.id != exclude_loan_id)
    return query.order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).first()


@bp.get("/data/peminjaman",endpoint="peminjaman")
def peminjaman():
    q=(request.args.get("q") or "").strip()
    status=(request.args.get("status") or "").strip().lower()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()
    all_companies=_vehicle_pt_choices()

    # Filter peminjaman selalu berdasarkan PT kendaraan (owner/aset), bukan berdasarkan semua PT peminjam.
    vehicle_ids=_filtered_vehicle_ids(q,company)
    lq=LoanTransaction.query.join(Vehicle,Vehicle.id==LoanTransaction.vehicle_id).filter(Vehicle.is_deleted==False)
    lq=_apply_vehicle_scope(lq)
    if vehicle_ids is not None:
        lq=lq.filter(LoanTransaction.vehicle_id.in_(vehicle_ids))
    if q:
        like=f"%{q}%"
        lq=lq.filter(or_(
            Vehicle.pt.ilike(like), Vehicle.active_name.ilike(like), Vehicle.new_asset_name.ilike(like),
            Vehicle.name_as_asset_pt.ilike(like), Vehicle.merk.ilike(like), Vehicle.type.ilike(like),
            Vehicle.plate_old.ilike(like), Vehicle.plate_new.ilike(like),
            LoanTransaction.borrower_company.ilike(like), LoanTransaction.borrower_name.ilike(like),
            LoanTransaction.purpose.ilike(like), LoanTransaction.note.ilike(like)
        ))
    if company:
        lq=lq.filter(func.lower(func.trim(Vehicle.pt))==company.strip().lower())
    # Status UI memakai hitungan tanggal seperti Data KIR:
    # overdue = telat, soon = dekat jatuh tempo, borrowed = aktif, done = selesai.
    if status == "done":
        lq = lq.filter(or_(func.lower(func.coalesce(LoanTransaction.status, "")) == "selesai", LoanTransaction.date_return_actual.isnot(None)))
    elif not status:
        lq = lq.filter(func.lower(func.coalesce(LoanTransaction.status,""))!="selesai", LoanTransaction.date_return_actual.is_(None))
    else:
        lq = lq.filter(func.lower(func.coalesce(LoanTransaction.status,""))!="selesai", LoanTransaction.date_return_actual.is_(None))

    items=lq.order_by(LoanTransaction.created_at.desc(),LoanTransaction.id.desc()).limit(1000).all()
    today_value = date.today()
    for it in items:
        if getattr(it, "date_return_actual", None) or _normalize_text(getattr(it, "status", None)).lower() == "selesai":
            it.status_view = "done"
        else:
            due = getattr(it, "date_return_plan", None)
            if due and due < today_value:
                it.status_view = "overdue"
            elif due and (due - today_value).days <= 7:
                it.status_view = "soon"
            else:
                it.status_view = "borrowed"
    if status and status != "done":
        items = [it for it in items if getattr(it, "status_view", "") == status]

    loan_place_map=_loan_place_map([it.id for it in items]) if items else {}

    # Dropdown kendaraan di modal tambah selalu load SEMUA kendaraan (tanpa filter PT tabel).
    # Filter berdasarkan "Dari PT" dilakukan di frontend JS saat user pilih PT pemilik kendaraan.
    vehicle_query=_apply_vehicle_scope(_active_vehicle_base_query())
    vehicle_choices=vehicle_query.order_by(Vehicle.pt.asc(),Vehicle.updated_at.desc(),Vehicle.id.desc()).limit(2000).all()
    highlight_id = _parse_int(request.args.get("highlight"))
    return render_template("data_peminjaman.html",items=items,q=q,status=status,company=company,companies=companies,all_companies=all_companies,vehicle_choices=vehicle_choices,loan_place_map=loan_place_map,today=date.today(),highlight_id=highlight_id,title="Data Peminjaman - Sinar Group")


@bp.get("/data/peminjaman/export")
@master_required
def peminjaman_export():
    """Bulk export Excel semua data peminjaman sesuai filter aktif (q, company, status).
    Menghasilkan dua sheet: Ringkasan per PT dan Riwayat Lengkap."""
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip().lower()
    company = _sanitize_company_filter((request.args.get("company") or "").strip())
    date_from = _parse_date(request.args.get("date_from"))
    date_to   = _parse_date(request.args.get("date_to"))

    # ── Query identik dengan halaman peminjaman ──
    vehicle_ids = _filtered_vehicle_ids(q, company)
    lq = LoanTransaction.query.join(Vehicle, Vehicle.id == LoanTransaction.vehicle_id).filter(Vehicle.is_deleted == False)
    lq = _apply_vehicle_scope(lq)
    if vehicle_ids is not None:
        lq = lq.filter(LoanTransaction.vehicle_id.in_(vehicle_ids))
    if q:
        like = f"%{q}%"
        lq = lq.filter(or_(
            Vehicle.pt.ilike(like), Vehicle.active_name.ilike(like), Vehicle.new_asset_name.ilike(like),
            Vehicle.name_as_asset_pt.ilike(like), Vehicle.merk.ilike(like), Vehicle.type.ilike(like),
            Vehicle.plate_old.ilike(like), Vehicle.plate_new.ilike(like),
            LoanTransaction.borrower_company.ilike(like), LoanTransaction.borrower_name.ilike(like),
            LoanTransaction.note.ilike(like)
        ))
    if company:
        lq = lq.filter(func.lower(func.trim(Vehicle.pt)) == company.strip().lower())
    if status:
        lq = lq.filter(func.lower(LoanTransaction.status) == status)
    # default: hanya aktif (belum selesai) — jika tidak ada status filter, export semua
    if date_from:
        lq = lq.filter(LoanTransaction.date_out >= date_from)
    if date_to:
        lq = lq.filter(LoanTransaction.date_out <= date_to)

    rows = lq.order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).all()

    if not rows:
        return jsonify({"ok": False, "message": "Tidak ada data untuk di-export."}), 404

    # ── Sheet 1: Ringkasan per PT ──
    from collections import defaultdict
    pt_summary = defaultdict(lambda: {"total": 0, "dipinjam": 0, "terlambat": 0, "selesai": 0})
    for r in rows:
        pt = (getattr(r.vehicle, "pt", None) or "-").strip()
        pt_summary[pt]["total"] += 1
        sl = _loan_status_label(r).lower()
        if "selesai" in sl:
            pt_summary[pt]["selesai"] += 1
        elif "terlambat" in sl:
            pt_summary[pt]["terlambat"] += 1
        else:
            pt_summary[pt]["dipinjam"] += 1

    summary_rows = []
    for pt, s in sorted(pt_summary.items()):
        summary_rows.append({
            "PT": pt,
            "Total Peminjaman": s["total"],
            "Sedang Dipinjam": s["dipinjam"],
            "Terlambat": s["terlambat"],
            "Selesai": s["selesai"],
        })
    summary_rows.append({
        "PT": "TOTAL",
        "Total Peminjaman": sum(s["total"] for s in pt_summary.values()),
        "Sedang Dipinjam": sum(s["dipinjam"] for s in pt_summary.values()),
        "Terlambat": sum(s["terlambat"] for s in pt_summary.values()),
        "Selesai": sum(s["selesai"] for s in pt_summary.values()),
    })
    df_summary = pd.DataFrame(summary_rows)

    # ── Sheet 2: Riwayat Lengkap ──
    history_rows = []
    for idx, r in enumerate(rows, 1):
        v = r.vehicle
        history_rows.append({
            "No": idx,
            "PT Kendaraan": (getattr(v, "pt", None) or "-").strip(),
            "Nama Aset": vehicle_display_name(v),
            "No Polisi": vehicle_plate(v),
            "Merk": _export_scalar(getattr(v, "merk", None)),
            "Type": _export_scalar(getattr(v, "type", None)),
            "Dari PT": _export_scalar(getattr(r, "borrower_company", None)),
            "Ke PT": _export_scalar(getattr(r, "borrower_name", None)),
            "Tanggal Pinjam": _export_date_value(getattr(r, "date_out", None)),
            "Rencana Kembali": _export_date_value(getattr(r, "date_return_plan", None)),
            "Aktual Kembali": _export_date_value(getattr(r, "date_return_actual", None)),
            "Status": _loan_status_label(r),
            "Catatan": _export_scalar(getattr(r, "note", None)),
        })
    df_history = pd.DataFrame(history_rows)

    # ── Build Excel ──
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Ringkasan PT")
        df_history.to_excel(writer, index=False, sheet_name="Riwayat Peminjaman")

        for ws in writer.book.worksheets:
            # Header styling
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True, color="FFD15F")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="1E4080")
            # Baris TOTAL di sheet ringkasan — bold
            if ws.title == "Ringkasan PT":
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = cell.font.copy(bold=True)
            _autosize_worksheet_columns(ws, min_width=14, max_width=45)

    output.seek(0)
    pt_clean = (company or "Semua_PT").replace(" ", "_")
    export_date = date.today().strftime("%d-%m-%Y")
    filename = f"Data_Peminjaman_{pt_clean}_{export_date}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@bp.get("/data/peminjaman/api/kendaraan-by-pt")
@master_required
def peminjaman_kendaraan_by_pt():
    """AJAX endpoint: return daftar kendaraan berdasarkan PT pemilik (Dari PT).
    Dipakai di modal Tambah Peminjaman untuk filter kendaraan real-time."""
    pt = _sanitize_company_filter((request.args.get("pt") or "").strip())
    q = (request.args.get("q") or "").strip()
    base = _apply_vehicle_scope(_active_vehicle_base_query())
    if pt:
        base = base.filter(func.lower(func.trim(Vehicle.pt)) == pt.strip().lower())
    if q:
        like = f"%{q}%"
        base = base.filter(or_(
            Vehicle.active_name.ilike(like), Vehicle.new_asset_name.ilike(like),
            Vehicle.name_as_asset_pt.ilike(like), Vehicle.merk.ilike(like),
            Vehicle.type.ilike(like), Vehicle.plate_old.ilike(like), Vehicle.plate_new.ilike(like)
        ))
    vehicles = base.order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).limit(500).all()
    results = []
    for v in vehicles:
        results.append({
            "id": v.id,
            "label": vehicle_display_name(v),
            "plate": vehicle_plate(v),
            "pt": (v.pt or "").strip(),
        })
    return jsonify({"ok": True, "vehicles": results})

@bp.post("/data/peminjaman/tambah")
@master_required
def peminjaman_tambah():
    vehicle_id=_parse_int(request.form.get("vehicle_id"))
    if not vehicle_id:
        flash("Kendaraan wajib dipilih dari data kendaraan utama.","danger")
        return redirect(url_for("main.peminjaman"))
    v=Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)
    session_db=Vehicle.query.session
    date_out=_parse_date_as_datetime(request.form.get("date_out")) or _today_as_datetime()
    date_return_plan=_parse_date(request.form.get("date_return_plan"))
    if not date_return_plan:
        flash("Tanggal rencana dikembalikan wajib diisi.","danger")
        return redirect(url_for("main.peminjaman"))
    if date_return_plan < (date_out.date() if hasattr(date_out, "date") else date_out):
        flash("Tanggal rencana dikembalikan tidak boleh lebih awal dari tanggal pinjam.","danger")
        return redirect(url_for("main.peminjaman"))
    borrower_company = _loan_owner_pt(v)
    if not borrower_company:
        flash("PT kendaraan belum terisi di data kendaraan. Isi kolom PT kendaraan dulu sebelum membuat peminjaman.","danger")
        return redirect(url_for("main.peminjaman"))
    borrower_name = _normalize_text(request.form.get("borrower_name"))
    if not borrower_name:
        flash("PT Peminjam (Ke PT) wajib dipilih.","danger")
        return redirect(url_for("main.peminjaman"))
    active_loan = _active_loan_for_vehicle(v.id)
    if active_loan:
        flash("Kendaraan ini masih dalam peminjaman aktif. Selesaikan peminjaman sebelumnya dulu.","danger")
        return redirect(url_for("main.peminjaman", company=borrower_company))
    
    row=LoanTransaction(
        vehicle_id=v.id,
        borrower_name=borrower_name or "-",
        borrower_company=borrower_company,

        date_out=date_out,
        date_return_plan=date_return_plan,
        date_return_actual=None,
        status="dipinjam",
        note=_normalize_text(request.form.get("note")) or None
    )
    session_db.add(row)
    session_db.flush()

    loan_history_row = LoanHistory(loan_id=row.id,vehicle_id=v.id,action_type="create",old_status=None,new_status=row.status,old_borrower=None,new_borrower=row.borrower_name,changed_by=session.get("user_name") or "SYSTEM",note="Peminjaman PT ke PT dibuat")
    _assign_model_timestamp(loan_history_row, "changed_at", ("created_at", "updated_at"))
    session_db.add(loan_history_row)
    try:audit("CREATE","LoanTransaction",row.id,note=f"Tambah peminjaman kendaraan {vehicle_plate(v)}")
    except Exception:pass
    return _safe_commit("Data peminjaman berhasil ditambahkan.","Gagal tambah peminjaman","main.peminjaman")

@bp.post("/data/peminjaman/<int:loan_id>/update")
@master_required
def peminjaman_update(loan_id:int):
    row=LoanTransaction.query.get_or_404(loan_id)
    if getattr(row, "vehicle", None):
        _assert_vehicle_scope(row.vehicle)
    session_db=Vehicle.query.session
    old_status=row.status
    old_borrower=row.borrower_name
    submitted_date_out=_parse_date_as_datetime(request.form.get("date_out")) or row.date_out
    submitted_return_plan=_parse_date(request.form.get("date_return_plan"))
    submitted_return_actual=_parse_date_as_datetime(request.form.get("date_return_actual"))
    payload=_request_payload()
    submitted_status=_normalize_text(payload.get("status")) or row.status
    if submitted_return_plan and submitted_date_out and submitted_return_plan < (submitted_date_out.date() if hasattr(submitted_date_out, "date") else submitted_date_out):
        flash("Tanggal rencana dikembalikan tidak boleh lebih awal dari tanggal pinjam.","danger")
        return redirect(url_for("main.peminjaman"))
    if submitted_return_actual and submitted_date_out and submitted_return_actual < submitted_date_out:
        flash("Tanggal aktual kembali tidak boleh lebih awal dari tanggal pinjam.","danger")
        return redirect(url_for("main.peminjaman"))
    borrower_company = _loan_owner_pt(getattr(row, "vehicle", None))
    if not borrower_company:
        flash("PT kendaraan belum terisi di data kendaraan. Isi kolom PT kendaraan dulu sebelum update peminjaman.","danger")
        return redirect(url_for("main.peminjaman"))
    borrower_name = _normalize_text(request.form.get("borrower_name"))
    if not borrower_name:
        flash("PT Peminjam (Ke PT) wajib dipilih.","danger")
        return redirect(url_for("main.peminjaman"))
    active_loan = _active_loan_for_vehicle(row.vehicle_id, exclude_loan_id=row.id)
    if active_loan and (_normalize_text(submitted_status).lower() != "selesai") and not submitted_return_actual:
        flash("Kendaraan ini masih punya peminjaman aktif lain. Tidak bisa membuat dua peminjaman aktif untuk kendaraan yang sama.","danger")
        return redirect(url_for("main.peminjaman", company=borrower_company))
    
    row.borrower_name=borrower_name or row.borrower_name
    row.borrower_company=borrower_company

    row.date_out=submitted_date_out
    row.date_return_plan=submitted_return_plan
    row.date_return_actual=submitted_return_actual
    row.status=submitted_status
    row.note=_normalize_text(request.form.get("note")) or None

    row.updated_at=_now_naive()
    loan_history_row = LoanHistory(loan_id=row.id,vehicle_id=row.vehicle_id,action_type="update",old_status=old_status,new_status=row.status,old_borrower=old_borrower,new_borrower=row.borrower_name,changed_by=session.get("user_name") or "SYSTEM",note="Update peminjaman PT ke PT")
    _assign_model_timestamp(loan_history_row, "changed_at", ("created_at", "updated_at"))
    session_db.add(loan_history_row)
    try:audit("UPDATE","LoanTransaction",row.id,note=f"Update peminjaman id={row.id}")
    except Exception:pass
    return _safe_commit("Data peminjaman berhasil diupdate.","Gagal update peminjaman","main.peminjaman")

@bp.post("/data/peminjaman/<int:loan_id>/selesai")
@master_required
def peminjaman_selesai(loan_id:int):
    row=LoanTransaction.query.get_or_404(loan_id)
    if getattr(row, "vehicle", None):
        _assert_vehicle_scope(row.vehicle)
    session_db=Vehicle.query.session
    old_status=row.status
    old_borrower=row.borrower_name
    date_return_actual=_parse_date_as_datetime(request.form.get("date_return_actual")) or _today_as_datetime()
    if row.date_out and date_return_actual < row.date_out:
        flash("Tanggal kembali aktual tidak boleh lebih awal dari tanggal pinjam.","danger")
        return redirect(url_for("main.peminjaman"))
    row.date_return_actual=date_return_actual
    row.status="selesai"
    row.updated_at=_now_naive()
    finish_note=_normalize_text(request.form.get("finish_note"))
    if finish_note:
        base_note=_normalize_text(row.note)
        row.note=(base_note+" | "+finish_note).strip(" |") if base_note else finish_note
    loan_history_row = LoanHistory(loan_id=row.id,vehicle_id=row.vehicle_id,action_type="finish",old_status=old_status,new_status="selesai",old_borrower=old_borrower,new_borrower=row.borrower_name,changed_by=session.get("user_name") or "SYSTEM",note="Peminjaman PT ke PT selesai / kendaraan dikembalikan")
    _assign_model_timestamp(loan_history_row, "changed_at", ("created_at", "updated_at"))
    session_db.add(loan_history_row)
    try:audit("FINISH","LoanTransaction",row.id,note=f"Selesaikan peminjaman id={row.id}")
    except Exception:pass
    return _safe_commit("Peminjaman selesai dan data masuk ke history peminjaman.","Gagal menyelesaikan peminjaman","main.peminjaman")

@bp.get("/history/kendaraan")
@master_required
def history_kendaraan():
    q=(request.args.get("q") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()
    query=_apply_vehicle_scope(VehicleChangeHistory.query.join(Vehicle,Vehicle.id==VehicleChangeHistory.vehicle_id).filter(Vehicle.is_deleted==False))
    if q:
        like=f"%{q}%"
        query=query.filter(or_(Vehicle.pt.ilike(like),Vehicle.active_name.ilike(like),Vehicle.new_asset_name.ilike(like),Vehicle.name_as_asset_pt.ilike(like),Vehicle.plate_old.ilike(like),Vehicle.plate_new.ilike(like),VehicleChangeHistory.field_label.ilike(like),VehicleChangeHistory.old_value.ilike(like),VehicleChangeHistory.new_value.ilike(like),VehicleChangeHistory.note.ilike(like)))
    if company:query=query.filter(func.lower(func.trim(Vehicle.pt))==company.strip().lower())
    rows=query.order_by(VehicleChangeHistory.changed_at.desc(),VehicleChangeHistory.id.desc()).limit(1000).all()
    items=[{"log":r,"vehicle":Vehicle.query.get(r.vehicle_id),"payload":None} for r in rows]
    return render_template("history_kendaraan.html",items=items,q=q,company=company,companies=companies,title="Riwayat Edit Kendaraan - Sinar Group")

@bp.get("/history/pembayaran")
@master_required
def history_pembayaran():
    _ensure_annual_tax_table()
    q=(request.args.get("q") or "").strip()
    jenis=(request.args.get("jenis") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()
    vehicles=_vehicle_query_filtered(q,company).order_by(Vehicle.updated_at.desc(),Vehicle.id.desc()).limit(1000).all()
    vehicle_map={v.id:v for v in vehicles}
    vehicle_ids=list(vehicle_map.keys())
    items=[]
    if vehicle_ids and jenis in ("","annual"):
        annual_rows=(db.session.execute(text("SELECT id, vehicle_id, tax_year, due_date, paid_date, amount, note FROM annual_tax_payments WHERE vehicle_id IN :ids AND paid_date IS NOT NULL ORDER BY paid_date DESC, id DESC").bindparams(bindparam("ids", expanding=True)),{"ids":vehicle_ids}).mappings().all())
        for r in annual_rows:
            v=vehicle_map.get(r["vehicle_id"])
            if not v:continue
            items.append({"kind":"annual","kind_label":"Pajak Tahunan","vehicle":v,"due_date":r["due_date"],"paid_date":r["paid_date"],"amount":r["amount"] or 0,"note":r["note"],"payment_status_label":"Berhasil lunas","payment_status_class":"safe"})
    if vehicle_ids and jenis in ("","five"):
        five_rows=(db.session.execute(text("SELECT id, vehicle_id, due_date, paid_date, amount, plate_before, plate_after, note FROM five_year_tax_payments WHERE vehicle_id IN :ids AND paid_date IS NOT NULL ORDER BY paid_date DESC, id DESC").bindparams(bindparam("ids", expanding=True)),{"ids":vehicle_ids}).mappings().all())
        for r in five_rows:
            v=vehicle_map.get(r["vehicle_id"])
            if not v:continue
            plate_before=_normalize_text(r.get("plate_before")) or _normalize_text(getattr(v,"plate_old",None))
            plate_after_saved=_normalize_text(r.get("plate_after"))
            current_plate_new=_normalize_text(getattr(v,"plate_new",None))
            plate_after=plate_after_saved or (current_plate_new if current_plate_new and current_plate_new != plate_before else None)
            is_plate_updated=bool(plate_after)
            items.append({
                "kind":"five",
                "kind_label":"Pajak 5 Tahunan",
                "vehicle":v,
                "due_date":r["due_date"],
                "paid_date":r["paid_date"],
                "amount":r["amount"] or 0,
                "note":r["note"],
                "plate_before":plate_before or None,
                "plate_after":plate_after or None,
                "payment_status_label":"Berhasil lunas" if is_plate_updated else "Lunas, menunggu update plat",
                "payment_status_class":"safe" if is_plate_updated else "soon",
                "plate_update_pending":not is_plate_updated,
            })
    items.sort(key=lambda x:(x.get("paid_date") or date.min,x.get("due_date") or date.min),reverse=True)
    return render_template("history_pembayaran.html",items=items[:1200],q=q,jenis=jenis,company=company,companies=companies,title="Riwayat Pembayaran - Sinar Group")

@bp.get("/history/servis")
@master_required
def history_servis():
    _ensure_service_support_tables()
    q=(request.args.get("q") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    status=(request.args.get("status") or "").strip()
    companies=_scoped_company_choices()
    vehicle_ids=_filtered_vehicle_ids(q,company)
    sq=ServiceRecord.query
    if vehicle_ids is not None:sq=sq.filter(ServiceRecord.vehicle_id.in_(vehicle_ids))
    rows=sq.order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).limit(1000).all()
    if q:
        ql=q.lower()
        rows=[r for r in rows if ql in " ".join([
            _normalize_text(vehicle_plate(getattr(r,'vehicle',None))),
            _normalize_text(vehicle_display_name(getattr(r,'vehicle',None))),
            _normalize_text(getattr(getattr(r,'vehicle',None),'merk',None)),
            _normalize_text(getattr(getattr(r,'vehicle',None),'type',None)),
            _normalize_text(getattr(r,'vendor',None)),
            _normalize_text(getattr(r,'note',None)),
        ]).lower()]
    if status:
        tmp=[]
        for r in rows:
            due=service_due_date(r)
            current_status=_service_status(due)
            if status=="overdue" and current_status=="overdue":tmp.append(r)
            elif status=="due" and current_status=="soon":tmp.append(r)
            elif status=="ok" and current_status=="safe":tmp.append(r)
        rows=tmp
    return render_template(
        "history_servis.html",
        service_rows=rows,
        q=q,
        status=status,
        company=company,
        companies=companies,
        title="Riwayat Servis - Sinar Group",
        service_due_date=service_due_date,
        service_items_for_record=_service_items_for_record,
        clean_service_note=_clean_service_note,
        service_record_type_label=_service_record_type_label,
        service_payment_label=_service_payment_label,
        service_total_cost=_service_total_cost,
        service_odometer_value=_service_odometer_value,
        service_processed_at=_service_processed_at,
        fmt_tgl_jam=format_tgl_jam_id_full,
    )

@bp.get("/history/kir")
@master_required
def history_kir():
    q=(request.args.get("q") or "").strip()
    status=(request.args.get("status") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()
    vehicle_ids=_filtered_vehicle_ids(q,company)
    kq=KirRecord.query
    if vehicle_ids is not None:kq=kq.filter(KirRecord.vehicle_id.in_(vehicle_ids))
    rows=kq.order_by(KirRecord.due_date.desc(),KirRecord.id.desc()).limit(1200).all()
    if status:
        filtered=[]
        for r in rows:
            current_status=_kir_status(getattr(r,"due_date",None))
            if status==current_status:filtered.append(r)
        rows=filtered
    return render_template("history_kir.html",kir_rows=rows[:900],q=q,status=status,company=company,companies=companies,title="Riwayat KIR - Sinar Group",today=date.today(),fmt_tgl_jam=format_tgl_jam_id_full,REMINDER_DAYS=DEFAULT_REMINDER_DAYS)

@bp.get("/history/user")
@master_required
def history_user():
    q=(request.args.get("q") or "").strip()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()
    query=_apply_vehicle_scope(VehicleChangeHistory.query.join(Vehicle,Vehicle.id==VehicleChangeHistory.vehicle_id).filter(Vehicle.is_deleted==False))
    if q:
        like=f"%{q}%"
        query=query.filter(or_(Vehicle.pt.ilike(like),Vehicle.active_name.ilike(like),Vehicle.new_asset_name.ilike(like),Vehicle.name_as_asset_pt.ilike(like),Vehicle.plate_old.ilike(like),Vehicle.plate_new.ilike(like),VehicleChangeHistory.field_label.ilike(like),VehicleChangeHistory.old_value.ilike(like),VehicleChangeHistory.new_value.ilike(like),VehicleChangeHistory.note.ilike(like)))
    if company:
        query=query.filter(func.lower(func.trim(Vehicle.pt))==company.strip().lower())
    rows=query.order_by(VehicleChangeHistory.changed_at.desc(),VehicleChangeHistory.id.desc()).limit(1000).all()
    
    # Kelompokkan baris history berdasarkan (vehicle_id, changed_at, changed_by)
    grouped_sessions = {}
    for r in rows:
        # Gunakan resolusi detik untuk pengelompokan (karena batch update biasanya di waktu yang sama)
        ts = r.changed_at.strftime('%Y-%m-%d %H:%M:%S') if r.changed_at else 'unknown'
        key = (r.vehicle_id, ts, r.changed_by)
        if key not in grouped_sessions:
            grouped_sessions[key] = {
                'vehicle_id': r.vehicle_id,
                'changed_at': r.changed_at,
                'changed_by': r.changed_by or 'system',
                'changed_fields': {}
            }
        
        field_name = r.field_name or 'unknown'
        grouped_sessions[key]['changed_fields'][field_name] = {
            'old': r.old_value or '',
            'new': r.new_value or ''
        }

    sessions=[]
    # Urutkan kembali berdasarkan waktu desc
    sorted_keys = sorted(grouped_sessions.keys(), key=lambda x: (x[1], x[0]), reverse=True)
    
    for key in sorted_keys:
        s = grouped_sessions[key]
        v = Vehicle.query.get(s['vehicle_id'])
        if not v:
            continue
            
        # Rekonstruksi snapshot data kendaraan pada saat riwayat ini terjadi.
        # Kita mulai dari data live (kondisi terbaru), lalu mundurkan nilainya 
        # dengan mengaplikasikan 'old_value' dari semua perubahan yang terjadi 
        # SETELAH sesi ini sampai ke waktu sesi ini.
        snapshot = _vehicle_snapshot(v, include_custom=True)
        
        # Cari semua perubahan untuk kendaraan ini yang terjadi SETELAH sesi ini
        later_changes = VehicleChangeHistory.query.filter(
            VehicleChangeHistory.vehicle_id == v.id,
            VehicleChangeHistory.changed_at > s['changed_at']
        ).order_by(VehicleChangeHistory.changed_at.asc()).all()
        
        # Mundurkan nilai snapshot ke waktu sesi ini dengan mengaplikasikan 'old_value'
        # dari perubahan-perubahan yang terjadi setelahnya (dari yang paling baru ke yang lama)
        # Tapi query di atas asc, jadi kita balik atau proses dari belakang.
        reversed_later = sorted(later_changes, key=lambda x: x.changed_at, reverse=True)
        for lc in reversed_later:
            if lc.field_name in snapshot:
                snapshot[lc.field_name] = lc.old_value

        s['snapshot'] = snapshot
        sessions.append(s)
    custom_columns=_vehicle_custom_columns()
    return render_template("history_user.html",sessions=sessions,q=q,company=company,companies=companies,custom_columns=custom_columns,title="Riwayat Kendaraan - Sinar Group")

@bp.get("/history/peminjaman")
@master_required
def history_peminjaman():
    q=(request.args.get("q") or "").strip()
    status=(request.args.get("status") or "").strip().lower()
    company=_sanitize_company_filter((request.args.get("company") or "").strip())
    companies=_scoped_company_choices()
    vehicle_ids=_filtered_vehicle_ids(q,company)
    query=LoanTransaction.query.join(Vehicle,Vehicle.id==LoanTransaction.vehicle_id).filter(
        Vehicle.is_deleted==False,
        or_(func.lower(LoanTransaction.status)=="selesai", LoanTransaction.date_return_actual.isnot(None))
    )
    query=_apply_vehicle_scope(query)
    if company:
        query=query.filter(func.lower(func.trim(Vehicle.pt))==company.strip().lower())
    if vehicle_ids is not None:
        query=query.filter(LoanTransaction.vehicle_id.in_(vehicle_ids))
    if status:
        query=query.filter(func.lower(LoanTransaction.status)==status)

    items=query.order_by(
        LoanTransaction.date_return_actual.desc(),
        LoanTransaction.updated_at.desc(),
        LoanTransaction.id.desc()
    ).limit(1000).all()
    loan_place_map=_loan_place_map([it.id for it in items]) if items else {}
    return render_template("history_peminjaman.html",items=items,q=q,status=status,company=company,companies=companies,loan_place_map=loan_place_map,today=date.today(),title="Riwayat Peminjaman - Sinar Group")

@bp.get("/login")
def login():
    if is_logged_in():
        return redirect(url_for("main.home"))
    return render_template("login.html", title="Login")

@bp.post("/login")
def login_post():
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""

    if not username or not password:
        flash("Username dan password wajib diisi.", "danger")
        return redirect(url_for("main.login"))

    user = User.query.filter(func.lower(User.username) == username.lower()).first()

    if not user or not user.check_password(password):
        flash("Username atau password salah.", "danger")
        return redirect(url_for("main.login"))

    if not user.is_active:
        flash("Akun tidak aktif.", "danger")
        return redirect(url_for("main.login"))

    session.clear()
    session["user_id"] = user.id
    session["user_name"] = user.full_name   # tampil di UI
    session["username"] = user.username     # identitas login
    session["user_role"] = user.role
    session["is_master"] = user.role == "master"

    try:
        audit("LOGIN", "User", user.id, note=f"Login username={user.username}")
    except:
        pass

    return redirect(url_for("main.home"))

@bp.get("/logout")
def logout():
    actor_name = session.get("user_name") or "SYSTEM"
    actor_username = session.get("username") or "-"
    actor_role = session.get("user_role") or "SYSTEM"

    try:
        audit("LOGOUT", "Session", 0, note=f"Logout user={actor_name} username={actor_username} role={actor_role}")
    except Exception:
        pass

    session.clear()
    flash("Berhasil logout.", "success")
    return redirect(url_for("main.login"))

@bp.post("/register")
def register_post():
    flash("Pembuatan akun hanya bisa dilakukan dari menu settings oleh master admin.","warning")
    return redirect(url_for("main.login"))

@bp.get("/audit")
@master_required
def audit_page():
    if not _is_master_session():
        flash("Hanya master admin yang bisa membuka audit.", "danger")
        return redirect(url_for("main.home"))
    raw_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(500).all()

    logs = []
    for l in raw_logs:
        actor_name = "-"
        actor_username = "-"
        actor_role = "SYSTEM"

        # Coba dari relasi actor dulu (kalau ada)
        if getattr(l, "actor", None):
            actor_name = (
                getattr(l.actor, "full_name", None)
                or getattr(l.actor, "name", None)
                or getattr(l.actor, "username", None)
                or "-"
            )
            actor_username = getattr(l.actor, "username", None) or "-"
            actor_role = (getattr(l.actor, "role", None) or "user").upper()
        else:
            # Parse dari field note — format: "... user=NAME username=USER role=ROLE"
            note = l.note or ""
            note_lower = note.lower()

            import re as _re
            # Logout: "Logout user=Budi username=budi role=admin"
            m_user = _re.search(r'user=([^\s]+)', note, _re.I)
            m_uname = _re.search(r'username=([^\s]+)', note, _re.I)
            m_role = _re.search(r'role=([^\s]+)', note, _re.I)

            if m_user:
                actor_name = m_user.group(1)
            if m_uname:
                actor_username = m_uname.group(1)
            if m_role:
                actor_role = m_role.group(1).upper()

            # Login: "Login username=master" — ambil username sebagai nama juga
            if "login username=" in note_lower and actor_name == "-":
                m_ln = _re.search(r'login username=([^\s]+)', note, _re.I)
                if m_ln:
                    actor_name = m_ln.group(1)
                    actor_username = m_ln.group(1)
                    actor_role = "USER"

            # Tambah via tabel inline: "... oleh NAMA"
            if actor_name == "-":
                m_oleh = _re.search(r'oleh\s+(.+?)(?:\s*$)', note, _re.I)
                if m_oleh:
                    actor_name = m_oleh.group(1).strip()

            # Fallback: action=LOGIN entity=User → actor dari note
            if actor_name == "-" and getattr(l, "action", "") == "LOGIN":
                actor_role = "USER"

        logs.append({
            "id": l.id,
            "created_at": l.created_at,
            "action": l.action,
            "entity_type": l.entity_type,
            "entity_id": l.entity_id,
            "vehicle_id": getattr(l, "vehicle_id", None),
            "note": l.note,
            "ip": l.ip,
            "actor_name": actor_name,
            "actor_username": actor_username,
            "actor_role": actor_role,
        })

    return render_template("audit.html", logs=logs, title="Audit Aktivitas")

def _settings_user_payload(user):
    return {
        "id": user.id,
        "name": (getattr(user, "full_name", None) or getattr(user, "username", None) or "-").strip(),
        "username": getattr(user, "username", None) or "-",
        "role": getattr(user, "role", None) or "admin",
        "active": bool(getattr(user, "is_active", False)),
        "pts": sorted([
            (getattr(item, "pt_name", None) or "").strip()
            for item in (getattr(user, "pt_accesses", []) or [])
            if (getattr(item, "pt_name", None) or "").strip()
        ])
    }

@bp.get("/settings")
@master_required
def settings_page():
    if not _is_master_session():
        flash("Hanya master admin yang bisa membuka settings.", "danger")
        return redirect(url_for("main.home"))
    users = User.query.order_by(User.role.desc(), User.full_name.asc(), User.username.asc()).all()
    admin_users = [_settings_user_payload(u) for u in users]
    pt_choices = _vehicle_pt_choices()
    return render_template(
        "settings.html",
        title="Pengaturan",
        admin_users=admin_users,
        pt_choices=pt_choices,
    )

@bp.post("/settings/users/create")
@master_required
def settings_user_create():
    if not _is_master_session():
        flash("Hanya master admin yang bisa menambah user.", "danger")
        return redirect(url_for("main.home"))

    full_name = (request.form.get("full_name") or "").strip()
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    role = (request.form.get("role") or "admin").strip().lower()
    pt_names = [x.strip() for x in request.form.getlist("pt_access") if x.strip()]

    if not full_name or not username or not password:
        flash("Nama lengkap, username, dan password wajib diisi.", "danger")
        return redirect(url_for("main.settings_page"))

    if role not in {"admin", "master"}:
        role = "admin"

    if role == "admin" and not pt_names:
        flash("Admin PT wajib memilih minimal 1 PT.", "danger")
        return redirect(url_for("main.settings_page"))

    existing_user = User.query.filter(func.lower(User.username) == username.lower()).first()
    if existing_user:
        flash("Username sudah terdaftar.", "danger")
        return redirect(url_for("main.settings_page"))

    user = User(
        username=username,
        full_name=full_name,
        role=role,
        is_active=True
    )
    user.set_password(password)
    db.session.add(user)
    db.session.flush()

    if role != "master":
        seen = set()
        for pt in pt_names:
            key = pt.lower()
            if key not in seen:
                seen.add(key)
                db.session.add(UserPtAccess(user_id=user.id, pt_name=pt))

    db.session.commit()

    try:
        audit("CREATE", "User", user.id, note=f"Buat user {full_name} (username={username}) role={role}")
    except Exception:
        pass

    flash(f"User {full_name} berhasil dibuat.", "success")
    return redirect(url_for("main.settings_page"))

@bp.post("/settings/users/<int:user_id>/reset-password")
@master_required
def settings_user_reset_password(user_id:int):
    if not _is_master_session():
        flash("Hanya master admin yang bisa reset password user.", "danger")
        return redirect(url_for("main.home"))

    user = User.query.get_or_404(user_id)
    new_password = (request.form.get("new_password") or "").strip()
    confirm_password = (request.form.get("confirm_password") or "").strip()

    if not new_password:
        flash(f"Password baru untuk {user.display_name()} wajib diisi.", "danger")
        return redirect(url_for("main.settings_page"))
    if new_password != confirm_password:
        flash(f"Konfirmasi password untuk {user.display_name()} tidak sama.", "danger")
        return redirect(url_for("main.settings_page"))

    user.set_password(new_password)
    db.session.commit()
    actor_name = session.get("user_name") or "SYSTEM"
    try:
        audit("UPDATE","User",user.id,note=f"Reset password user {user.display_name()} ({user.username}) oleh {actor_name}")
    except Exception:
        pass
    flash(f"Password user {user.display_name()} berhasil direset.", "success")
    return redirect(url_for("main.settings_page"))

@bp.post("/settings/change-password/<int:user_id>", endpoint="settings_change_password")
@master_required
def settings_change_password(user_id:int):
    return settings_user_reset_password(user_id)

@bp.post("/settings/users/<int:user_id>/toggle")
@master_required
def settings_toggle_user(user_id:int):
    if not _is_master_session():
        flash("Hanya master admin yang bisa mengubah status user.", "danger")
        return redirect(url_for("main.home"))
    user = User.query.get_or_404(user_id)
    user.is_active = not bool(user.is_active)
    db.session.commit()
    try:
        audit("UPDATE","User",user.id,note=f"Set aktif={user.is_active} untuk {user.display_name()}")
    except Exception:
        pass
    flash(f"Status user {user.display_name()} berhasil diperbarui.", "success")
    return redirect(url_for("main.settings_page"))


@bp.post("/settings/users/<int:user_id>/delete")
@master_required
def settings_user_delete(user_id:int):
    if not _is_master_session():
        flash("Hanya master admin yang bisa menghapus user.", "danger")
        return redirect(url_for("main.home"))
    user = User.query.get_or_404(user_id)
    me = get_me()
    if me and me.id == user_id:
        flash("Tidak bisa menghapus akun yang sedang digunakan.", "danger")
        return redirect(url_for("main.settings_page"))
    name = user.display_name()
    username = user.username
    try:
        UserPtAccess.query.filter_by(user_id=user_id).delete()
        db.session.delete(user)
        db.session.commit()
        actor_name = session.get("user_name") or "SYSTEM"
        try:
            audit("DELETE","User",user_id,note=f"Hapus user {name} ({username}) oleh {actor_name}")
        except Exception:
            pass
        flash(f"User {name} berhasil dihapus.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Gagal menghapus user: {e}", "danger")
    return redirect(url_for("main.settings_page"))


@bp.post("/api/kendaraan/<int:vehicle_id>/terjual")
@master_required
def api_kendaraan_mark_terjual(vehicle_id:int):
    v = Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)

    # Support multipart/form-data (ada foto) ATAU JSON (tanpa foto)
    if request.content_type and 'multipart' in request.content_type:
        data = request.form
        get_val = lambda k: _normalize_text(data.get(k))
        sold_date_str = get_val("date")
        sold_to       = get_val("sold_to")
        note          = get_val("note")
        input_terbilang = get_val("terbilang")
        nik           = get_val("nik")
        npwp          = get_val("npwp")
        foto_files = []
        for key in ("foto", "fotos", "foto[]"):
            foto_files.extend([f for f in request.files.getlist(key) if f and f.filename])
        foto_mode = get_val("foto_mode").lower()
    else:
        data = request.get_json() or {}
        get_val = lambda k: _normalize_text(data.get(k))
        sold_date_str = get_val("date")
        sold_to       = get_val("sold_to")
        note          = get_val("note")
        input_terbilang = get_val("terbilang")
        nik           = get_val("nik")
        npwp          = get_val("npwp")
        foto_files = []
        foto_mode = get_val("foto_mode").lower()

    sold_date   = _parse_date(sold_date_str)
    price_value = _parse_float(data.get("price"))

    if not sold_date or price_value in (None, 0) or not sold_to:
        return jsonify({"success": False, "message": "Tanggal, harga jual, dan pembeli wajib diisi."}), 400

    price_int = int(round(price_value))
    terbilang = input_terbilang or f"{price_int:,}".replace(",", ".")

    existing_note = _normalize_text(v.tambahan_keterangan)
    existing_sold_info = _parse_sold_info(existing_note)

    # Buang blok TERJUAL lama saja, info umum/catatan lain tetap aman.
    if existing_note:
        existing_note = re.sub(r"\s*\|?\s*TERJUAL pada .*?$", "", existing_note, flags=re.IGNORECASE | re.DOTALL).strip(" |")

    # Simpan banyak foto jika ada.
    # Default: upload baru akan DITAMBAHKAN ke foto lama.
    # Untuk mengganti semua foto lama, kirim foto_mode=replace / ganti.
    foto_filenames = []
    existing_foto_urls = list(existing_sold_info.get("foto_urls") or [])
    if data.get("hapus_foto") in (True, "1", "true", "yes", "on"):
        existing_foto_urls = []
    if foto_mode in {"replace", "ganti", "reset"}:
        existing_foto_urls = []

    for foto_file in foto_files:
        foto_filename = _save_foto(foto_file, vehicle_id)
        if foto_filename:
            foto_filenames.append(foto_filename)
            existing_foto_urls.append(url_for("main.kendaraan_foto", filename=foto_filename, _external=False))

    foto_urls = []
    seen_foto = set()
    for foto_url_item in existing_foto_urls:
        normalized_foto = _normalize_foto_url(foto_url_item)
        if normalized_foto and normalized_foto not in seen_foto:
            seen_foto.add(normalized_foto)
            foto_urls.append(normalized_foto)
    foto_url = foto_urls[0] if foto_urls else None
    foto_filename = foto_filenames[0] if foto_filenames else None

    # Bangun info_terjual setelah foto final tersedia.
    info_terjual = f"TERJUAL pada {sold_date.strftime('%Y-%m-%d')} kepada {sold_to} seharga Rp {price_int:,} ({terbilang})".replace(",", ".")
    if nik:
        info_terjual += f". NIK/KTP: {nik}"
    if npwp:
        info_terjual += f". NPWP: {npwp}"
    if note:
        info_terjual += f". Catatan: {note}"
    foto_joined = _join_foto_urls(foto_urls)
    if foto_joined:
        info_terjual += f". Foto: {foto_joined}"

    v.tambahan_keterangan = f"{existing_note} | {info_terjual}".strip(" |") if existing_note else info_terjual
    v.status = "TERJUAL"
    v.updated_at = _now_naive()

    try:
        db.session.commit()
        try:
            audit("UPDATE", "Vehicle", v.id, note=f"Unit {vehicle_plate(v)} ditandai TERJUAL kepada {sold_to}" + (f" | NIK: {nik}" if nik else "") + (f" | NPWP: {npwp}" if npwp else ""))
        except Exception:
            pass
        return jsonify({
            "success": True,
            "message": f"Unit {vehicle_plate(v)} berhasil dipindahkan ke laporan terjual.",
            "sold_date": sold_date.strftime('%Y-%m-%d'),
            "price": price_int,
            "sold_to": sold_to,
            "foto_url": foto_url,
            "foto_urls": foto_urls,
            "foto_filename": foto_filename,
            "foto_filenames": foto_filenames,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

# ===== FOTO PEMBELI TERJUAL =====
ALLOWED_IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.pdf'}
MAX_FOTO_SIZE = 10 * 1024 * 1024  # 10 MB

def _foto_upload_dir() -> Path:
    """Folder penyimpanan foto — dibuat otomatis jika belum ada."""
    try:
        base = Path(current_app.root_path)
    except RuntimeError:
        base = Path(__file__).parent
    folder = base / "static" / "uploads" / "terjual_foto"
    folder.mkdir(parents=True, exist_ok=True)
    return folder

def _save_foto(file_storage, vehicle_id: int) -> str | None:
    """Simpan file foto, return nama file yang tersimpan atau None."""
    if not file_storage or not file_storage.filename:
        return None
    ext = Path(file_storage.filename).suffix.lower()
    if ext not in ALLOWED_IMAGE_EXTS:
        return None
    # Baca dulu untuk cek ukuran
    data = file_storage.read()
    if len(data) > MAX_FOTO_SIZE:
        return None
    filename = f"terjual_{vehicle_id}_{secrets.token_hex(8)}{ext}"
    save_path = _foto_upload_dir() / filename
    save_path.write_bytes(data)
    return filename

@bp.get("/kendaraan/foto/<path:filename>")
@master_required
def kendaraan_foto(filename: str):
    """Serve foto pembeli/identitas unit terjual dari folder upload mana pun yang valid."""
    clean_filename = os.path.basename(str(filename).replace('\\', '/'))
    path = _resolve_foto_path(clean_filename)
    if path:
        return send_from_directory(str(path.parent), path.name)
    return send_from_directory(str(_foto_upload_dir()), clean_filename)

# ---------------------------------------------------------------------------
# PDF Token – akses PDF dari hyperlink Excel tanpa perlu login
# Token di-sign dengan SECRET_KEY app, berlaku 24 jam.
# ---------------------------------------------------------------------------

def _pdf_serializer():
    secret = current_app.config.get("SECRET_KEY") or "fallback-secret-key-change-me"
    return URLSafeTimedSerializer(secret, salt="pdf-download")

def _generate_pdf_token(filename: str) -> str:
    """Buat signed token untuk satu filename PDF."""
    return _pdf_serializer().dumps(filename)

def _verify_pdf_token(token: str, max_age: int = 86400) -> str | None:
    """Verifikasi token, return filename atau None kalau invalid/expired."""
    try:
        return _pdf_serializer().loads(token, max_age=max_age)
    except (BadSignature, SignatureExpired):
        return None


@bp.get("/kendaraan/foto-pdf/<token>")
def kendaraan_foto_pdf_public(token: str):
    """Serve PDF tanpa login menggunakan signed token.

    Route ini sengaja tidak pakai @master_required supaya hyperlink di Excel
    bisa langsung dibuka tanpa harus login dulu. Keamanannya dijaga oleh token
    yang di-sign dan punya masa berlaku (default 24 jam).
    """
    filename = _verify_pdf_token(token)
    if not filename:
        from flask import abort
        abort(403)
    clean_filename = os.path.basename(str(filename).replace('\\', '/'))
    if not clean_filename.lower().endswith('.pdf'):
        from flask import abort
        abort(403)
    path = _resolve_foto_path(clean_filename)
    if path and path.exists():
        return send_from_directory(str(path.parent), path.name, mimetype='application/pdf')
    upload_dir = _foto_upload_dir()
    pdf_path = upload_dir / clean_filename
    if pdf_path.exists():
        return send_from_directory(str(upload_dir), clean_filename, mimetype='application/pdf')
    from flask import abort
    abort(404)


@bp.get("/api/kendaraan/terjual")
@master_required
def api_kendaraan_terjual():
    pt = (request.args.get("pt") or "").strip()
    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))

    query = _apply_vehicle_scope(
        Vehicle.query.filter(
            Vehicle.is_deleted == False,
            func.lower(func.trim(Vehicle.status)).like("%terjual%")
        )
    )
    if pt:
        query = query.filter(func.lower(func.trim(Vehicle.pt)) == pt.strip().lower())

    vehicles = query.order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).all()

    items = []
    for v in vehicles:
        tgl = None
        ket = v.tambahan_keterangan or ""
        # coba ambil tanggal terjual dari keterangan atau updated_at
        sold_info = _parse_sold_info(ket)
        tgl = sold_info["sold_date"]

        if date_from or date_to:
            ref = tgl if tgl else (v.updated_at.date() if v.updated_at else None)
            if ref:
                if date_from and ref < date_from:
                    continue
                if date_to and ref > date_to:
                    continue
        
        items.append({
            "id": v.id,
            "no": len(items) + 1,
            "pt": v.pt or "-",
            "active_name": v.active_name or v.name_as_asset_pt or "-",
            "merk": v.merk or "",
            "type": v.type or "",
            "plate": v.plate_new or v.plate_old or "-",
            "tgl_terjual": tgl.strftime("%Y-%m-%d") if tgl else "-",
            "harga_jual": rupiah(sold_info["price"]),
            "price_raw": int(float(sold_info["price"] or 0)),
            "terbilang": sold_info["terbilang"],
            "pembeli": sold_info["sold_to"],
            "nik": sold_info["nik"],
            "npwp": sold_info["npwp"],
            "foto_url": sold_info["foto_url"],
            "foto_urls": sold_info.get("foto_urls") or [],
            "foto_count": len(sold_info.get("foto_urls") or []),
            "foto_is_pdf": bool(sold_info.get("foto_url") and Path(str(sold_info.get("foto_url"))).suffix.lower() == ".pdf"),
            "foto_filename": os.path.basename(str(sold_info.get("foto_url") or "").split("?", 1)[0].split("#", 1)[0]),
            "keterangan": sold_info["note"],
        })
    return jsonify({"success": True, "items": items})


def _pdf_page_to_png_bytes(pdf_path: Path) -> bytes | None:
    """Render halaman pertama PDF ke PNG bytes menggunakan PyMuPDF.

    Hasilnya bisa langsung di-embed ke Excel sebagai gambar biasa,
    sehingga file Excel self-contained dan bisa dibuka di mana saja
    tanpa butuh server/internet.
    """
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(str(pdf_path))
        if doc.page_count == 0:
            return None
        page = doc[0]
        # Render dengan resolusi cukup supaya teks PDF masih terbaca di sel kecil
        mat = fitz.Matrix(1.5, 1.5)
        pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
        return pix.tobytes("png")
    except Exception:
        return None


def _put_laporan_terjual_foto_to_excel(ws, row_idx: int, col_idx: int, foto_url: str | None) -> None:
    """Tempel foto/preview-PDF ke export Excel Laporan Terjual.

    JPG/JPEG/PNG/WEBP  -> embed gambar langsung ke sel (self-contained).
    PDF                -> render halaman pertama jadi PNG lalu embed,
                          sehingga Excel bisa dibuka di mana saja tanpa
                          butuh koneksi ke server.
    """
    cell = ws.cell(row=row_idx, column=col_idx)
    normalized_url = _normalize_foto_url(foto_url)
    if not normalized_url:
        cell.value = '-'
        return

    foto_path = _resolve_foto_path(normalized_url)
    ext = (foto_path.suffix.lower() if foto_path else Path(str(normalized_url)).suffix.lower())

    if ext == '.pdf':
        # Konversi halaman pertama PDF ke PNG lalu embed supaya Excel self-contained.
        if foto_path and foto_path.exists():
            png_bytes = _pdf_page_to_png_bytes(foto_path)
            if png_bytes:
                from io import BytesIO as _BytesIO
                bio = _BytesIO(png_bytes)
                bio.seek(0)
                if not hasattr(ws, '_foto_buffers'):
                    ws._foto_buffers = []
                ws._foto_buffers.append(bio)
                img = ExcelImage(bio)
                img.width = 95
                img.height = 72
                ws.row_dimensions[row_idx].height = 60
                ws.add_image(img, cell.coordinate)
                cell.value = None
                return
        # Fallback kalau file PDF tidak ditemukan atau gagal render
        cell.value = 'PDF tidak ditemukan'
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, 30)
        return

    if not foto_path or not foto_path.exists():
        cell.value = 'Foto tidak ditemukan'
        return

    if ext not in {'.jpg', '.jpeg', '.png', '.webp'}:
        cell.value = 'Format foto tidak didukung'
        return

    try:
        try:
            from PIL import Image as PILImage
        except ModuleNotFoundError:
            import sys, subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'Pillow'])
            from PIL import Image as PILImage

        from io import BytesIO as _BytesIO

        with PILImage.open(str(foto_path)) as pil_img:
            pil_img = pil_img.convert('RGB')
            pil_img.thumbnail((120, 90))
            bio = _BytesIO()
            pil_img.save(bio, format='PNG')
            bio.seek(0)

        if not hasattr(ws, '_foto_buffers'):
            ws._foto_buffers = []
        ws._foto_buffers.append(bio)

        img = ExcelImage(bio)
        img.width = 95
        img.height = 72
        ws.row_dimensions[row_idx].height = 60
        ws.add_image(img, cell.coordinate)
        cell.value = None
    except ModuleNotFoundError as exc:
        cell.value = 'Install Pillow dulu: pip install Pillow'
    except Exception as exc:
        cell.value = f'Foto gagal diproses: {type(exc).__name__}'

@bp.get("/api/kendaraan/terjual/export")
@master_required
def api_kendaraan_terjual_export():
    pt = (request.args.get("pt") or "").strip()
    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))

    query = _apply_vehicle_scope(
        Vehicle.query.filter(
            Vehicle.is_deleted == False,
            func.lower(func.trim(Vehicle.status)).like("%terjual%")
        )
    )
    if pt:
        query = query.filter(func.lower(func.trim(Vehicle.pt)) == pt.strip().lower())

    vehicles = query.order_by(Vehicle.updated_at.desc(), Vehicle.id.desc()).all()

    export_rows = []
    max_foto_count = 1
    for v in vehicles:
        ket = v.tambahan_keterangan or ""
        sold_info = _parse_sold_info(ket)
        tgl_terjual_str = sold_info["sold_date"].strftime("%Y-%m-%d") if sold_info["sold_date"] else "-"

        if date_from or date_to:
            ref = sold_info["sold_date"] if sold_info["sold_date"] else (v.updated_at.date() if v.updated_at else None)
            if ref:
                if date_from and ref < date_from:
                    continue
                if date_to and ref > date_to:
                    continue

        foto_urls = sold_info.get("foto_urls") or ([] if not sold_info.get("foto_url") else [sold_info.get("foto_url")])
        max_foto_count = max(max_foto_count, len(foto_urls) or 1)
        export_rows.append((v, sold_info, tgl_terjual_str, foto_urls))

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Terjual"

    header_fill = PatternFill("solid", fgColor="1e4080")
    header_font = Font(bold=True, color="FFD15F", size=10)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    foto_headers = [f"Foto {i}" for i in range(1, max_foto_count + 1)]
    headers = [
        "No", "PT", "Nama Aktiva", "Merk", "Type", "No Polisi",
        "Tgl Terjual", "Harga Jual", "Terbilang", "Pembeli",
        "NIK/KTP", "NPWP", *foto_headers, "Keterangan"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    foto_start_col = 13
    keterangan_col = foto_start_col + max_foto_count

    export_no = 1
    for v, sold_info, tgl_terjual_str, foto_urls in export_rows:
        current_row = ws.max_row + 1
        ws.append([
            export_no,
            v.pt or "-",
            v.active_name or v.name_as_asset_pt or "-",
            v.merk or "-",
            v.type or "-",
            v.plate_new or v.plate_old or "-",
            tgl_terjual_str,
            sold_info["price"],
            sold_info["terbilang"],
            sold_info["sold_to"],
            sold_info["nik"],
            sold_info["npwp"],
            *([""] * max_foto_count),
            sold_info["note"],
        ])

        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = border

        for foto_idx in range(max_foto_count):
            foto_col = foto_start_col + foto_idx
            _put_laporan_terjual_foto_to_excel(
                ws,
                current_row,
                foto_col,
                foto_urls[foto_idx] if foto_idx < len(foto_urls) else None
            )
        export_no += 1

    base_widths = {
        'A': 6, 'B': 16, 'C': 24, 'D': 14, 'E': 16, 'F': 15,
        'G': 14, 'H': 16, 'I': 30, 'J': 20, 'K': 18, 'L': 18,
    }
    for col_letter, width in base_widths.items():
        ws.column_dimensions[col_letter].width = width
    for col_idx in range(foto_start_col, foto_start_col + max_foto_count):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions[get_column_letter(keterangan_col)].width = 28

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for row_idx in range(2, ws.max_row + 1):
        if ws.row_dimensions[row_idx].height is None:
            ws.row_dimensions[row_idx].height = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"laporan_terjual_{date.today().strftime('%d-%m-%Y')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



def _safe_user_history_rows(vehicle_id:int, limit:int=100):
    """Ambil riwayat user tanpa bergantung pada nama kolom tanggal tertentu.

    Di beberapa database/model kolomnya bernama changed_at, sedangkan versi lama
    route pernah memanggil UserHistory.change_date sehingga memicu AttributeError.
    Helper ini membaca kolom yang benar dari model lalu tetap memberi key
    change_date untuk template.
    """
    rows = []
    try:
        date_attr = (
            getattr(UserHistory, "change_date", None)
            or getattr(UserHistory, "changed_at", None)
            or getattr(UserHistory, "created_at", None)
            or getattr(UserHistory, "updated_at", None)
            or getattr(UserHistory, "id")
        )
        query = UserHistory.query.filter_by(vehicle_id=vehicle_id).order_by(date_attr.desc(), UserHistory.id.desc()).limit(limit)
        for r in query.all():
            d = _history_timestamp_value(r, "change_date", "changed_at", "created_at", "updated_at")
            rows.append({
                "id": getattr(r, "id", None),
                "change_date": d or getattr(r, "change_date", None) or getattr(r, "changed_at", None),
                "changed_at": d or getattr(r, "changed_at", None) or getattr(r, "change_date", None),
                "user_lama": getattr(r, "user_lama", None) or getattr(r, "old_user", None) or "",
                "user_baru": getattr(r, "user_baru", None) or getattr(r, "new_user", None) or "",
                "pt_pemakai_lama": getattr(r, "pt_pemakai_lama", None) or "",
                "pt_pemakai_baru": getattr(r, "pt_pemakai_baru", None) or "",
                "changed_by": getattr(r, "changed_by", None) or "system",
                "note": getattr(r, "note", None) or "",
            })
    except Exception as e:
        print(f"DEBUG: UserHistory safe loader ignored: {e}")
        rows = []
    return rows

@bp.get("/health")
def health():
    return {"ok":True,"app":"fleet-asset-vehicle"}

def _vehicle_detail_last_snapshot_rows(vehicle:Vehicle):
    """Baris untuk Riwayat Perubahan Kendaraan di halaman detail kendaraan.

    Permintaan tampilan detail: yang ditampilkan bukan semua log perubahan,
    tetapi kondisi TERAKHIR kendaraan untuk SEMUA kolom kendaraan.
    Waktu/oleh tetap diambil dari sesi edit terakhir jika ada.
    """
    latest_change = (
        VehicleChangeHistory.query
        .filter_by(vehicle_id=vehicle.id)
        .order_by(VehicleChangeHistory.changed_at.desc(), VehicleChangeHistory.id.desc())
        .first()
    )
    latest_at = getattr(latest_change, "changed_at", None) or getattr(vehicle, "updated_at", None) or getattr(vehicle, "created_at", None)
    latest_by = getattr(latest_change, "changed_by", None) or "system"

    snapshot = _vehicle_snapshot(vehicle, include_custom=True)
    labels = _label_map(include_custom=True)
    rows = []
    for field_name, field_label in labels.items():
        value = snapshot.get(field_name)
        rows.append({
            "vehicle_id": vehicle.id,
            "field_name": field_name,
            "field_label": field_label,
            "old_value": "-",
            "new_value": _export_scalar(value),
            "change_type": "latest_snapshot",
            "changed_by": latest_by,
            "changed_at": latest_at,
            "note": "Data terakhir kendaraan",
        })
    return rows



def _detail_export_date_value(value):
    dt_value = _coerce_datetime_value(value)
    if dt_value:
        return format_tgl_id_full(dt_value.date()) or "-"
    return format_tgl_id_full(value) if value else "-"


def _detail_export_in_range(value, date_from=None, date_to=None):
    if not value:
        return False if (date_from or date_to) else True
    dt_value = _coerce_datetime_value(value)
    d = dt_value.date() if dt_value else value
    if date_from and d < date_from:
        return False
    if date_to and d > date_to:
        return False
    return True


@bp.get("/kendaraan/<int:vehicle_id>/detail_page/export")
@bp.get("/kendaraan/<int:vehicle_id>/export")
@bp.get("/api/kendaraan/detail/export/<int:vehicle_id>")
@bp.get("/api/kendaraan/export/<int:vehicle_id>")
@master_required
def vehicle_detail_export_excel(vehicle_id:int):
    """Export Excel dari halaman detail kendaraan.

    Dipakai oleh detail kendaraan yang dibuka dari Data Pajak, Data KIR, Data Servis,
    maupun Data Peminjaman. Endpoint ini sengaja berdiri sendiri agar route detail
    tidak mengubah tampilan/logic lama, hanya menambahkan kemampuan export Excel.
    """
    v = Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)

    date_from = _parse_date(request.args.get("date_from"))
    date_to = _parse_date(request.args.get("date_to"))
    if date_from and date_to and date_from > date_to:
        return jsonify({"success": False, "message": "Tanggal awal tidak boleh lebih besar dari tanggal akhir."}), 400

    vehicle_name = vehicle_display_name(v)
    plate = vehicle_plate(v)
    custom_values = _vehicle_custom_values_map([v.id]).get(v.id, {})
    custom_columns = _vehicle_custom_columns()

    asset_owner = _company_name(getattr(v, "asset_owner_company", None))
    pt_pemakai = _company_name(getattr(v, "pt_pemakai_company", None))

    summary_rows = [
        ("Nama Aset", getattr(v, "active_name", None) or getattr(v, "new_asset_name", None) or getattr(v, "name_as_asset_pt", None) or vehicle_name),
        ("No Polisi", plate),
        ("PT", _export_scalar(getattr(v, "pt", None))),
        ("PT Pemilik Aset", _export_scalar(asset_owner)),
        ("PT Pemakai", _export_scalar(pt_pemakai)),
        ("Merk", _export_scalar(getattr(v, "merk", None))),
        ("Tipe", _export_scalar(getattr(v, "type", None))),
        ("Jenis", _export_scalar(getattr(v, "jenis", None))),
        ("Tahun Pemakaian", _export_scalar(getattr(v, "year_of_use", None))),
        ("User Lama", _export_scalar(getattr(v, "user_old", None))),
        ("User Baru", _export_scalar(getattr(v, "user_new", None))),
        ("Status", _export_scalar(getattr(v, "status", None))),
        ("Kondisi Terkini", _export_scalar(getattr(v, "kondisi_terkini", None))),
        ("Lokasi", _export_scalar(getattr(v, "lokasi", None))),
        ("Tambahan Keterangan", _export_scalar(getattr(v, "tambahan_keterangan", None))),
        ("Tanggal Awal Filter", format_tgl_id_full(date_from) if date_from else "Semua"),
        ("Tanggal Akhir Filter", format_tgl_id_full(date_to) if date_to else "Semua"),
        ("Tanggal Export", format_tgl_jam_id_full(_now_naive())),
    ]
    for col in custom_columns:
        summary_rows.append((col.get("column_label") or col.get("column_key"), _export_scalar(custom_values.get(col.get("column_key")))))
    df_summary = pd.DataFrame(summary_rows, columns=["Informasi", "Nilai"])

    annual_rows = []
    try:
        rows = db.session.execute(
            text("""
                SELECT tax_year, due_date, paid_date, amount, note
                FROM annual_tax_payments
                WHERE vehicle_id = :vid
                ORDER BY COALESCE(paid_date, due_date) DESC, id DESC
            """),
            {"vid": v.id},
        ).mappings().all()
        for idx, r in enumerate([x for x in rows if _detail_export_in_range(x.get("paid_date") or x.get("due_date"), date_from, date_to)], 1):
            annual_rows.append({
                "No": idx,
                "Tahun Pajak": r.get("tax_year") or "-",
                "Exp / Jatuh Tempo": _detail_export_date_value(r.get("due_date")),
                "Tanggal Bayar": _detail_export_date_value(r.get("paid_date")),
                "Nominal": rupiah(r.get("amount")) if r.get("amount") not in (None, "") else "-",
                "Status": "Lunas" if _is_paid_complete(r.get("paid_date"), r.get("amount")) else "Belum Bayar",
                "Catatan": _export_scalar(r.get("note")),
            })
    except Exception:
        annual_rows = []
    df_annual = pd.DataFrame(annual_rows or [{"No":"-","Tahun Pajak":"-","Exp / Jatuh Tempo":"-","Tanggal Bayar":"-","Nominal":"-","Status":"-","Catatan":"-"}])

    five_rows = []
    try:
        rows = db.session.execute(
            text("""
                SELECT due_date, paid_date, amount, plate_before, plate_after, note
                FROM five_year_tax_payments
                WHERE vehicle_id = :vid
                ORDER BY COALESCE(paid_date, due_date) DESC, id DESC
            """),
            {"vid": v.id},
        ).mappings().all()
        for idx, r in enumerate([x for x in rows if _detail_export_in_range(x.get("paid_date") or x.get("due_date"), date_from, date_to)], 1):
            five_rows.append({
                "No": idx,
                "Exp / Jatuh Tempo": _detail_export_date_value(r.get("due_date")),
                "Tanggal Bayar": _detail_export_date_value(r.get("paid_date")),
                "Nominal": rupiah(r.get("amount")) if r.get("amount") not in (None, "") else "-",
                "Status": "Lunas" if _is_paid_complete(r.get("paid_date"), r.get("amount")) else "Belum Bayar",
                "Plat Lama": _export_scalar(r.get("plate_before")),
                "Plat Baru": _export_scalar(r.get("plate_after")),
                "Catatan": _export_scalar(r.get("note")),
            })
    except Exception:
        five_rows = []
    df_five = pd.DataFrame(five_rows or [{"No":"-","Exp / Jatuh Tempo":"-","Tanggal Bayar":"-","Nominal":"-","Status":"-","Plat Lama":"-","Plat Baru":"-","Catatan":"-"}])

    kir_rows = []
    kirs = KirRecord.query.filter_by(vehicle_id=v.id).order_by(KirRecord.done_date.desc(), KirRecord.due_date.desc(), KirRecord.id.desc()).all()
    for idx, r in enumerate([x for x in kirs if _detail_export_in_range(getattr(x, "done_date", None) or getattr(x, "due_date", None), date_from, date_to)], 1):
        due_date = getattr(r, "due_date", None)
        kir_rows.append({
            "No": idx,
            "Tanggal KIR": _detail_export_date_value(getattr(r, "done_date", None)),
            "Jatuh Tempo KIR": _detail_export_date_value(due_date),
            "Status KIR": _export_scalar(_kir_status(due_date) if due_date else "empty"),
            "Hasil KIR": _export_scalar(getattr(r, "result", None)),
            "Status Proses": _export_scalar(getattr(r, "status", None)),
            "Catatan": _export_scalar(getattr(r, "note", None)),
        })
    df_kir = pd.DataFrame(kir_rows or [{"No":"-","Tanggal KIR":"-","Jatuh Tempo KIR":"-","Status KIR":"-","Hasil KIR":"-","Status Proses":"-","Catatan":"-"}])

    service_rows = []
    services = ServiceRecord.query.filter_by(vehicle_id=v.id).order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).all()
    for idx, r in enumerate([x for x in services if _detail_export_in_range(getattr(x, "service_date", None), date_from, date_to)], 1):
        service_rows.append({
            "No": idx,
            "Tanggal Servis": _detail_export_date_value(getattr(r, "service_date", None)),
            "Next Servis": _detail_export_date_value(service_due_date(r)),
            "Kategori": _export_scalar(_service_record_type_label(r)),
            "Vendor": _export_scalar(getattr(r, "vendor", None)),
            "Odometer": _export_scalar(_service_odometer_value(r)),
            "Pembayaran": _export_scalar(_service_payment_label(r)),
            "Total Biaya": rupiah(_service_total_cost(r)),
            "Rincian Item": _export_scalar(_service_items_summary(r)),
            "Catatan": _export_scalar(_clean_service_note(getattr(r, "note", None))),
        })
    df_service = pd.DataFrame(service_rows or [{"No":"-","Tanggal Servis":"-","Next Servis":"-","Kategori":"-","Vendor":"-","Odometer":"-","Pembayaran":"-","Total Biaya":"-","Rincian Item":"-","Catatan":"-"}])

    loan_rows = []
    loans = LoanTransaction.query.filter_by(vehicle_id=v.id).order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).all()
    loan_place_lookup = _loan_place_map([r.id for r in loans]) if loans else {}
    for idx, r in enumerate([x for x in loans if _detail_export_in_range(getattr(x, "date_out", None) or getattr(x, "date_return_plan", None) or getattr(x, "date_return_actual", None), date_from, date_to)], 1):
        loan_rows.append({
            "No": idx,
            "Dari PT": _export_scalar(getattr(r, "borrower_company", None)),
            "Ke PT / Peminjam": _export_scalar(getattr(r, "borrower_name", None) or getattr(r, "borrower", None)),
            "Tempat": _export_scalar(loan_place_lookup.get(r.id)),
            "Tanggal Pinjam": _detail_export_date_value(getattr(r, "date_out", None)),
            "Rencana Kembali": _detail_export_date_value(getattr(r, "date_return_plan", None)),
            "Aktual Kembali": _detail_export_date_value(getattr(r, "date_return_actual", None)),
            "Tujuan": _export_scalar(getattr(r, "purpose", None)),
            "Status": _export_scalar(_loan_status_label(r)),
            "Catatan": _export_scalar(getattr(r, "note", None)),
        })
    df_loan = pd.DataFrame(loan_rows or [{"No":"-","Dari PT":"-","Ke PT / Peminjam":"-","Tempat":"-","Tanggal Pinjam":"-","Rencana Kembali":"-","Aktual Kembali":"-","Tujuan":"-","Status":"-","Catatan":"-"}])

    vehicle_history_rows = []
    for idx, r in enumerate(_vehicle_detail_last_snapshot_rows(v), 1):
        vehicle_history_rows.append({
            "No": idx,
            "Tanggal": format_tgl_jam_id_full(r.get("changed_at")) if r.get("changed_at") else "-",
            "Kolom": _export_scalar(r.get("field_label")),
            "Nilai Terakhir": _export_scalar(r.get("new_value")),
            "Diubah Oleh": _export_scalar(r.get("changed_by")),
            "Catatan": _export_scalar(r.get("note")),
        })
    df_vehicle_history = pd.DataFrame(vehicle_history_rows or [{"No":"-","Tanggal":"-","Kolom":"-","Nilai Terakhir":"-","Diubah Oleh":"-","Catatan":"-"}])

    user_history_rows = []
    for idx, r in enumerate(_safe_user_history_rows(v.id, limit=1000), 1):
        d = _history_timestamp_value(r, "change_date", "changed_at", "created_at", "updated_at")
        if not _detail_export_in_range(d, date_from, date_to):
            continue
        user_history_rows.append({
            "No": idx,
            "Tanggal": format_tgl_jam_id_full(d) if d else "-",
            "User Lama": _export_scalar(getattr(r, "user_lama", None)),
            "User Baru": _export_scalar(getattr(r, "user_baru", None)),
            "Diubah Oleh": _export_scalar(getattr(r, "changed_by", None) or "system"),
            "Catatan": _export_scalar(getattr(r, "note", None)),
        })
    df_user_history = pd.DataFrame(user_history_rows or [{"No":"-","Tanggal":"-","User Lama":"-","User Baru":"-","Diubah Oleh":"-","Catatan":"-"}])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Ringkasan Kendaraan")
        df_annual.to_excel(writer, index=False, sheet_name="Pajak Tahunan")
        df_five.to_excel(writer, index=False, sheet_name="Pajak 5 Tahunan")
        df_kir.to_excel(writer, index=False, sheet_name="KIR")
        df_service.to_excel(writer, index=False, sheet_name="Servis")
        df_loan.to_excel(writer, index=False, sheet_name="Peminjaman")
        df_vehicle_history.to_excel(writer, index=False, sheet_name="Riwayat Kendaraan")
        df_user_history.to_excel(writer, index=False, sheet_name="Riwayat User")

        for ws in writer.book.worksheets:
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True, color="FFD15F")
                    cell.fill = cell.fill.copy(fill_type="solid", fgColor="1E4080")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            _autosize_worksheet_columns(ws, min_width=14, max_width=42)
            ws.freeze_panes = "A2"

    output.seek(0)
    safe_name = _safe_export_filename_part(vehicle_name, "KENDARAAN")
    safe_plate = _safe_export_filename_part(plate, "-")
    period_suffix = []
    if date_from:
        period_suffix.append(date_from.strftime("%Y-%m-%d"))
    if date_to:
        period_suffix.append(date_to.strftime("%Y-%m-%d"))
    period_text = f"_{'_to_'.join(period_suffix)}" if period_suffix else "_semua-periode"
    filename = f"Detail_Kendaraan_{safe_name.replace(' ', '_')}_{safe_plate.replace(' ', '_')}{period_text}_{date.today().strftime('%d-%m-%Y')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.get("/kendaraan/<int:vehicle_id>/detail_page")
@master_required
def vehicle_detail_page(vehicle_id:int):
    v=Vehicle.query.get_or_404(vehicle_id)
    _assert_vehicle_scope(v)

    # Di halaman detail kendaraan, masing-masing riwayat cukup tampil data terakhir.
    # Pakai id.desc() sebagai tie-breaker agar hasil edit/input terbaru langsung muncul.
    kir_rows=KirRecord.query.filter_by(vehicle_id=v.id).order_by(KirRecord.due_date.desc(), KirRecord.id.desc()).limit(1).all()
    service_rows=ServiceRecord.query.filter_by(vehicle_id=v.id).order_by(ServiceRecord.service_date.desc(), ServiceRecord.id.desc()).limit(1).all()
    loan_rows=LoanTransaction.query.filter_by(vehicle_id=v.id).order_by(LoanTransaction.created_at.desc(), LoanTransaction.id.desc()).limit(1).all()
    annual_rows = []
    five_rows = []
    try:
        annual_rows = db.session.execute(text("SELECT * FROM annual_tax_payments WHERE vehicle_id = :vid ORDER BY COALESCE(paid_date, due_date) DESC, id DESC LIMIT 1"), {'vid': v.id}).mappings().all()
        five_rows = db.session.execute(text("SELECT * FROM five_year_tax_payments WHERE vehicle_id = :vid ORDER BY COALESCE(paid_date, due_date) DESC, id DESC LIMIT 1"), {'vid': v.id}).mappings().all()
    except Exception:
        pass

    # Riwayat perubahan kendaraan = snapshot terakhir semua kolom, bukan hanya field yang terakhir diedit.
    vehicle_changes=_vehicle_detail_last_snapshot_rows(v)
    user_rows = _safe_user_history_rows(v.id, limit=1)

    # Kolom tambahan dari Data Kendaraan harus selalu ikut tampil di Detail Kendaraan.
    # Diambil fresh tiap request supaya setelah tambah/edit/hapus kolom atau update nilai,
    # halaman detail langsung mengikuti data terbaru dari vehicle_custom_*.
    custom_columns = _vehicle_custom_columns()
    custom_values = _vehicle_custom_values_map([v.id]).get(v.id, {})

    return render_template(
        "kendaraan_detail.html",
        vehicle=v,
        annual_rows=annual_rows,
        five_rows=five_rows,
        kir_rows=kir_rows,
        service_rows=service_rows,
        loan_rows=loan_rows,
        vehicle_changes=vehicle_changes,
        user_rows=user_rows,
        custom_columns=custom_columns,
        custom_values=custom_values,
        title=f"Resume Kendaraan - {vehicle_plate(v)}",
        detail_export_url=url_for("main.vehicle_detail_export_excel", vehicle_id=v.id),
        pajak_export_url=url_for("main.api_pajak_export", vehicle_id=v.id),
        kir_export_url=url_for("main.api_kir_export", vehicle_id=v.id),
        servis_export_url=url_for("main.api_servis_export_excel", vehicle_id=v.id),
        peminjaman_export_url=url_for("main.api_peminjaman_export", vehicle_id=v.id),
        back_url=_detail_back_url(),
    )